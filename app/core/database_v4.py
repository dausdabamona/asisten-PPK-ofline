"""
PPK DOCUMENT FACTORY v4.0 - Enhanced Database Manager
======================================================
Complete database schema with:
- Master Pegawai & Pejabat
- Paket Pejabat Assignment
- Survey Harga Detail
- Harga Lifecycle (HPS → Kontrak)
- Workflow Enforcement
- Audit Trail
"""

import sqlite3
import os
import json
from datetime import datetime, date
from typing import Dict, List, Optional, Any, Tuple
from contextlib import contextmanager

from .config import DATABASE_PATH, TAHUN_ANGGARAN, SATKER_DEFAULT

# ============================================================================
# ENHANCED DATABASE SCHEMA v4.0
# ============================================================================

SCHEMA_V4_SQL = """
-- ============================================================================
-- MASTER DATA TABLES
-- ============================================================================

-- Satker / Unit Kerja
CREATE TABLE IF NOT EXISTS satker (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    kode TEXT NOT NULL UNIQUE,
    nama TEXT NOT NULL,
    nama_pendek TEXT,
    alamat TEXT,
    kota TEXT,
    kode_pos TEXT,
    provinsi TEXT,
    telepon TEXT,
    fax TEXT,
    email TEXT,
    website TEXT,
    kementerian TEXT,
    eselon1 TEXT,
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

-- Pegawai (Enhanced)
CREATE TABLE IF NOT EXISTS pegawai (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nip TEXT UNIQUE,
    nama TEXT NOT NULL,
    gelar_depan TEXT,
    gelar_belakang TEXT,
    pangkat TEXT,
    golongan TEXT,
    jabatan TEXT,
    unit_kerja TEXT,
    email TEXT,
    telepon TEXT,
    npwp TEXT,
    no_rekening TEXT,
    nama_bank TEXT,
    -- Role flags
    is_ppk INTEGER DEFAULT 0,
    is_ppspm INTEGER DEFAULT 0,
    is_bendahara INTEGER DEFAULT 0,
    is_pemeriksa INTEGER DEFAULT 0,
    is_pejabat_pengadaan INTEGER DEFAULT 0,
    -- Additional
    signature_path TEXT,
    foto_path TEXT,
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_pegawai_nip ON pegawai(nip);
CREATE INDEX IF NOT EXISTS idx_pegawai_nama ON pegawai(nama);
CREATE INDEX IF NOT EXISTS idx_pegawai_active ON pegawai(is_active);

-- Penyedia / Vendor
CREATE TABLE IF NOT EXISTS penyedia (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nama TEXT NOT NULL,
    nama_direktur TEXT,
    jabatan_direktur TEXT DEFAULT 'Direktur',
    alamat TEXT,
    kota TEXT,
    npwp TEXT,
    no_rekening TEXT,
    nama_bank TEXT,
    nama_rekening TEXT,
    telepon TEXT,
    email TEXT,
    is_pkp INTEGER DEFAULT 1,
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

-- ============================================================================
-- PAKET PENGADAAN (PROCUREMENT PACKAGE) - Enhanced
-- ============================================================================

CREATE TABLE IF NOT EXISTS paket (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    kode TEXT UNIQUE,
    nama TEXT NOT NULL,
    tahun_anggaran INTEGER NOT NULL,
    jenis_pengadaan TEXT,
    metode_pengadaan TEXT,
    lokasi TEXT,
    sumber_dana TEXT,
    kode_akun TEXT,
    
    -- Nilai
    nilai_pagu REAL,
    nilai_hps REAL,
    nilai_hps_final REAL,           -- HPS setelah ditetapkan
    nilai_kontrak REAL,
    nilai_kontrak_final REAL,       -- Kontrak setelah negosiasi
    nilai_ppn REAL,
    nilai_pph REAL,
    jenis_pph TEXT,
    tarif_pph REAL,
    overhead_profit REAL DEFAULT 0.10,  -- Default 10%
    
    -- Waktu
    tanggal_mulai DATE,
    tanggal_selesai DATE,
    jangka_waktu INTEGER,
    
    -- Pihak terkait (legacy - akan diganti paket_pejabat)
    ppk_id INTEGER,
    ppspm_id INTEGER,
    bendahara_id INTEGER,
    penyedia_id INTEGER,
    
    -- Status workflow
    current_stage TEXT DEFAULT 'SPESIFIKASI',
    status TEXT DEFAULT 'draft',
    metode_hps TEXT DEFAULT 'RATA',
    
    -- Workflow flags
    spesifikasi_locked INTEGER DEFAULT 0,
    survey_locked INTEGER DEFAULT 0,
    hps_locked INTEGER DEFAULT 0,
    kak_locked INTEGER DEFAULT 0,
    kontrak_draft_locked INTEGER DEFAULT 0,
    kontrak_final_locked INTEGER DEFAULT 0,
    
    -- Metadata
    data_json TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (ppk_id) REFERENCES pegawai(id),
    FOREIGN KEY (ppspm_id) REFERENCES pegawai(id),
    FOREIGN KEY (bendahara_id) REFERENCES pegawai(id),
    FOREIGN KEY (penyedia_id) REFERENCES penyedia(id)
);

-- ============================================================================
-- PAKET PEJABAT (Role Assignment per Paket)
-- ============================================================================

CREATE TABLE IF NOT EXISTS paket_pejabat (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    pegawai_id INTEGER NOT NULL,
    peran TEXT NOT NULL,  -- PPK, PEJABAT_PENGADAAN, PPSPM, BENDAHARA, KETUA_PPHP, ANGGOTA_PPHP
    urutan INTEGER DEFAULT 1,  -- Untuk multi-select (anggota)
    tanggal_penetapan DATE,
    nomor_sk TEXT,
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    FOREIGN KEY (pegawai_id) REFERENCES pegawai(id),
    UNIQUE(paket_id, pegawai_id, peran)
);

CREATE INDEX IF NOT EXISTS idx_paket_pejabat_paket ON paket_pejabat(paket_id);
CREATE INDEX IF NOT EXISTS idx_paket_pejabat_peran ON paket_pejabat(peran);

-- ============================================================================
-- ITEM BARANG / BILL OF QUANTITY - Enhanced
-- ============================================================================

CREATE TABLE IF NOT EXISTS item_barang (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    nomor_urut INTEGER NOT NULL,
    kategori TEXT,
    kelompok TEXT,
    uraian TEXT NOT NULL,
    spesifikasi TEXT,
    satuan TEXT,
    volume REAL DEFAULT 0,
    
    -- Harga Survey
    harga_survey1 REAL,
    harga_survey2 REAL,
    harga_survey3 REAL,
    harga_rata REAL,
    
    -- Harga HPS
    harga_dasar REAL DEFAULT 0,       -- Harga satuan HPS
    overhead_profit REAL,              -- Overhead per item (jika berbeda)
    harga_hps_satuan REAL,            -- Harga HPS per satuan (setelah overhead)
    total_hps REAL DEFAULT 0,         -- Volume x Harga HPS
    
    -- Harga Kontrak Final
    harga_kontrak_satuan REAL,        -- Harga setelah negosiasi
    total_kontrak REAL,               -- Volume x Harga Kontrak
    
    -- Selisih
    selisih_harga REAL,               -- HPS - Kontrak per satuan
    selisih_total REAL,               -- Total HPS - Total Kontrak
    
    -- Legacy
    total REAL DEFAULT 0,
    keterangan TEXT,
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_item_barang_paket ON item_barang(paket_id);

-- ============================================================================
-- SURVEY HARGA DETAIL (Per Item per Sumber)
-- ============================================================================

CREATE TABLE IF NOT EXISTS survey_harga_detail (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    item_id INTEGER NOT NULL,
    sumber_ke INTEGER NOT NULL,       -- 1, 2, 3
    
    -- Informasi Sumber
    jenis_survey TEXT NOT NULL,       -- LURING, DARING, KONTRAK, BROSUR, KATALOG
    nama_sumber TEXT,                 -- Nama toko/marketplace/kontrak
    alamat TEXT,
    
    -- Untuk LURING (Toko Fisik)
    kota TEXT,
    telepon TEXT,
    tanggal_survey DATE,
    surveyor TEXT,
    nip_surveyor TEXT,
    
    -- Untuk DARING (Marketplace)
    platform TEXT,                    -- Tokopedia, Shopee, dll
    link_produk TEXT,
    tanggal_akses DATE,
    
    -- Untuk KONTRAK (Kontrak Sebelumnya)
    nomor_kontrak TEXT,
    tahun_kontrak INTEGER,
    instansi TEXT,
    
    -- Harga
    harga REAL NOT NULL,
    keterangan TEXT,
    
    -- Bukti
    bukti_path TEXT,                  -- Path to uploaded file
    
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    FOREIGN KEY (item_id) REFERENCES item_barang(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_survey_detail_paket ON survey_harga_detail(paket_id);
CREATE INDEX IF NOT EXISTS idx_survey_detail_item ON survey_harga_detail(item_id);

-- ============================================================================
-- SURVEY TOKO / SUMBER HARGA (Global per Paket)
-- ============================================================================

CREATE TABLE IF NOT EXISTS survey_toko (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    nomor_urut INTEGER NOT NULL,
    nama_toko TEXT NOT NULL,
    alamat TEXT,
    kota TEXT,
    telepon TEXT,
    email TEXT,
    website TEXT,
    jenis_survey TEXT,
    tanggal_survey DATE,
    nama_surveyor TEXT,
    nip_surveyor TEXT,
    platform_online TEXT,
    link_produk TEXT,
    nomor_kontrak_referensi TEXT,
    tahun_kontrak_referensi INTEGER,
    instansi_referensi TEXT,
    keterangan TEXT,
    is_active INTEGER DEFAULT 1,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_survey_toko_paket ON survey_toko(paket_id);

-- ============================================================================
-- HARGA LIFECYCLE (Track Price Changes)
-- ============================================================================

CREATE TABLE IF NOT EXISTS harga_lifecycle (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    item_id INTEGER,                  -- NULL for paket-level
    
    tahap TEXT NOT NULL,              -- SURVEY, HPS, NEGOSIASI, KONTRAK_FINAL
    tanggal TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    -- Values
    harga_satuan REAL,
    total REAL,
    
    -- Metadata
    keterangan TEXT,
    created_by TEXT,
    dokumen_ref TEXT,                 -- Nomor dokumen terkait
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    FOREIGN KEY (item_id) REFERENCES item_barang(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_harga_lifecycle_paket ON harga_lifecycle(paket_id);

-- ============================================================================
-- REVISI HARGA (dari Pejabat Pengadaan)
-- ============================================================================

CREATE TABLE IF NOT EXISTS revisi_harga (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    
    -- Tanggal & Nomor
    tanggal_revisi DATE,
    nomor_ba_klarifikasi TEXT,        -- Nomor BA Klarifikasi
    
    -- Summary
    total_hps_awal REAL,
    total_kontrak_hasil REAL,
    selisih REAL,
    persentase_selisih REAL,
    
    -- Status
    status TEXT DEFAULT 'draft',      -- draft, submitted, approved, rejected
    catatan TEXT,
    
    -- Approval
    approved_by INTEGER,
    approved_at TIMESTAMP,
    
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    FOREIGN KEY (approved_by) REFERENCES pegawai(id)
);

-- ============================================================================
-- BUKTI SURVEY (Attachments)
-- ============================================================================

CREATE TABLE IF NOT EXISTS bukti_survey (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    item_id INTEGER,
    survey_detail_id INTEGER,
    
    jenis TEXT,                       -- FOTO, SCREENSHOT, BROSUR, KONTRAK, NOTA
    filename TEXT NOT NULL,
    filepath TEXT NOT NULL,
    filesize INTEGER,
    mimetype TEXT,
    keterangan TEXT,
    
    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    uploaded_by TEXT,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    FOREIGN KEY (item_id) REFERENCES item_barang(id) ON DELETE CASCADE,
    FOREIGN KEY (survey_detail_id) REFERENCES survey_harga_detail(id) ON DELETE CASCADE
);

-- ============================================================================
-- DOCUMENT TIMELINE & NUMBERING
-- ============================================================================

CREATE TABLE IF NOT EXISTS dokumen_timeline (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    doc_type TEXT NOT NULL,
    nomor_dokumen TEXT,
    tanggal_dokumen DATE,
    tanggal_input TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    catatan TEXT,
    is_locked INTEGER DEFAULT 0,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    UNIQUE(paket_id, doc_type)
);

CREATE INDEX IF NOT EXISTS idx_dokumen_timeline_paket ON dokumen_timeline(paket_id);

CREATE TABLE IF NOT EXISTS nomor_counter (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun INTEGER NOT NULL,
    doc_type TEXT NOT NULL,
    prefix TEXT,
    last_number INTEGER DEFAULT 0,
    format_template TEXT,
    
    UNIQUE(tahun, doc_type)
);

-- ============================================================================
-- TIM PEMERIKSA (PPHP Team)
-- ============================================================================

CREATE TABLE IF NOT EXISTS tim_pemeriksa (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    pegawai_id INTEGER NOT NULL,
    jabatan_tim TEXT,
    urutan INTEGER,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    FOREIGN KEY (pegawai_id) REFERENCES pegawai(id),
    UNIQUE(paket_id, pegawai_id)
);

-- ============================================================================
-- WORKFLOW STATE
-- ============================================================================

CREATE TABLE IF NOT EXISTS workflow_stage (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    stage_code TEXT NOT NULL,
    stage_order INTEGER NOT NULL,
    status TEXT DEFAULT 'pending',
    started_at TIMESTAMP,
    completed_at TIMESTAMP,
    completed_by TEXT,
    notes TEXT,
    
    -- Validation results
    validation_passed INTEGER DEFAULT 0,
    validation_errors TEXT,           -- JSON array of errors
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE,
    UNIQUE(paket_id, stage_code)
);

-- ============================================================================
-- AUDIT TRAIL
-- ============================================================================

CREATE TABLE IF NOT EXISTS audit_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    user_id INTEGER,
    user_name TEXT,
    action TEXT NOT NULL,             -- CREATE, UPDATE, DELETE, LOCK, APPROVE
    table_name TEXT,
    record_id INTEGER,
    old_values TEXT,                  -- JSON
    new_values TEXT,                  -- JSON
    ip_address TEXT,
    notes TEXT
);

CREATE INDEX IF NOT EXISTS idx_audit_log_table ON audit_log(table_name, record_id);
CREATE INDEX IF NOT EXISTS idx_audit_log_timestamp ON audit_log(timestamp);

-- ============================================================================
-- DOCUMENT TRACKING
-- ============================================================================

CREATE TABLE IF NOT EXISTS dokumen (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    paket_id INTEGER NOT NULL,
    doc_type TEXT NOT NULL,
    nomor TEXT,
    tanggal DATE,
    filename TEXT,
    filepath TEXT,
    template_used TEXT,
    template_version INTEGER,
    data_json TEXT,
    status TEXT DEFAULT 'draft',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    created_by TEXT,
    
    FOREIGN KEY (paket_id) REFERENCES paket(id) ON DELETE CASCADE
);

-- ============================================================================
-- TEMPLATE MANAGEMENT
-- ============================================================================

CREATE TABLE IF NOT EXISTS template (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL,
    name TEXT NOT NULL,
    type TEXT NOT NULL,
    filename TEXT NOT NULL,
    filepath TEXT NOT NULL,
    version INTEGER DEFAULT 1,
    is_active INTEGER DEFAULT 1,
    placeholders TEXT,
    description TEXT,
    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    uploaded_by TEXT
);

CREATE TABLE IF NOT EXISTS template_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    template_id INTEGER NOT NULL,
    version INTEGER NOT NULL,
    filename TEXT NOT NULL,
    filepath TEXT NOT NULL,
    replaced_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    replaced_by TEXT,
    
    FOREIGN KEY (template_id) REFERENCES template(id)
);

CREATE TABLE IF NOT EXISTS doc_counter (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    doc_type TEXT NOT NULL,
    tahun INTEGER NOT NULL,
    counter INTEGER DEFAULT 0,
    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(doc_type, tahun)
);

-- ============================================================================
-- PERJALANAN DINAS (Official Travel)
-- ============================================================================

CREATE TABLE IF NOT EXISTS perjalanan_dinas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Info kegiatan
    nama_kegiatan TEXT NOT NULL,
    maksud_perjalanan TEXT,
    nomor_surat_tugas TEXT,
    nomor_sppd TEXT,

    -- Pelaksana
    pelaksana_nama TEXT NOT NULL,
    pelaksana_nip TEXT,
    pelaksana_pangkat TEXT,
    pelaksana_golongan TEXT,
    pelaksana_jabatan TEXT,

    -- Tujuan
    kota_asal TEXT,
    kota_tujuan TEXT,
    provinsi_tujuan TEXT,
    alamat_tujuan TEXT,

    -- Waktu
    tanggal_surat_tugas DATE,
    tanggal_berangkat DATE,
    tanggal_kembali DATE,
    lama_perjalanan INTEGER DEFAULT 1,

    -- Anggaran DIPA
    sumber_dana TEXT DEFAULT 'DIPA',
    kode_akun TEXT,

    -- Biaya
    biaya_transport REAL DEFAULT 0,
    biaya_uang_harian REAL DEFAULT 0,
    biaya_penginapan REAL DEFAULT 0,
    biaya_representasi REAL DEFAULT 0,
    biaya_lain_lain REAL DEFAULT 0,
    uang_muka REAL DEFAULT 0,

    -- Pejabat
    ppk_nama TEXT,
    ppk_nip TEXT,
    ppk_jabatan TEXT,
    bendahara_nama TEXT,
    bendahara_nip TEXT,

    -- Status
    status TEXT DEFAULT 'draft',
    doc_count INTEGER DEFAULT 0,

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_pd_tahun ON perjalanan_dinas(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_pd_pelaksana ON perjalanan_dinas(pelaksana_nama);

-- ============================================================================
-- SWAKELOLA (Self-managed Procurement)
-- ============================================================================

CREATE TABLE IF NOT EXISTS swakelola (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Info kegiatan
    nama_kegiatan TEXT NOT NULL,
    tipe_swakelola INTEGER DEFAULT 0,
    tipe_swakelola_text TEXT,
    deskripsi TEXT,
    output_kegiatan TEXT,

    -- Nomor dokumen
    nomor_kak TEXT,
    nomor_sk_tim TEXT,

    -- Waktu
    tanggal_sk_tim DATE,
    tanggal_mulai DATE,
    tanggal_selesai DATE,
    jangka_waktu INTEGER DEFAULT 30,

    -- Anggaran DIPA
    sumber_dana TEXT DEFAULT 'DIPA',
    kode_akun TEXT,
    pagu_swakelola REAL DEFAULT 0,

    -- Pemegang Uang Muka
    pum_nama TEXT,
    pum_nip TEXT,
    pum_jabatan TEXT,
    uang_muka REAL DEFAULT 0,
    tanggal_uang_muka DATE,

    -- Realisasi / Rampung
    total_realisasi REAL DEFAULT 0,
    tanggal_rampung DATE,

    -- Tim Pelaksana
    ketua_nama TEXT,
    ketua_nip TEXT,
    ketua_jabatan TEXT,
    sekretaris_nama TEXT,
    sekretaris_nip TEXT,
    anggota_tim TEXT,

    -- Pejabat
    ppk_nama TEXT,
    ppk_nip TEXT,
    ppk_jabatan TEXT,
    bendahara_nama TEXT,
    bendahara_nip TEXT,

    -- Status
    status TEXT DEFAULT 'draft',

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_sw_tahun ON swakelola(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_sw_kegiatan ON swakelola(nama_kegiatan);

-- ============================================================================
-- PJLP (Penyedia Jasa Lainnya Perorangan)
-- ============================================================================

CREATE TABLE IF NOT EXISTS pjlp (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Info Kontrak
    nomor_kontrak TEXT,
    tanggal_kontrak DATE,
    nama_pekerjaan TEXT NOT NULL,
    lingkup_pekerjaan TEXT,
    lokasi_pekerjaan TEXT,

    -- Data Penyedia/Tenaga PJLP
    nama_pjlp TEXT NOT NULL,
    nik TEXT,
    npwp TEXT,
    alamat TEXT,
    telepon TEXT,
    email TEXT,
    no_rekening TEXT,
    nama_bank TEXT,

    -- Periode Kontrak
    tanggal_mulai DATE,
    tanggal_selesai DATE,
    jangka_waktu INTEGER DEFAULT 12,

    -- Nilai Kontrak
    honor_bulanan REAL DEFAULT 0,
    total_nilai_kontrak REAL DEFAULT 0,

    -- Anggaran DIPA
    sumber_dana TEXT DEFAULT 'DIPA',
    kode_akun TEXT,

    -- Pejabat
    ppk_nama TEXT,
    ppk_nip TEXT,
    ppk_jabatan TEXT,
    bendahara_nama TEXT,
    bendahara_nip TEXT,

    -- Status
    status TEXT DEFAULT 'aktif',

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_pjlp_tahun ON pjlp(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_pjlp_nama ON pjlp(nama_pjlp);
CREATE INDEX IF NOT EXISTS idx_pjlp_status ON pjlp(status);

-- ============================================================================
-- PEMBAYARAN PJLP (Monthly Payment Tracking)
-- ============================================================================

CREATE TABLE IF NOT EXISTS pembayaran_pjlp (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    pjlp_id INTEGER NOT NULL,

    -- Periode
    bulan INTEGER NOT NULL,
    tahun INTEGER NOT NULL,

    -- Dokumen
    nomor_kuitansi TEXT,
    tanggal_kuitansi DATE,
    nomor_spp TEXT,
    tanggal_spp DATE,

    -- Nilai Pembayaran
    nilai_bruto REAL DEFAULT 0,
    potongan_pajak REAL DEFAULT 0,
    potongan_lain REAL DEFAULT 0,
    nilai_netto REAL DEFAULT 0,

    -- Monitoring
    kehadiran_hari INTEGER DEFAULT 0,
    total_hari_kerja INTEGER DEFAULT 22,
    catatan_kinerja TEXT,

    -- Status
    status TEXT DEFAULT 'draft',
    tanggal_bayar DATE,

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

    FOREIGN KEY (pjlp_id) REFERENCES pjlp(id) ON DELETE CASCADE,
    UNIQUE(pjlp_id, bulan, tahun)
);

CREATE INDEX IF NOT EXISTS idx_pembayaran_pjlp_periode ON pembayaran_pjlp(tahun, bulan);
CREATE INDEX IF NOT EXISTS idx_pembayaran_pjlp_status ON pembayaran_pjlp(status);

-- ============================================================================
-- SK KPA (Surat Keputusan Kuasa Pengguna Anggaran)
-- ============================================================================

CREATE TABLE IF NOT EXISTS sk_kpa (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,
    nomor_sk TEXT NOT NULL,
    tanggal_sk DATE NOT NULL,
    perihal TEXT NOT NULL,
    jenis_pembayaran TEXT NOT NULL,  -- 'swakelola', 'honorarium', 'jamuan_tamu', 'pjlp'
    referensi_id INTEGER,  -- ID dari tabel terkait

    -- Data KPA
    kpa_nama TEXT,
    kpa_nip TEXT,
    kpa_jabatan TEXT,

    -- Data PPK
    ppk_nama TEXT,
    ppk_nip TEXT,
    ppk_jabatan TEXT,

    -- Data Bendahara
    bendahara_nama TEXT,
    bendahara_nip TEXT,

    -- Nilai
    nilai_pembayaran REAL DEFAULT 0,

    -- Keterangan
    dasar_pembayaran TEXT,
    keterangan TEXT,

    status TEXT DEFAULT 'draft',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_sk_kpa_tahun ON sk_kpa(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_sk_kpa_jenis ON sk_kpa(jenis_pembayaran);

-- ============================================================================
-- HONORARIUM
-- ============================================================================

CREATE TABLE IF NOT EXISTS honorarium (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Info Kegiatan
    nama_kegiatan TEXT NOT NULL,
    jenis_honorarium TEXT NOT NULL,  -- 'narasumber', 'moderator', 'panitia', 'tim_kerja', 'lainnya'
    kategori TEXT DEFAULT 'reguler',  -- 'reguler', 'insidentil'

    -- Dokumen
    nomor_sk_kpa TEXT,
    tanggal_sk_kpa DATE,
    nomor_spt TEXT,
    tanggal_spt DATE,
    nomor_kuitansi TEXT,
    tanggal_kuitansi DATE,

    -- Periode (untuk reguler)
    bulan INTEGER,
    periode_mulai DATE,
    periode_selesai DATE,

    -- Sumber Dana
    sumber_dana TEXT DEFAULT 'DIPA',
    kode_akun TEXT,
    mak TEXT,

    -- Pejabat
    kpa_nama TEXT,
    kpa_nip TEXT,
    ppk_nama TEXT,
    ppk_nip TEXT,
    bendahara_nama TEXT,
    bendahara_nip TEXT,

    -- Total
    total_bruto REAL DEFAULT 0,
    total_pajak REAL DEFAULT 0,
    total_netto REAL DEFAULT 0,

    status TEXT DEFAULT 'draft',
    keterangan TEXT,

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

-- Detail Penerima Honorarium
CREATE TABLE IF NOT EXISTS honorarium_detail (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    honorarium_id INTEGER NOT NULL,

    -- Data Penerima
    nama TEXT NOT NULL,
    nip TEXT,
    jabatan TEXT,
    pangkat_golongan TEXT,
    npwp TEXT,
    no_rekening TEXT,
    nama_bank TEXT,

    -- Peran dalam kegiatan
    peran TEXT,  -- 'narasumber', 'moderator', 'panitia', dll

    -- Nilai
    jumlah_jp INTEGER DEFAULT 1,  -- Jam Pelajaran / Kegiatan
    tarif_per_jp REAL DEFAULT 0,
    nilai_bruto REAL DEFAULT 0,
    pph21 REAL DEFAULT 0,
    potongan_lain REAL DEFAULT 0,
    nilai_netto REAL DEFAULT 0,

    keterangan TEXT,

    FOREIGN KEY (honorarium_id) REFERENCES honorarium(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_honorarium_tahun ON honorarium(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_honorarium_kategori ON honorarium(kategori);

-- ============================================================================
-- JAMUAN TAMU
-- ============================================================================

CREATE TABLE IF NOT EXISTS jamuan_tamu (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Info Kegiatan
    nama_kegiatan TEXT NOT NULL,
    tanggal_kegiatan DATE NOT NULL,
    tempat TEXT,
    kategori TEXT DEFAULT 'reguler',  -- 'reguler', 'insidentil'

    -- Tamu
    nama_tamu TEXT,
    instansi_tamu TEXT,
    jabatan_tamu TEXT,
    jumlah_tamu INTEGER DEFAULT 1,

    -- Dokumen
    nomor_sk_kpa TEXT,
    tanggal_sk_kpa DATE,
    nomor_nota_dinas TEXT,
    tanggal_nota_dinas DATE,
    nomor_kuitansi TEXT,
    tanggal_kuitansi DATE,

    -- Sumber Dana
    sumber_dana TEXT DEFAULT 'DIPA',
    kode_akun TEXT,
    mak TEXT,

    -- Pejabat
    kpa_nama TEXT,
    kpa_nip TEXT,
    ppk_nama TEXT,
    ppk_nip TEXT,
    bendahara_nama TEXT,
    bendahara_nip TEXT,

    -- Biaya
    biaya_konsumsi REAL DEFAULT 0,
    biaya_akomodasi REAL DEFAULT 0,
    biaya_transportasi REAL DEFAULT 0,
    biaya_lainnya REAL DEFAULT 0,
    total_biaya REAL DEFAULT 0,

    status TEXT DEFAULT 'draft',
    keterangan TEXT,

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_jamuan_tamu_tahun ON jamuan_tamu(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_jamuan_tamu_tanggal ON jamuan_tamu(tanggal_kegiatan);

-- ============================================================================
-- PAGU ANGGARAN / FA DETAIL (POK/DIPA)
-- ============================================================================

CREATE TABLE IF NOT EXISTS pagu_anggaran (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Kode Hierarki
    kode_program TEXT,
    kode_kegiatan TEXT,
    kode_kro TEXT,
    kode_ro TEXT,
    kode_komponen TEXT,
    kode_sub_komponen TEXT,
    kode_akun TEXT,
    kode_detail TEXT,
    kode_full TEXT,  -- Kode lengkap gabungan

    -- Level hierarki: 1=Program, 2=Kegiatan, 3=KRO, 4=RO, 5=Komponen, 6=SubKomponen, 7=Akun, 8=Detail
    level_kode INTEGER DEFAULT 8,
    parent_id INTEGER,

    -- Uraian
    uraian TEXT NOT NULL,

    -- Volume dan Satuan
    volume REAL DEFAULT 0,
    satuan TEXT,

    -- Harga
    harga_satuan REAL DEFAULT 0,
    jumlah REAL DEFAULT 0,

    -- Realisasi
    realisasi REAL DEFAULT 0,
    sisa REAL DEFAULT 0,
    persen_realisasi REAL DEFAULT 0,

    -- Sumber Dana
    sumber_dana TEXT DEFAULT 'RM',  -- RM, PNBP, BLU, PLN, HLN

    -- Flag
    is_locked INTEGER DEFAULT 0,
    is_blokir INTEGER DEFAULT 0,

    -- Nomor MAK (Mata Anggaran Kegiatan) = kode_akun.kode_detail
    nomor_mak TEXT,

    -- Keterangan
    keterangan TEXT,

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

    -- Unique constraint untuk update/replace
    UNIQUE(tahun_anggaran, kode_full)
);

CREATE INDEX IF NOT EXISTS idx_pagu_tahun ON pagu_anggaran(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_pagu_kode_full ON pagu_anggaran(kode_full);
CREATE INDEX IF NOT EXISTS idx_pagu_nomor_mak ON pagu_anggaran(nomor_mak);
CREATE INDEX IF NOT EXISTS idx_pagu_kode_akun ON pagu_anggaran(kode_akun);
CREATE INDEX IF NOT EXISTS idx_pagu_level ON pagu_anggaran(level_kode);
CREATE INDEX IF NOT EXISTS idx_pagu_parent ON pagu_anggaran(parent_id);

-- Tabel untuk tracking realisasi per bulan
CREATE TABLE IF NOT EXISTS realisasi_anggaran (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    pagu_id INTEGER NOT NULL,
    bulan INTEGER NOT NULL,
    tahun INTEGER NOT NULL,

    nilai_realisasi REAL DEFAULT 0,
    nomor_sp2d TEXT,
    tanggal_sp2d DATE,
    keterangan TEXT,

    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

    FOREIGN KEY (pagu_id) REFERENCES pagu_anggaran(id) ON DELETE CASCADE,
    UNIQUE(pagu_id, bulan, tahun)
);

CREATE INDEX IF NOT EXISTS idx_realisasi_periode ON realisasi_anggaran(tahun, bulan);

-- ============================================================================
-- HONORARIUM PENGELOLA KEUANGAN
-- ============================================================================

CREATE TABLE IF NOT EXISTS honorarium_pengelola (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun INTEGER NOT NULL,
    bulan INTEGER NOT NULL,
    jabatan TEXT NOT NULL,  -- KPA, PPK, PPSPM, Bendahara, Operator, Staf
    pegawai_id INTEGER NOT NULL,
    jumlah REAL DEFAULT 0,
    pajak REAL DEFAULT 0,
    netto REAL DEFAULT 0,
    pagu_id INTEGER,  -- Reference to pagu_anggaran
    keterangan TEXT,
    status TEXT DEFAULT 'draft',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

    FOREIGN KEY (pegawai_id) REFERENCES pegawai(id),
    FOREIGN KEY (pagu_id) REFERENCES pagu_anggaran(id)
);

CREATE INDEX IF NOT EXISTS idx_hon_pengelola_tahun_bulan ON honorarium_pengelola(tahun, bulan);
CREATE INDEX IF NOT EXISTS idx_hon_pengelola_jabatan ON honorarium_pengelola(jabatan);

-- ============================================================================
-- DIPA (Daftar Isian Pelaksanaan Anggaran) - Import from CSV
-- ============================================================================

CREATE TABLE IF NOT EXISTS dipa (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tahun_anggaran INTEGER NOT NULL,

    -- Kode Satker & Hierarki
    kode_satker TEXT,
    kode_program TEXT,
    kode_kegiatan TEXT,
    kode_output TEXT,
    kode_akun TEXT,

    -- Uraian
    uraian_item TEXT NOT NULL,

    -- Volume & Satuan
    volume REAL DEFAULT 0,
    satuan TEXT,
    harga_satuan REAL DEFAULT 0,
    total REAL DEFAULT 0,

    -- POK (Pagu Operasional Kegiatan) per bulan
    pok_nilai_1 REAL DEFAULT 0,   -- Januari
    pok_nilai_2 REAL DEFAULT 0,   -- Februari
    pok_nilai_3 REAL DEFAULT 0,   -- Maret
    pok_nilai_4 REAL DEFAULT 0,   -- April
    pok_nilai_5 REAL DEFAULT 0,   -- Mei
    pok_nilai_6 REAL DEFAULT 0,   -- Juni
    pok_nilai_7 REAL DEFAULT 0,   -- Juli
    pok_nilai_8 REAL DEFAULT 0,   -- Agustus
    pok_nilai_9 REAL DEFAULT 0,   -- September
    pok_nilai_10 REAL DEFAULT 0,  -- Oktober
    pok_nilai_11 REAL DEFAULT 0,  -- November
    pok_nilai_12 REAL DEFAULT 0,  -- Desember

    -- Realisasi
    realisasi REAL DEFAULT 0,
    sisa REAL DEFAULT 0,

    -- Kode lengkap gabungan untuk lookup
    kode_full TEXT,

    -- Metadata
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

    -- Unique constraint untuk update/replace
    UNIQUE(tahun_anggaran, kode_satker, kode_akun, uraian_item)
);

CREATE INDEX IF NOT EXISTS idx_dipa_tahun ON dipa(tahun_anggaran);
CREATE INDEX IF NOT EXISTS idx_dipa_kode_akun ON dipa(kode_akun);
CREATE INDEX IF NOT EXISTS idx_dipa_kode_satker ON dipa(kode_satker);
CREATE INDEX IF NOT EXISTS idx_dipa_kode_full ON dipa(kode_full);
"""

# ============================================================================
# PERAN PEJABAT
# ============================================================================

PERAN_PEJABAT = {
    'PPK': 'Pejabat Pembuat Komitmen',
    'PEJABAT_PENGADAAN': 'Pejabat Pengadaan',
    'PPSPM': 'Pejabat Penandatangan SPM',
    'BENDAHARA': 'Bendahara Pengeluaran',
    'KETUA_PPHP': 'Ketua Pejabat Pemeriksa Hasil Pekerjaan',
    'ANGGOTA_PPHP': 'Anggota Pejabat Pemeriksa Hasil Pekerjaan',
    'STAF_PPK': 'Staf PPK',
}

# ============================================================================
# WORKFLOW STAGES - Enhanced
# ============================================================================

WORKFLOW_STAGES_V4 = [
    {'code': 'SPESIFIKASI', 'name': 'Spesifikasi & BOQ', 'order': 1, 
     'required_before': [], 'can_generate': ['SPESIFIKASI_TEKNIS', 'BOQ']},
    {'code': 'SURVEY', 'name': 'Survey Harga', 'order': 2, 
     'required_before': ['SPESIFIKASI'], 'can_generate': ['BA_SURVEY', 'LAMPIRAN_SURVEY']},
    {'code': 'HPS', 'name': 'Penetapan HPS', 'order': 3, 
     'required_before': ['SURVEY'], 'can_generate': ['REKAP_HPS', 'PENETAPAN_HPS']},
    {'code': 'KAK', 'name': 'Kerangka Acuan Kerja', 'order': 4, 
     'required_before': ['HPS'], 'can_generate': ['KAK']},
    {'code': 'KONTRAK_DRAFT', 'name': 'Rancangan Kontrak', 'order': 5, 
     'required_before': ['KAK'], 'can_generate': ['SPK_DRAFT', 'SSUK', 'SSKK']},
    {'code': 'NOTA_DINAS', 'name': 'Nota Dinas ke PP', 'order': 6, 
     'required_before': ['KONTRAK_DRAFT'], 'can_generate': ['NOTA_DINAS_PP']},
    {'code': 'REVISI_HARGA', 'name': 'Revisi/Klarifikasi Harga', 'order': 7, 
     'required_before': ['NOTA_DINAS'], 'can_generate': ['BA_KLARIFIKASI', 'HPS_REVISI']},
    {'code': 'KONTRAK_FINAL', 'name': 'Kontrak Final', 'order': 8, 
     'required_before': ['REVISI_HARGA'], 'can_generate': ['SPK_FINAL']},
    {'code': 'SPMK', 'name': 'Surat Perintah Mulai Kerja', 'order': 9, 
     'required_before': ['KONTRAK_FINAL'], 'can_generate': ['SPMK']},
    {'code': 'BAHP', 'name': 'Berita Acara Pemeriksaan', 'order': 10, 
     'required_before': ['SPMK'], 'can_generate': ['BAHP']},
    {'code': 'BAST', 'name': 'Berita Acara Serah Terima', 'order': 11, 
     'required_before': ['BAHP'], 'can_generate': ['BAST']},
    {'code': 'SPP', 'name': 'Surat Permintaan Pembayaran', 'order': 12, 
     'required_before': ['BAST'], 'can_generate': ['SPP', 'SSP', 'KUITANSI']},
]

# ============================================================================
# DATABASE MANAGER CLASS
# ============================================================================

class DatabaseManagerV4:
    """Enhanced Database Manager for PPK Document Factory v4.0"""
    
    def __init__(self, db_path: str = None):
        self.db_path = db_path or DATABASE_PATH
        self._init_db()
    
    def _init_db(self):
        """Initialize database and create/update tables"""
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        
        with self.get_connection() as conn:
            conn.executescript(SCHEMA_V4_SQL)
            
            # Migrations
            self._run_migrations(conn)
            
            # Insert default satker if not exists
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM satker")
            if cursor.fetchone()[0] == 0:
                self._insert_default_satker(cursor)
            
            conn.commit()
    
    def _run_migrations(self, conn):
        """Run database migrations"""
        cursor = conn.cursor()
        
        # Migration: Add new columns to pegawai if not exist
        cursor.execute("PRAGMA table_info(pegawai)")
        columns = [col[1] for col in cursor.fetchall()]
        
        migrations = [
            ('email', "ALTER TABLE pegawai ADD COLUMN email TEXT"),
            ('telepon', "ALTER TABLE pegawai ADD COLUMN telepon TEXT"),
            ('is_pejabat_pengadaan', "ALTER TABLE pegawai ADD COLUMN is_pejabat_pengadaan INTEGER DEFAULT 0"),
            ('foto_path', "ALTER TABLE pegawai ADD COLUMN foto_path TEXT"),
        ]
        
        for col, sql in migrations:
            if col not in columns:
                try:
                    cursor.execute(sql)
                except:
                    pass
        
        # Migration: Add new columns to paket if not exist
        cursor.execute("PRAGMA table_info(paket)")
        paket_columns = [col[1] for col in cursor.fetchall()]
        
        paket_migrations = [
            ('nilai_hps_final', "ALTER TABLE paket ADD COLUMN nilai_hps_final REAL"),
            ('nilai_kontrak_final', "ALTER TABLE paket ADD COLUMN nilai_kontrak_final REAL"),
            ('overhead_profit', "ALTER TABLE paket ADD COLUMN overhead_profit REAL DEFAULT 0.10"),
            ('spesifikasi_locked', "ALTER TABLE paket ADD COLUMN spesifikasi_locked INTEGER DEFAULT 0"),
            ('survey_locked', "ALTER TABLE paket ADD COLUMN survey_locked INTEGER DEFAULT 0"),
            ('hps_locked', "ALTER TABLE paket ADD COLUMN hps_locked INTEGER DEFAULT 0"),
            ('kak_locked', "ALTER TABLE paket ADD COLUMN kak_locked INTEGER DEFAULT 0"),
            ('kontrak_draft_locked', "ALTER TABLE paket ADD COLUMN kontrak_draft_locked INTEGER DEFAULT 0"),
            ('kontrak_final_locked', "ALTER TABLE paket ADD COLUMN kontrak_final_locked INTEGER DEFAULT 0"),
            ('metode_hps', "ALTER TABLE paket ADD COLUMN metode_hps TEXT DEFAULT 'RATA'"),
        ]
        
        for col, sql in paket_migrations:
            if col not in paket_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass
        
        # Migration: Add new columns to item_barang
        cursor.execute("PRAGMA table_info(item_barang)")
        item_columns = [col[1] for col in cursor.fetchall()]
        
        item_migrations = [
            ('harga_hps_satuan', "ALTER TABLE item_barang ADD COLUMN harga_hps_satuan REAL"),
            ('total_hps', "ALTER TABLE item_barang ADD COLUMN total_hps REAL"),
            ('harga_kontrak_satuan', "ALTER TABLE item_barang ADD COLUMN harga_kontrak_satuan REAL"),
            ('total_kontrak', "ALTER TABLE item_barang ADD COLUMN total_kontrak REAL"),
            ('selisih_harga', "ALTER TABLE item_barang ADD COLUMN selisih_harga REAL"),
            ('selisih_total', "ALTER TABLE item_barang ADD COLUMN selisih_total REAL"),
            ('overhead_profit', "ALTER TABLE item_barang ADD COLUMN overhead_profit REAL"),
        ]
        
        for col, sql in item_migrations:
            if col not in item_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add pejabat keuangan columns to satker
        cursor.execute("PRAGMA table_info(satker)")
        satker_columns = [col[1] for col in cursor.fetchall()]

        satker_migrations = [
            ('kpa_id', "ALTER TABLE satker ADD COLUMN kpa_id INTEGER REFERENCES pegawai(id)"),
            ('ppk_id', "ALTER TABLE satker ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('ppspm_id', "ALTER TABLE satker ADD COLUMN ppspm_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE satker ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in satker_migrations:
            if col not in satker_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add SK KPA columns to swakelola
        cursor.execute("PRAGMA table_info(swakelola)")
        swakelola_columns = [col[1] for col in cursor.fetchall()]

        swakelola_migrations = [
            ('nomor_sk_kpa', "ALTER TABLE swakelola ADD COLUMN nomor_sk_kpa TEXT"),
            ('tanggal_sk_kpa', "ALTER TABLE swakelola ADD COLUMN tanggal_sk_kpa DATE"),
            ('perihal_sk_kpa', "ALTER TABLE swakelola ADD COLUMN perihal_sk_kpa TEXT"),
        ]

        for col, sql in swakelola_migrations:
            if col not in swakelola_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to perjalanan_dinas for normalization
        cursor.execute("PRAGMA table_info(perjalanan_dinas)")
        pd_columns = [col[1] for col in cursor.fetchall()]

        pd_migrations = [
            ('pelaksana_id', "ALTER TABLE perjalanan_dinas ADD COLUMN pelaksana_id INTEGER REFERENCES pegawai(id)"),
            ('ppk_id', "ALTER TABLE perjalanan_dinas ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE perjalanan_dinas ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in pd_migrations:
            if col not in pd_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to swakelola for normalization
        swakelola_norm_migrations = [
            ('ketua_id', "ALTER TABLE swakelola ADD COLUMN ketua_id INTEGER REFERENCES pegawai(id)"),
            ('sekretaris_id', "ALTER TABLE swakelola ADD COLUMN sekretaris_id INTEGER REFERENCES pegawai(id)"),
            ('pum_id', "ALTER TABLE swakelola ADD COLUMN pum_id INTEGER REFERENCES pegawai(id)"),
            ('ppk_id', "ALTER TABLE swakelola ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE swakelola ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in swakelola_norm_migrations:
            if col not in swakelola_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to jamuan_tamu for normalization
        cursor.execute("PRAGMA table_info(jamuan_tamu)")
        jt_columns = [col[1] for col in cursor.fetchall()]

        jt_migrations = [
            ('kpa_id', "ALTER TABLE jamuan_tamu ADD COLUMN kpa_id INTEGER REFERENCES pegawai(id)"),
            ('ppk_id', "ALTER TABLE jamuan_tamu ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE jamuan_tamu ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in jt_migrations:
            if col not in jt_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to honorarium for normalization
        cursor.execute("PRAGMA table_info(honorarium)")
        hon_columns = [col[1] for col in cursor.fetchall()]

        hon_migrations = [
            ('kpa_id', "ALTER TABLE honorarium ADD COLUMN kpa_id INTEGER REFERENCES pegawai(id)"),
            ('ppk_id', "ALTER TABLE honorarium ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE honorarium ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in hon_migrations:
            if col not in hon_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to honorarium_detail for normalization
        cursor.execute("PRAGMA table_info(honorarium_detail)")
        hond_columns = [col[1] for col in cursor.fetchall()]

        hond_migrations = [
            ('pegawai_id', "ALTER TABLE honorarium_detail ADD COLUMN pegawai_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in hond_migrations:
            if col not in hond_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to sk_kpa for normalization
        cursor.execute("PRAGMA table_info(sk_kpa)")
        sk_columns = [col[1] for col in cursor.fetchall()]

        sk_migrations = [
            ('kpa_id', "ALTER TABLE sk_kpa ADD COLUMN kpa_id INTEGER REFERENCES pegawai(id)"),
            ('ppk_id', "ALTER TABLE sk_kpa ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE sk_kpa ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in sk_migrations:
            if col not in sk_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key to honorarium_pengelola for normalization
        cursor.execute("PRAGMA table_info(honorarium_pengelola)")
        hp_columns = [col[1] for col in cursor.fetchall()]

        hp_migrations = [
            ('pegawai_id', "ALTER TABLE honorarium_pengelola ADD COLUMN pegawai_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in hp_migrations:
            if col not in hp_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add foreign key columns to pjlp for normalization
        cursor.execute("PRAGMA table_info(pjlp)")
        pjlp_columns = [col[1] for col in cursor.fetchall()]

        pjlp_migrations = [
            ('ppk_id', "ALTER TABLE pjlp ADD COLUMN ppk_id INTEGER REFERENCES pegawai(id)"),
            ('bendahara_id', "ALTER TABLE pjlp ADD COLUMN bendahara_id INTEGER REFERENCES pegawai(id)"),
        ]

        for col, sql in pjlp_migrations:
            if col not in pjlp_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

        # Migration: Add nomor_mak column to pagu_anggaran
        cursor.execute("PRAGMA table_info(pagu_anggaran)")
        pagu_columns = [col[1] for col in cursor.fetchall()]

        pagu_migrations = [
            ('nomor_mak', "ALTER TABLE pagu_anggaran ADD COLUMN nomor_mak TEXT"),
        ]

        for col, sql in pagu_migrations:
            if col not in pagu_columns:
                try:
                    cursor.execute(sql)
                except:
                    pass

    def _insert_default_satker(self, cursor):
        """Insert default satker data"""
        cursor.execute("""
            INSERT INTO satker (kode, nama, nama_pendek, alamat, kota, kode_pos,
                provinsi, telepon, fax, email, website, kementerian, eselon1)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            SATKER_DEFAULT['kode'],
            SATKER_DEFAULT['nama'],
            SATKER_DEFAULT['nama_pendek'],
            SATKER_DEFAULT['alamat'],
            SATKER_DEFAULT['kota'],
            SATKER_DEFAULT['kode_pos'],
            SATKER_DEFAULT['provinsi'],
            SATKER_DEFAULT['telepon'],
            SATKER_DEFAULT['fax'],
            SATKER_DEFAULT['email'],
            SATKER_DEFAULT['website'],
            SATKER_DEFAULT['kementerian'],
            SATKER_DEFAULT['eselon1'],
        ))

    # =========================================================================
    # SATKER OPERATIONS
    # =========================================================================

    def get_satker(self) -> Dict:
        """Get default satker data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM satker LIMIT 1")
            row = cursor.fetchone()
            return dict(row) if row else {}

    def update_satker(self, data: Dict) -> bool:
        """Update satker data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE satker SET
                    nama = ?, nama_pendek = ?, alamat = ?, kota = ?, kode_pos = ?,
                    provinsi = ?, telepon = ?, fax = ?, email = ?, website = ?,
                    kementerian = ?, eselon1 = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = 1
            """, (
                data.get('nama'), data.get('nama_pendek'), data.get('alamat'),
                data.get('kota'), data.get('kode_pos'), data.get('provinsi'),
                data.get('telepon'), data.get('fax'), data.get('email'),
                data.get('website'), data.get('kementerian'), data.get('eselon1'),
            ))
            conn.commit()
            return cursor.rowcount > 0

    def update_satker_pejabat(self, kpa_id: int = None, ppk_id: int = None,
                              ppspm_id: int = None, bendahara_id: int = None) -> bool:
        """Update pejabat keuangan for satker"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE satker SET
                    kpa_id = ?, ppk_id = ?, ppspm_id = ?, bendahara_id = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = 1
            """, (kpa_id, ppk_id, ppspm_id, bendahara_id))
            conn.commit()
            return cursor.rowcount > 0

    def get_satker_pejabat(self) -> Dict:
        """Get satker pejabat with pegawai details"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT s.id, s.kode, s.nama, s.nama_pendek,
                    s.kpa_id, s.ppk_id, s.ppspm_id, s.bendahara_id,
                    kpa.nama as kpa_nama, kpa.nip as kpa_nip, kpa.jabatan as kpa_jabatan,
                    kpa.pangkat as kpa_pangkat, kpa.golongan as kpa_golongan,
                    ppk.nama as ppk_nama, ppk.nip as ppk_nip, ppk.jabatan as ppk_jabatan,
                    ppk.pangkat as ppk_pangkat, ppk.golongan as ppk_golongan,
                    ppspm.nama as ppspm_nama, ppspm.nip as ppspm_nip, ppspm.jabatan as ppspm_jabatan,
                    ppspm.pangkat as ppspm_pangkat, ppspm.golongan as ppspm_golongan,
                    bendahara.nama as bendahara_nama, bendahara.nip as bendahara_nip,
                    bendahara.jabatan as bendahara_jabatan, bendahara.pangkat as bendahara_pangkat,
                    bendahara.golongan as bendahara_golongan
                FROM satker s
                LEFT JOIN pegawai kpa ON s.kpa_id = kpa.id
                LEFT JOIN pegawai ppk ON s.ppk_id = ppk.id
                LEFT JOIN pegawai ppspm ON s.ppspm_id = ppspm.id
                LEFT JOIN pegawai bendahara ON s.bendahara_id = bendahara.id
                LIMIT 1
            """)
            row = cursor.fetchone()
            return dict(row) if row else {}

    @contextmanager
    def get_connection(self):
        """Get database connection with context manager"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()
    
    # =========================================================================
    # PEGAWAI OPERATIONS
    # =========================================================================
    
    def get_all_pegawai(self, active_only: bool = True, search: str = None) -> List[Dict]:
        """Get all pegawai with optional search"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            sql = "SELECT * FROM pegawai"
            params = []
            conditions = []
            
            if active_only:
                conditions.append("is_active = 1")
            
            if search:
                conditions.append("(nama LIKE ? OR nip LIKE ? OR jabatan LIKE ?)")
                search_pattern = f"%{search}%"
                params.extend([search_pattern, search_pattern, search_pattern])
            
            if conditions:
                sql += " WHERE " + " AND ".join(conditions)
            
            sql += " ORDER BY nama"
            
            cursor.execute(sql, params)
            return [dict(row) for row in cursor.fetchall()]
    
    def get_pegawai(self, pegawai_id: int) -> Optional[Dict]:
        """Get single pegawai by ID"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM pegawai WHERE id = ?", (pegawai_id,))
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def get_pegawai_by_nip(self, nip: str) -> Optional[Dict]:
        """Get pegawai by NIP"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM pegawai WHERE nip = ?", (nip,))
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def create_pegawai(self, data: Dict) -> int:
        """Create new pegawai"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO pegawai (
                    nip, nama, gelar_depan, gelar_belakang, pangkat, golongan,
                    jabatan, unit_kerja, email, telepon, npwp, no_rekening, nama_bank,
                    is_ppk, is_ppspm, is_bendahara, is_pemeriksa, is_pejabat_pengadaan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('nip'), data.get('nama'), data.get('gelar_depan'),
                data.get('gelar_belakang'), data.get('pangkat'), data.get('golongan'),
                data.get('jabatan'), data.get('unit_kerja'), data.get('email'),
                data.get('telepon'), data.get('npwp'), data.get('no_rekening'),
                data.get('nama_bank'), data.get('is_ppk', 0), data.get('is_ppspm', 0),
                data.get('is_bendahara', 0), data.get('is_pemeriksa', 0),
                data.get('is_pejabat_pengadaan', 0)
            ))
            
            conn.commit()
            return cursor.lastrowid
    
    def update_pegawai(self, pegawai_id: int, data: Dict) -> bool:
        """Update pegawai"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE pegawai SET
                    nip = ?, nama = ?, gelar_depan = ?, gelar_belakang = ?,
                    pangkat = ?, golongan = ?, jabatan = ?, unit_kerja = ?,
                    email = ?, telepon = ?, npwp = ?, no_rekening = ?, nama_bank = ?,
                    is_ppk = ?, is_ppspm = ?, is_bendahara = ?, is_pemeriksa = ?,
                    is_pejabat_pengadaan = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nip'), data.get('nama'), data.get('gelar_depan'),
                data.get('gelar_belakang'), data.get('pangkat'), data.get('golongan'),
                data.get('jabatan'), data.get('unit_kerja'), data.get('email'),
                data.get('telepon'), data.get('npwp'), data.get('no_rekening'),
                data.get('nama_bank'), data.get('is_ppk', 0), data.get('is_ppspm', 0),
                data.get('is_bendahara', 0), data.get('is_pemeriksa', 0),
                data.get('is_pejabat_pengadaan', 0), pegawai_id
            ))
            
            conn.commit()
            return cursor.rowcount > 0
    
    def deactivate_pegawai(self, pegawai_id: int) -> bool:
        """Soft delete pegawai"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE pegawai SET is_active = 0, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (pegawai_id,))
            conn.commit()
            return cursor.rowcount > 0
    
    def get_pegawai_by_role(self, role: str) -> List[Dict]:
        """Get pegawai by role flag"""
        role_column = {
            'ppk': 'is_ppk',
            'ppspm': 'is_ppspm',
            'bendahara': 'is_bendahara',
            'pemeriksa': 'is_pemeriksa',
            'pejabat_pengadaan': 'is_pejabat_pengadaan',
        }.get(role.lower())
        
        if not role_column:
            return []
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f"""
                SELECT * FROM pegawai 
                WHERE {role_column} = 1 AND is_active = 1
                ORDER BY nama
            """)
            return [dict(row) for row in cursor.fetchall()]
    
    def bulk_import_pegawai(self, pegawai_list: List[Dict]) -> Tuple[int, int, List[str]]:
        """
        Bulk import pegawai from list
        Returns: (imported, updated, errors)
        """
        imported = 0
        updated = 0
        errors = []
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            for i, data in enumerate(pegawai_list):
                try:
                    nip = data.get('nip', '').strip()
                    nama = data.get('nama', '').strip()
                    
                    if not nama:
                        errors.append(f"Baris {i+1}: Nama wajib diisi")
                        continue
                    
                    # Check if exists
                    if nip:
                        cursor.execute("SELECT id FROM pegawai WHERE nip = ?", (nip,))
                        existing = cursor.fetchone()
                        
                        if existing:
                            # Update
                            cursor.execute("""
                                UPDATE pegawai SET
                                    nama = ?, pangkat = ?, golongan = ?, jabatan = ?,
                                    unit_kerja = ?, email = ?, telepon = ?,
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE nip = ?
                            """, (
                                nama, data.get('pangkat'), data.get('golongan'),
                                data.get('jabatan'), data.get('unit_kerja'),
                                data.get('email'), data.get('telepon'), nip
                            ))
                            updated += 1
                            continue
                    
                    # Insert new
                    cursor.execute("""
                        INSERT INTO pegawai (nip, nama, pangkat, golongan, jabatan, unit_kerja, email, telepon)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        nip or None, nama, data.get('pangkat'), data.get('golongan'),
                        data.get('jabatan'), data.get('unit_kerja'),
                        data.get('email'), data.get('telepon')
                    ))
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Baris {i+1}: {str(e)}")
            
            conn.commit()
        
        return imported, updated, errors

    # =========================================================================
    # PENYEDIA (VENDOR) OPERATIONS
    # =========================================================================

    def get_all_penyedia(self, active_only: bool = True) -> List[Dict]:
        """Get all penyedia/vendor"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            if active_only:
                cursor.execute("SELECT * FROM penyedia WHERE is_active = 1 ORDER BY nama")
            else:
                cursor.execute("SELECT * FROM penyedia ORDER BY nama")
            return [dict(row) for row in cursor.fetchall()]

    def get_penyedia(self, penyedia_id: int) -> Optional[Dict]:
        """Get single penyedia by ID"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM penyedia WHERE id = ?", (penyedia_id,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def create_penyedia(self, data: Dict) -> int:
        """Create new penyedia"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO penyedia (
                    nama, nama_direktur, jabatan_direktur, alamat, kota,
                    npwp, no_rekening, nama_bank, nama_rekening,
                    telepon, email, is_pkp
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('nama'),
                data.get('nama_direktur'),
                data.get('jabatan_direktur', 'Direktur'),
                data.get('alamat'),
                data.get('kota'),
                data.get('npwp'),
                data.get('no_rekening'),
                data.get('nama_bank'),
                data.get('nama_rekening'),
                data.get('telepon'),
                data.get('email'),
                1 if data.get('is_pkp', True) else 0
            ))
            conn.commit()
            return cursor.lastrowid

    def update_penyedia(self, penyedia_id: int, data: Dict) -> bool:
        """Update penyedia"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE penyedia SET
                    nama = ?, nama_direktur = ?, jabatan_direktur = ?,
                    alamat = ?, kota = ?, npwp = ?, no_rekening = ?,
                    nama_bank = ?, nama_rekening = ?, telepon = ?, email = ?,
                    is_pkp = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nama'),
                data.get('nama_direktur'),
                data.get('jabatan_direktur', 'Direktur'),
                data.get('alamat'),
                data.get('kota'),
                data.get('npwp'),
                data.get('no_rekening'),
                data.get('nama_bank'),
                data.get('nama_rekening'),
                data.get('telepon'),
                data.get('email'),
                1 if data.get('is_pkp', True) else 0,
                penyedia_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def deactivate_penyedia(self, penyedia_id: int) -> bool:
        """Deactivate penyedia (soft delete)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE penyedia SET is_active = 0, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (penyedia_id,))
            conn.commit()
            return cursor.rowcount > 0

    def import_penyedia(self, filepath: str, format_type: str = 'excel') -> Tuple[int, int, List[str]]:
        """
        Import penyedia from Excel/JSON file
        Returns: (success_count, error_count, error_messages)
        """
        success = 0
        errors_count = 0
        errors = []

        try:
            if format_type == 'excel':
                from openpyxl import load_workbook
                wb = load_workbook(filepath)
                ws = wb.active

                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip() if cell.value else '')

                # Map column indices
                col_map = {}
                for idx, h in enumerate(headers):
                    h_lower = h.lower().replace(' ', '_').replace('.', '')
                    if h_lower in ['no', 'nomor']:
                        continue  # Skip row number
                    elif h_lower in ['nama_perusahaan', 'nama', 'perusahaan', 'company']:
                        col_map['nama'] = idx
                    elif h_lower in ['nama_direktur', 'direktur', 'pimpinan', 'nama_pimpinan']:
                        col_map['nama_direktur'] = idx
                    elif h_lower in ['jabatan_direktur', 'jabatan_pimpinan', 'jabatan']:
                        col_map['jabatan_direktur'] = idx
                    elif h_lower in ['alamat', 'address']:
                        col_map['alamat'] = idx
                    elif h_lower in ['kota', 'city', 'kabupaten']:
                        col_map['kota'] = idx
                    elif h_lower in ['npwp']:
                        col_map['npwp'] = idx
                    elif h_lower in ['no_rekening', 'norekening', 'rekening', 'account']:
                        col_map['no_rekening'] = idx
                    elif h_lower in ['bank', 'nama_bank']:
                        col_map['nama_bank'] = idx
                    elif h_lower in ['nama_rekening', 'namarekening', 'atas_nama']:
                        col_map['nama_rekening'] = idx
                    elif h_lower in ['telepon', 'telp', 'phone', 'hp']:
                        col_map['telepon'] = idx
                    elif h_lower in ['email', 'e-mail']:
                        col_map['email'] = idx
                    elif h_lower in ['pkp', 'is_pkp']:
                        col_map['is_pkp'] = idx

                # Read data rows
                with self.get_connection() as conn:
                    cursor = conn.cursor()

                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(v is None for v in row):
                            continue

                        try:
                            data = {}
                            for field, col_idx in col_map.items():
                                if col_idx < len(row):
                                    val = row[col_idx]
                                    data[field] = str(val).strip() if val else ''

                            # Skip if nama is empty
                            if not data.get('nama'):
                                continue

                            # Check PKP column
                            pkp_val = data.get('is_pkp', '').lower()
                            is_pkp = 1 if pkp_val in ['ya', 'yes', '1', 'true', 'pkp', 'v', 'x'] else 0

                            # Check if exists by NPWP
                            npwp = data.get('npwp', '').strip()
                            existing = None
                            if npwp:
                                cursor.execute("SELECT id FROM penyedia WHERE npwp = ?", (npwp,))
                                existing = cursor.fetchone()

                            if existing:
                                # Update existing
                                cursor.execute("""
                                    UPDATE penyedia SET
                                        nama = ?, nama_direktur = ?, jabatan_direktur = ?,
                                        alamat = ?, kota = ?, no_rekening = ?,
                                        nama_bank = ?, nama_rekening = ?, telepon = ?, email = ?,
                                        is_pkp = ?, is_active = 1, updated_at = CURRENT_TIMESTAMP
                                    WHERE npwp = ?
                                """, (
                                    data.get('nama'),
                                    data.get('nama_direktur'),
                                    data.get('jabatan_direktur', 'Direktur'),
                                    data.get('alamat'),
                                    data.get('kota'),
                                    data.get('no_rekening'),
                                    data.get('nama_bank'),
                                    data.get('nama_rekening'),
                                    data.get('telepon'),
                                    data.get('email'),
                                    is_pkp,
                                    npwp
                                ))
                            else:
                                # Insert new
                                cursor.execute("""
                                    INSERT INTO penyedia (
                                        nama, nama_direktur, jabatan_direktur, alamat, kota,
                                        npwp, no_rekening, nama_bank, nama_rekening,
                                        telepon, email, is_pkp
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (
                                    data.get('nama'),
                                    data.get('nama_direktur'),
                                    data.get('jabatan_direktur', 'Direktur'),
                                    data.get('alamat'),
                                    data.get('kota'),
                                    npwp or None,
                                    data.get('no_rekening'),
                                    data.get('nama_bank'),
                                    data.get('nama_rekening'),
                                    data.get('telepon'),
                                    data.get('email'),
                                    is_pkp
                                ))
                            success += 1

                        except Exception as e:
                            errors_count += 1
                            errors.append(f"Baris {row_idx}: {str(e)}")

                    conn.commit()

            else:  # JSON format
                import json
                with open(filepath, 'r', encoding='utf-8') as f:
                    data_list = json.load(f)

                if isinstance(data_list, dict):
                    data_list = data_list.get('penyedia', [])

                with self.get_connection() as conn:
                    cursor = conn.cursor()

                    for i, data in enumerate(data_list):
                        try:
                            if not data.get('nama'):
                                continue

                            npwp = data.get('npwp', '').strip()
                            existing = None
                            if npwp:
                                cursor.execute("SELECT id FROM penyedia WHERE npwp = ?", (npwp,))
                                existing = cursor.fetchone()

                            if existing:
                                self.update_penyedia(existing[0], data)
                            else:
                                self.create_penyedia(data)
                            success += 1

                        except Exception as e:
                            errors_count += 1
                            errors.append(f"Data {i+1}: {str(e)}")

                    conn.commit()

        except Exception as e:
            errors.append(f"Error membaca file: {str(e)}")

        return success, errors_count, errors

    def export_penyedia(self, filepath: str, format_type: str = 'excel') -> bool:
        """Export penyedia to Excel/JSON file"""
        try:
            penyedia_list = self.get_all_penyedia(active_only=False)

            if format_type == 'excel':
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Data Penyedia"

                # Headers
                headers = ['Nama Perusahaan', 'Nama Direktur', 'Jabatan', 'Alamat', 'Kota',
                          'NPWP', 'No Rekening', 'Bank', 'Nama Rekening', 'Telepon', 'Email', 'PKP', 'Status']

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                # Data
                for row, p in enumerate(penyedia_list, 2):
                    ws.cell(row=row, column=1, value=p.get('nama', '')).border = border
                    ws.cell(row=row, column=2, value=p.get('nama_direktur', '')).border = border
                    ws.cell(row=row, column=3, value=p.get('jabatan_direktur', '')).border = border
                    ws.cell(row=row, column=4, value=p.get('alamat', '')).border = border
                    ws.cell(row=row, column=5, value=p.get('kota', '')).border = border
                    ws.cell(row=row, column=6, value=p.get('npwp', '')).border = border
                    ws.cell(row=row, column=7, value=p.get('no_rekening', '')).border = border
                    ws.cell(row=row, column=8, value=p.get('nama_bank', '')).border = border
                    ws.cell(row=row, column=9, value=p.get('nama_rekening', '')).border = border
                    ws.cell(row=row, column=10, value=p.get('telepon', '')).border = border
                    ws.cell(row=row, column=11, value=p.get('email', '')).border = border
                    ws.cell(row=row, column=12, value='Ya' if p.get('is_pkp') else 'Tidak').border = border
                    ws.cell(row=row, column=13, value='Aktif' if p.get('is_active') else 'Non-aktif').border = border

                # Auto-width columns
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 40)

                wb.save(filepath)

            else:  # JSON format
                import json
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump({'penyedia': penyedia_list}, f, indent=2, default=str, ensure_ascii=False)

            return True
        except Exception as e:
            print(f"Error exporting penyedia: {e}")
            return False

    # =========================================================================
    # PAKET PEJABAT OPERATIONS
    # =========================================================================
    
    def get_paket_pejabat(self, paket_id: int) -> List[Dict]:
        """Get all pejabat assigned to a paket"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT pp.*, p.nip, p.nama, p.pangkat, p.golongan, p.jabatan
                FROM paket_pejabat pp
                JOIN pegawai p ON pp.pegawai_id = p.id
                WHERE pp.paket_id = ? AND pp.is_active = 1
                ORDER BY pp.peran, pp.urutan
            """, (paket_id,))
            return [dict(row) for row in cursor.fetchall()]
    
    def get_paket_pejabat_by_role(self, paket_id: int, peran: str) -> List[Dict]:
        """Get pejabat by role for a paket"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT pp.*, p.nip, p.nama, p.pangkat, p.golongan, p.jabatan,
                       p.gelar_depan, p.gelar_belakang
                FROM paket_pejabat pp
                JOIN pegawai p ON pp.pegawai_id = p.id
                WHERE pp.paket_id = ? AND pp.peran = ? AND pp.is_active = 1
                ORDER BY pp.urutan
            """, (paket_id, peran))
            return [dict(row) for row in cursor.fetchall()]
    
    def set_paket_pejabat(self, paket_id: int, peran: str, pegawai_id: int, 
                         urutan: int = 1, tanggal_penetapan: str = None,
                         nomor_sk: str = None) -> int:
        """Set pejabat for a paket role"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # For single-select roles (PPK, PPSPM, etc), deactivate existing
            if peran not in ['ANGGOTA_PPHP']:
                cursor.execute("""
                    UPDATE paket_pejabat SET is_active = 0
                    WHERE paket_id = ? AND peran = ?
                """, (paket_id, peran))
            
            # Check if already exists
            cursor.execute("""
                SELECT id FROM paket_pejabat
                WHERE paket_id = ? AND pegawai_id = ? AND peran = ?
            """, (paket_id, pegawai_id, peran))
            
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    UPDATE paket_pejabat SET 
                        is_active = 1, urutan = ?,
                        tanggal_penetapan = ?, nomor_sk = ?
                    WHERE id = ?
                """, (urutan, tanggal_penetapan, nomor_sk, existing[0]))
                result = existing[0]
            else:
                cursor.execute("""
                    INSERT INTO paket_pejabat (paket_id, pegawai_id, peran, urutan, 
                                              tanggal_penetapan, nomor_sk)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (paket_id, pegawai_id, peran, urutan, tanggal_penetapan, nomor_sk))
                result = cursor.lastrowid
            
            conn.commit()
            return result
    
    def remove_paket_pejabat(self, paket_id: int, peran: str, pegawai_id: int = None):
        """Remove pejabat from paket role"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            if pegawai_id:
                cursor.execute("""
                    UPDATE paket_pejabat SET is_active = 0
                    WHERE paket_id = ? AND peran = ? AND pegawai_id = ?
                """, (paket_id, peran, pegawai_id))
            else:
                cursor.execute("""
                    UPDATE paket_pejabat SET is_active = 0
                    WHERE paket_id = ? AND peran = ?
                """, (paket_id, peran))
            
            conn.commit()
    
    # =========================================================================
    # SURVEY HARGA DETAIL OPERATIONS
    # =========================================================================
    
    def get_survey_harga_detail(self, paket_id: int, item_id: int = None) -> List[Dict]:
        """Get survey harga details"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            if item_id:
                cursor.execute("""
                    SELECT * FROM survey_harga_detail
                    WHERE paket_id = ? AND item_id = ? AND is_active = 1
                    ORDER BY sumber_ke
                """, (paket_id, item_id))
            else:
                cursor.execute("""
                    SELECT * FROM survey_harga_detail
                    WHERE paket_id = ? AND is_active = 1
                    ORDER BY item_id, sumber_ke
                """, (paket_id,))
            
            return [dict(row) for row in cursor.fetchall()]
    
    def add_survey_harga_detail(self, paket_id: int, item_id: int, data: Dict) -> int:
        """Add survey harga detail"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO survey_harga_detail (
                    paket_id, item_id, sumber_ke, jenis_survey, nama_sumber,
                    alamat, kota, telepon, tanggal_survey, surveyor, nip_surveyor,
                    platform, link_produk, tanggal_akses,
                    nomor_kontrak, tahun_kontrak, instansi,
                    harga, keterangan, bukti_path
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                paket_id, item_id, data.get('sumber_ke', 1),
                data.get('jenis_survey'), data.get('nama_sumber'),
                data.get('alamat'), data.get('kota'), data.get('telepon'),
                data.get('tanggal_survey'), data.get('surveyor'), data.get('nip_surveyor'),
                data.get('platform'), data.get('link_produk'), data.get('tanggal_akses'),
                data.get('nomor_kontrak'), data.get('tahun_kontrak'), data.get('instansi'),
                data.get('harga', 0), data.get('keterangan'), data.get('bukti_path')
            ))
            
            conn.commit()
            return cursor.lastrowid
    
    def count_survey_per_item(self, paket_id: int) -> Dict[int, int]:
        """Count surveys per item"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT item_id, COUNT(*) as count
                FROM survey_harga_detail
                WHERE paket_id = ? AND is_active = 1
                GROUP BY item_id
            """, (paket_id,))
            return {row['item_id']: row['count'] for row in cursor.fetchall()}
    
    # =========================================================================
    # REVISI HARGA OPERATIONS
    # =========================================================================
    
    def create_revisi_harga(self, paket_id: int, data: Dict) -> int:
        """Create revisi harga record"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO revisi_harga (
                    paket_id, tanggal_revisi, nomor_ba_klarifikasi,
                    total_hps_awal, total_kontrak_hasil, selisih, persentase_selisih,
                    catatan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                paket_id, data.get('tanggal_revisi'), data.get('nomor_ba_klarifikasi'),
                data.get('total_hps_awal'), data.get('total_kontrak_hasil'),
                data.get('selisih'), data.get('persentase_selisih'),
                data.get('catatan')
            ))
            
            conn.commit()
            return cursor.lastrowid
    
    def get_revisi_harga(self, paket_id: int) -> Optional[Dict]:
        """Get latest revisi harga for paket"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM revisi_harga
                WHERE paket_id = ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (paket_id,))
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def approve_revisi_harga(self, revisi_id: int, approver_id: int) -> bool:
        """Approve revisi harga"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE revisi_harga SET
                    status = 'approved',
                    approved_by = ?,
                    approved_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (approver_id, revisi_id))
            conn.commit()
            return cursor.rowcount > 0
    
    # =========================================================================
    # HARGA LIFECYCLE OPERATIONS
    # =========================================================================
    
    def log_harga_change(self, paket_id: int, item_id: int, tahap: str,
                        harga_satuan: float, total: float, keterangan: str = None,
                        dokumen_ref: str = None) -> int:
        """Log harga change for audit trail"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO harga_lifecycle (
                    paket_id, item_id, tahap, harga_satuan, total,
                    keterangan, dokumen_ref
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (paket_id, item_id, tahap, harga_satuan, total, keterangan, dokumen_ref))
            conn.commit()
            return cursor.lastrowid
    
    def get_harga_history(self, paket_id: int, item_id: int = None) -> List[Dict]:
        """Get harga change history"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            if item_id:
                cursor.execute("""
                    SELECT * FROM harga_lifecycle
                    WHERE paket_id = ? AND item_id = ?
                    ORDER BY tanggal
                """, (paket_id, item_id))
            else:
                cursor.execute("""
                    SELECT * FROM harga_lifecycle
                    WHERE paket_id = ?
                    ORDER BY tanggal
                """, (paket_id,))
            
            return [dict(row) for row in cursor.fetchall()]
    
    # =========================================================================
    # WORKFLOW ENFORCEMENT
    # =========================================================================
    
    def check_stage_requirements(self, paket_id: int, stage_code: str) -> Tuple[bool, List[str]]:
        """
        Check if stage requirements are met
        Returns: (can_proceed, list_of_missing_requirements)
        """
        stage_config = next((s for s in WORKFLOW_STAGES_V4 if s['code'] == stage_code), None)
        if not stage_config:
            return False, ["Stage tidak dikenal"]
        
        missing = []
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Check required_before stages
            for required_stage in stage_config.get('required_before', []):
                cursor.execute("""
                    SELECT status FROM workflow_stage
                    WHERE paket_id = ? AND stage_code = ?
                """, (paket_id, required_stage))
                row = cursor.fetchone()
                
                if not row or row['status'] != 'completed':
                    stage_name = next((s['name'] for s in WORKFLOW_STAGES_V4 
                                      if s['code'] == required_stage), required_stage)
                    missing.append(f"Stage '{stage_name}' harus diselesaikan terlebih dahulu")
            
            # Stage-specific validations
            if stage_code == 'SURVEY':
                # Check spesifikasi has items
                cursor.execute("""
                    SELECT COUNT(*) FROM item_barang
                    WHERE paket_id = ? AND is_active = 1
                """, (paket_id,))
                if cursor.fetchone()[0] == 0:
                    missing.append("Belum ada item barang/spesifikasi")
            
            elif stage_code == 'HPS':
                # Check all items have at least 3 survey sources
                cursor.execute("""
                    SELECT ib.id, ib.uraian, COUNT(shd.id) as survey_count
                    FROM item_barang ib
                    LEFT JOIN survey_harga_detail shd ON ib.id = shd.item_id AND shd.is_active = 1
                    WHERE ib.paket_id = ? AND ib.is_active = 1
                    GROUP BY ib.id
                    HAVING survey_count < 3
                """, (paket_id,))
                
                incomplete_items = cursor.fetchall()
                for item in incomplete_items:
                    missing.append(f"Item '{item['uraian']}' belum memiliki 3 sumber survey")
            
            elif stage_code == 'KONTRAK_FINAL':
                # Check revisi harga approved
                cursor.execute("""
                    SELECT status FROM revisi_harga
                    WHERE paket_id = ?
                    ORDER BY created_at DESC LIMIT 1
                """, (paket_id,))
                row = cursor.fetchone()
                if not row or row['status'] != 'approved':
                    missing.append("Revisi harga belum disetujui")
            
            elif stage_code in ['SPP', 'SSP', 'KUITANSI']:
                # Check kontrak_final_locked
                cursor.execute("""
                    SELECT kontrak_final_locked, nilai_kontrak_final
                    FROM paket WHERE id = ?
                """, (paket_id,))
                row = cursor.fetchone()
                if not row or not row['kontrak_final_locked']:
                    missing.append("Kontrak Final belum dikunci")
                if not row or not row['nilai_kontrak_final']:
                    missing.append("Nilai Kontrak Final belum ditetapkan")
        
        return len(missing) == 0, missing
    
    def lock_stage(self, paket_id: int, stage_code: str) -> bool:
        """Lock a stage (mark as completed and prevent changes)"""
        lock_column = {
            'SPESIFIKASI': 'spesifikasi_locked',
            'SURVEY': 'survey_locked',
            'HPS': 'hps_locked',
            'KAK': 'kak_locked',
            'KONTRAK_DRAFT': 'kontrak_draft_locked',
            'KONTRAK_FINAL': 'kontrak_final_locked',
        }.get(stage_code)
        
        if not lock_column:
            return False
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Update workflow_stage
            cursor.execute("""
                UPDATE workflow_stage SET
                    status = 'completed',
                    completed_at = CURRENT_TIMESTAMP
                WHERE paket_id = ? AND stage_code = ?
            """, (paket_id, stage_code))
            
            # Update paket lock flag
            cursor.execute(f"""
                UPDATE paket SET {lock_column} = 1, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (paket_id,))
            
            conn.commit()
            return True
    
    def is_stage_locked(self, paket_id: int, stage_code: str) -> bool:
        """Check if stage is locked"""
        lock_column = {
            'SPESIFIKASI': 'spesifikasi_locked',
            'SURVEY': 'survey_locked',
            'HPS': 'hps_locked',
            'KAK': 'kak_locked',
            'KONTRAK_DRAFT': 'kontrak_draft_locked',
            'KONTRAK_FINAL': 'kontrak_final_locked',
        }.get(stage_code)
        
        if not lock_column:
            return False
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f"SELECT {lock_column} FROM paket WHERE id = ?", (paket_id,))
            row = cursor.fetchone()
            return bool(row and row[0])
    
    # =========================================================================
    # AUDIT TRAIL
    # =========================================================================
    
    def log_audit(self, action: str, table_name: str, record_id: int,
                 old_values: Dict = None, new_values: Dict = None,
                 user_name: str = None, notes: str = None):
        """Log audit trail entry"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO audit_log (
                    user_name, action, table_name, record_id,
                    old_values, new_values, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                user_name, action, table_name, record_id,
                json.dumps(old_values) if old_values else None,
                json.dumps(new_values) if new_values else None,
                notes
            ))
            conn.commit()

    # =========================================================================
    # PERJALANAN DINAS (Official Travel)
    # =========================================================================

    def create_perjalanan_dinas(self, data: Dict) -> int:
        """Create new perjalanan dinas"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO perjalanan_dinas (
                    tahun_anggaran, nama_kegiatan, maksud_perjalanan,
                    nomor_surat_tugas, nomor_sppd,
                    pelaksana_id, pelaksana_nama, pelaksana_nip, pelaksana_pangkat,
                    pelaksana_golongan, pelaksana_jabatan,
                    kota_asal, kota_tujuan, provinsi_tujuan, alamat_tujuan,
                    tanggal_surat_tugas, tanggal_berangkat, tanggal_kembali, lama_perjalanan,
                    sumber_dana, kode_akun,
                    biaya_transport, biaya_uang_harian, biaya_penginapan,
                    biaya_representasi, biaya_lain_lain, uang_muka,
                    ppk_id, ppk_nama, ppk_nip, ppk_jabatan,
                    bendahara_id, bendahara_nama, bendahara_nip,
                    status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran', TAHUN_ANGGARAN),
                data.get('nama_kegiatan'),
                data.get('maksud_perjalanan'),
                data.get('nomor_surat_tugas'),
                data.get('nomor_sppd'),
                data.get('pelaksana_id'),
                data.get('pelaksana_nama'),
                data.get('pelaksana_nip'),
                data.get('pelaksana_pangkat'),
                data.get('pelaksana_golongan'),
                data.get('pelaksana_jabatan'),
                data.get('kota_asal'),
                data.get('kota_tujuan'),
                data.get('provinsi_tujuan'),
                data.get('alamat_tujuan'),
                data.get('tanggal_surat_tugas'),
                data.get('tanggal_berangkat'),
                data.get('tanggal_kembali'),
                data.get('lama_perjalanan', 1),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('biaya_transport', 0),
                data.get('biaya_uang_harian', 0),
                data.get('biaya_penginapan', 0),
                data.get('biaya_representasi', 0),
                data.get('biaya_lain_lain', 0),
                data.get('uang_muka', 0),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('status', 'draft')
            ))
            conn.commit()
            return cursor.lastrowid

    def update_perjalanan_dinas(self, pd_id: int, data: Dict) -> bool:
        """Update perjalanan dinas"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE perjalanan_dinas SET
                    nama_kegiatan = ?, maksud_perjalanan = ?,
                    nomor_surat_tugas = ?, nomor_sppd = ?,
                    pelaksana_id = ?, pelaksana_nama = ?, pelaksana_nip = ?, pelaksana_pangkat = ?,
                    pelaksana_golongan = ?, pelaksana_jabatan = ?,
                    kota_asal = ?, kota_tujuan = ?, provinsi_tujuan = ?, alamat_tujuan = ?,
                    tanggal_surat_tugas = ?, tanggal_berangkat = ?, tanggal_kembali = ?, lama_perjalanan = ?,
                    sumber_dana = ?, kode_akun = ?,
                    biaya_transport = ?, biaya_uang_harian = ?, biaya_penginapan = ?,
                    biaya_representasi = ?, biaya_lain_lain = ?, uang_muka = ?,
                    ppk_id = ?, ppk_nama = ?, ppk_nip = ?, ppk_jabatan = ?,
                    bendahara_id = ?, bendahara_nama = ?, bendahara_nip = ?,
                    status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nama_kegiatan'),
                data.get('maksud_perjalanan'),
                data.get('nomor_surat_tugas'),
                data.get('nomor_sppd'),
                data.get('pelaksana_id'),
                data.get('pelaksana_nama'),
                data.get('pelaksana_nip'),
                data.get('pelaksana_pangkat'),
                data.get('pelaksana_golongan'),
                data.get('pelaksana_jabatan'),
                data.get('kota_asal'),
                data.get('kota_tujuan'),
                data.get('provinsi_tujuan'),
                data.get('alamat_tujuan'),
                data.get('tanggal_surat_tugas'),
                data.get('tanggal_berangkat'),
                data.get('tanggal_kembali'),
                data.get('lama_perjalanan', 1),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('biaya_transport', 0),
                data.get('biaya_uang_harian', 0),
                data.get('biaya_penginapan', 0),
                data.get('biaya_representasi', 0),
                data.get('biaya_lain_lain', 0),
                data.get('uang_muka', 0),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('status', 'draft'),
                pd_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def get_perjalanan_dinas(self, pd_id: int) -> Optional[Dict]:
        """Get single perjalanan dinas by ID with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT pd.*,
                    -- Pelaksana dari pegawai jika ada pelaksana_id
                    COALESCE(pel.nama, pd.pelaksana_nama) as pelaksana_nama,
                    COALESCE(pel.nip, pd.pelaksana_nip) as pelaksana_nip,
                    COALESCE(pel.pangkat, pd.pelaksana_pangkat) as pelaksana_pangkat,
                    COALESCE(pel.golongan, pd.pelaksana_golongan) as pelaksana_golongan,
                    COALESCE(pel.jabatan, pd.pelaksana_jabatan) as pelaksana_jabatan,
                    -- PPK dari pegawai jika ada ppk_id
                    COALESCE(ppk.nama, pd.ppk_nama) as ppk_nama,
                    COALESCE(ppk.nip, pd.ppk_nip) as ppk_nip,
                    COALESCE(ppk.jabatan, pd.ppk_jabatan) as ppk_jabatan,
                    -- Bendahara dari pegawai jika ada bendahara_id
                    COALESCE(bend.nama, pd.bendahara_nama) as bendahara_nama,
                    COALESCE(bend.nip, pd.bendahara_nip) as bendahara_nip
                FROM perjalanan_dinas pd
                LEFT JOIN pegawai pel ON pd.pelaksana_id = pel.id
                LEFT JOIN pegawai ppk ON pd.ppk_id = ppk.id
                LEFT JOIN pegawai bend ON pd.bendahara_id = bend.id
                WHERE pd.id = ?
            """, (pd_id,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def get_all_perjalanan_dinas(self, tahun: int = None) -> List[Dict]:
        """Get all perjalanan dinas with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            sql = """
                SELECT pd.*,
                    COALESCE(pel.nama, pd.pelaksana_nama) as pelaksana_nama,
                    COALESCE(pel.nip, pd.pelaksana_nip) as pelaksana_nip
                FROM perjalanan_dinas pd
                LEFT JOIN pegawai pel ON pd.pelaksana_id = pel.id
            """
            if tahun:
                sql += " WHERE pd.tahun_anggaran = ? ORDER BY pd.created_at DESC"
                cursor.execute(sql, (tahun,))
            else:
                sql += " ORDER BY pd.created_at DESC"
                cursor.execute(sql)
            return [dict(row) for row in cursor.fetchall()]

    def delete_perjalanan_dinas(self, pd_id: int) -> bool:
        """Delete perjalanan dinas"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM perjalanan_dinas WHERE id = ?", (pd_id,))
            conn.commit()
            return cursor.rowcount > 0

    # =========================================================================
    # SWAKELOLA (Self-managed Procurement)
    # =========================================================================

    def create_swakelola(self, data: Dict) -> int:
        """Create new swakelola activity"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO swakelola (
                    tahun_anggaran, nama_kegiatan, tipe_swakelola, tipe_swakelola_text,
                    deskripsi, output_kegiatan, nomor_kak, nomor_sk_tim,
                    nomor_sk_kpa, tanggal_sk_kpa, perihal_sk_kpa,
                    tanggal_sk_tim, tanggal_mulai, tanggal_selesai, jangka_waktu,
                    sumber_dana, kode_akun, pagu_swakelola,
                    pum_id, pum_nama, pum_nip, pum_jabatan, uang_muka, tanggal_uang_muka,
                    total_realisasi, tanggal_rampung,
                    ketua_id, ketua_nama, ketua_nip, ketua_jabatan,
                    sekretaris_id, sekretaris_nama, sekretaris_nip, anggota_tim,
                    ppk_id, ppk_nama, ppk_nip, ppk_jabatan,
                    bendahara_id, bendahara_nama, bendahara_nip,
                    status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran', TAHUN_ANGGARAN),
                data.get('nama_kegiatan'),
                data.get('tipe_swakelola', 0),
                data.get('tipe_swakelola_text'),
                data.get('deskripsi'),
                data.get('output_kegiatan'),
                data.get('nomor_kak'),
                data.get('nomor_sk_tim'),
                data.get('nomor_sk_kpa'),
                data.get('tanggal_sk_kpa'),
                data.get('perihal_sk_kpa'),
                data.get('tanggal_sk_tim'),
                data.get('tanggal_mulai'),
                data.get('tanggal_selesai'),
                data.get('jangka_waktu', 30),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('pagu_swakelola', 0),
                data.get('pum_id'),
                data.get('pum_nama'),
                data.get('pum_nip'),
                data.get('pum_jabatan'),
                data.get('uang_muka', 0),
                data.get('tanggal_uang_muka'),
                data.get('total_realisasi', 0),
                data.get('tanggal_rampung'),
                data.get('ketua_id'),
                data.get('ketua_nama'),
                data.get('ketua_nip'),
                data.get('ketua_jabatan'),
                data.get('sekretaris_id'),
                data.get('sekretaris_nama'),
                data.get('sekretaris_nip'),
                data.get('anggota_tim'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('status', 'draft')
            ))
            conn.commit()
            return cursor.lastrowid

    def update_swakelola(self, sw_id: int, data: Dict) -> bool:
        """Update swakelola activity"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE swakelola SET
                    nama_kegiatan = ?, tipe_swakelola = ?, tipe_swakelola_text = ?,
                    deskripsi = ?, output_kegiatan = ?, nomor_kak = ?, nomor_sk_tim = ?,
                    nomor_sk_kpa = ?, tanggal_sk_kpa = ?, perihal_sk_kpa = ?,
                    tanggal_sk_tim = ?, tanggal_mulai = ?, tanggal_selesai = ?, jangka_waktu = ?,
                    sumber_dana = ?, kode_akun = ?, pagu_swakelola = ?,
                    pum_id = ?, pum_nama = ?, pum_nip = ?, pum_jabatan = ?, uang_muka = ?, tanggal_uang_muka = ?,
                    total_realisasi = ?, tanggal_rampung = ?,
                    ketua_id = ?, ketua_nama = ?, ketua_nip = ?, ketua_jabatan = ?,
                    sekretaris_id = ?, sekretaris_nama = ?, sekretaris_nip = ?, anggota_tim = ?,
                    ppk_id = ?, ppk_nama = ?, ppk_nip = ?, ppk_jabatan = ?,
                    bendahara_id = ?, bendahara_nama = ?, bendahara_nip = ?,
                    status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nama_kegiatan'),
                data.get('tipe_swakelola', 0),
                data.get('tipe_swakelola_text'),
                data.get('deskripsi'),
                data.get('output_kegiatan'),
                data.get('nomor_kak'),
                data.get('nomor_sk_tim'),
                data.get('nomor_sk_kpa'),
                data.get('tanggal_sk_kpa'),
                data.get('perihal_sk_kpa'),
                data.get('tanggal_sk_tim'),
                data.get('tanggal_mulai'),
                data.get('tanggal_selesai'),
                data.get('jangka_waktu', 30),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('pagu_swakelola', 0),
                data.get('pum_id'),
                data.get('pum_nama'),
                data.get('pum_nip'),
                data.get('pum_jabatan'),
                data.get('uang_muka', 0),
                data.get('tanggal_uang_muka'),
                data.get('total_realisasi', 0),
                data.get('tanggal_rampung'),
                data.get('ketua_id'),
                data.get('ketua_nama'),
                data.get('ketua_nip'),
                data.get('ketua_jabatan'),
                data.get('sekretaris_id'),
                data.get('sekretaris_nama'),
                data.get('sekretaris_nip'),
                data.get('anggota_tim'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('status', 'draft'),
                sw_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def get_swakelola(self, sw_id: int) -> Optional[Dict]:
        """Get single swakelola by ID with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT sw.*,
                    -- PUM dari pegawai jika ada pum_id
                    COALESCE(pum.nama, sw.pum_nama) as pum_nama,
                    COALESCE(pum.nip, sw.pum_nip) as pum_nip,
                    COALESCE(pum.jabatan, sw.pum_jabatan) as pum_jabatan,
                    -- Ketua dari pegawai jika ada ketua_id
                    COALESCE(ketua.nama, sw.ketua_nama) as ketua_nama,
                    COALESCE(ketua.nip, sw.ketua_nip) as ketua_nip,
                    COALESCE(ketua.jabatan, sw.ketua_jabatan) as ketua_jabatan,
                    -- Sekretaris dari pegawai jika ada sekretaris_id
                    COALESCE(sekr.nama, sw.sekretaris_nama) as sekretaris_nama,
                    COALESCE(sekr.nip, sw.sekretaris_nip) as sekretaris_nip,
                    -- PPK dari pegawai jika ada ppk_id
                    COALESCE(ppk.nama, sw.ppk_nama) as ppk_nama,
                    COALESCE(ppk.nip, sw.ppk_nip) as ppk_nip,
                    COALESCE(ppk.jabatan, sw.ppk_jabatan) as ppk_jabatan,
                    -- Bendahara dari pegawai jika ada bendahara_id
                    COALESCE(bend.nama, sw.bendahara_nama) as bendahara_nama,
                    COALESCE(bend.nip, sw.bendahara_nip) as bendahara_nip
                FROM swakelola sw
                LEFT JOIN pegawai pum ON sw.pum_id = pum.id
                LEFT JOIN pegawai ketua ON sw.ketua_id = ketua.id
                LEFT JOIN pegawai sekr ON sw.sekretaris_id = sekr.id
                LEFT JOIN pegawai ppk ON sw.ppk_id = ppk.id
                LEFT JOIN pegawai bend ON sw.bendahara_id = bend.id
                WHERE sw.id = ?
            """, (sw_id,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def get_all_swakelola(self, tahun: int = None) -> List[Dict]:
        """Get all swakelola activities with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            sql = """
                SELECT sw.*,
                    COALESCE(ketua.nama, sw.ketua_nama) as ketua_nama
                FROM swakelola sw
                LEFT JOIN pegawai ketua ON sw.ketua_id = ketua.id
            """
            if tahun:
                sql += " WHERE sw.tahun_anggaran = ?"
                cursor.execute(sql + " ORDER BY sw.created_at DESC", (tahun,))
            else:
                cursor.execute(sql + " ORDER BY sw.created_at DESC")
            return [dict(row) for row in cursor.fetchall()]

    def delete_swakelola(self, sw_id: int) -> bool:
        """Delete swakelola"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM swakelola WHERE id = ?", (sw_id,))
            conn.commit()
            return cursor.rowcount > 0

    # ========================================================================
    # PJLP (Penyedia Jasa Lainnya Perorangan) METHODS
    # ========================================================================

    def create_pjlp(self, data: Dict) -> int:
        """Create new PJLP contract"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO pjlp (
                    tahun_anggaran, nomor_kontrak, tanggal_kontrak,
                    nama_pekerjaan, nama_pjlp, nik, npwp, alamat, telepon, email,
                    no_rekening, nama_bank, tanggal_mulai, tanggal_selesai,
                    jangka_waktu, honor_bulanan, total_nilai_kontrak,
                    sumber_dana, kode_akun,
                    ppk_id, ppk_nama, ppk_nip, ppk_jabatan,
                    bendahara_id, bendahara_nama, bendahara_nip, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran'),
                data.get('nomor_kontrak'),
                data.get('tanggal_kontrak'),
                data.get('nama_pekerjaan'),
                data.get('nama_pjlp'),
                data.get('nik'),
                data.get('npwp'),
                data.get('alamat'),
                data.get('telepon'),
                data.get('email'),
                data.get('no_rekening'),
                data.get('nama_bank'),
                data.get('tanggal_mulai'),
                data.get('tanggal_selesai'),
                data.get('jangka_waktu', 12),
                data.get('honor_bulanan', 0),
                data.get('total_nilai_kontrak', 0),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('status', 'aktif')
            ))
            conn.commit()
            return cursor.lastrowid

    def update_pjlp(self, pjlp_id: int, data: Dict) -> bool:
        """Update PJLP contract"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE pjlp SET
                    tahun_anggaran = ?, nomor_kontrak = ?, tanggal_kontrak = ?,
                    nama_pekerjaan = ?, nama_pjlp = ?, nik = ?, npwp = ?,
                    alamat = ?, telepon = ?, email = ?, no_rekening = ?, nama_bank = ?,
                    tanggal_mulai = ?, tanggal_selesai = ?, jangka_waktu = ?,
                    honor_bulanan = ?, total_nilai_kontrak = ?, sumber_dana = ?,
                    kode_akun = ?,
                    ppk_id = ?, ppk_nama = ?, ppk_nip = ?, ppk_jabatan = ?,
                    bendahara_id = ?, bendahara_nama = ?, bendahara_nip = ?, status = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('tahun_anggaran'),
                data.get('nomor_kontrak'),
                data.get('tanggal_kontrak'),
                data.get('nama_pekerjaan'),
                data.get('nama_pjlp'),
                data.get('nik'),
                data.get('npwp'),
                data.get('alamat'),
                data.get('telepon'),
                data.get('email'),
                data.get('no_rekening'),
                data.get('nama_bank'),
                data.get('tanggal_mulai'),
                data.get('tanggal_selesai'),
                data.get('jangka_waktu', 12),
                data.get('honor_bulanan', 0),
                data.get('total_nilai_kontrak', 0),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('status', 'aktif'),
                pjlp_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def get_pjlp(self, pjlp_id: int) -> Optional[Dict]:
        """Get single PJLP by ID with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.*,
                    COALESCE(ppk.nama, p.ppk_nama) as ppk_nama,
                    COALESCE(ppk.nip, p.ppk_nip) as ppk_nip,
                    COALESCE(ppk.jabatan, p.ppk_jabatan) as ppk_jabatan,
                    COALESCE(bend.nama, p.bendahara_nama) as bendahara_nama,
                    COALESCE(bend.nip, p.bendahara_nip) as bendahara_nip
                FROM pjlp p
                LEFT JOIN pegawai ppk ON p.ppk_id = ppk.id
                LEFT JOIN pegawai bend ON p.bendahara_id = bend.id
                WHERE p.id = ?
            """, (pjlp_id,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def get_all_pjlp(self, tahun: int = None, status: str = None) -> List[Dict]:
        """Get all PJLP contracts"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = "SELECT * FROM pjlp WHERE 1=1"
            params = []
            if tahun:
                query += " AND tahun_anggaran = ?"
                params.append(tahun)
            if status:
                query += " AND status = ?"
                params.append(status)
            query += " ORDER BY created_at DESC"
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]

    def delete_pjlp(self, pjlp_id: int) -> bool:
        """Delete PJLP contract"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pjlp WHERE id = ?", (pjlp_id,))
            conn.commit()
            return cursor.rowcount > 0

    # ========================================================================
    # PEMBAYARAN PJLP (Monthly Payment) METHODS
    # ========================================================================

    def create_pembayaran_pjlp(self, data: Dict) -> int:
        """Create new PJLP monthly payment"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO pembayaran_pjlp (
                    pjlp_id, bulan, tahun, nomor_kuitansi, tanggal_kuitansi,
                    nomor_spp, tanggal_spp, nilai_bruto, potongan_pajak,
                    potongan_lain, nilai_netto, kehadiran_hari, total_hari_kerja,
                    catatan_kinerja, status, tanggal_bayar
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('pjlp_id'),
                data.get('bulan'),
                data.get('tahun'),
                data.get('nomor_kuitansi'),
                data.get('tanggal_kuitansi'),
                data.get('nomor_spp'),
                data.get('tanggal_spp'),
                data.get('nilai_bruto', 0),
                data.get('potongan_pajak', 0),
                data.get('potongan_lain', 0),
                data.get('nilai_netto', 0),
                data.get('kehadiran_hari', 0),
                data.get('total_hari_kerja', 22),
                data.get('catatan_kinerja'),
                data.get('status', 'draft'),
                data.get('tanggal_bayar')
            ))
            conn.commit()
            return cursor.lastrowid

    def update_pembayaran_pjlp(self, payment_id: int, data: Dict) -> bool:
        """Update PJLP monthly payment"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE pembayaran_pjlp SET
                    nomor_kuitansi = ?, tanggal_kuitansi = ?,
                    nomor_spp = ?, tanggal_spp = ?,
                    nilai_bruto = ?, potongan_pajak = ?, potongan_lain = ?,
                    nilai_netto = ?, kehadiran_hari = ?, total_hari_kerja = ?,
                    catatan_kinerja = ?, status = ?, tanggal_bayar = ?
                WHERE id = ?
            """, (
                data.get('nomor_kuitansi'),
                data.get('tanggal_kuitansi'),
                data.get('nomor_spp'),
                data.get('tanggal_spp'),
                data.get('nilai_bruto', 0),
                data.get('potongan_pajak', 0),
                data.get('potongan_lain', 0),
                data.get('nilai_netto', 0),
                data.get('kehadiran_hari', 0),
                data.get('total_hari_kerja', 22),
                data.get('catatan_kinerja'),
                data.get('status', 'draft'),
                data.get('tanggal_bayar'),
                payment_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def get_pembayaran_pjlp(self, payment_id: int) -> Optional[Dict]:
        """Get single payment by ID"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM pembayaran_pjlp WHERE id = ?", (payment_id,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return None

    def get_pembayaran_by_pjlp(self, pjlp_id: int) -> List[Dict]:
        """Get all payments for a PJLP contract"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM pembayaran_pjlp
                WHERE pjlp_id = ?
                ORDER BY tahun DESC, bulan DESC
            """, (pjlp_id,))
            return [dict(row) for row in cursor.fetchall()]

    def get_pembayaran_by_bulan(self, bulan: int, tahun: int) -> List[Dict]:
        """Get all PJLP payments for a specific month"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.*, pj.nama_pjlp, pj.nama_pekerjaan, pj.honor_bulanan
                FROM pembayaran_pjlp p
                JOIN pjlp pj ON p.pjlp_id = pj.id
                WHERE p.bulan = ? AND p.tahun = ?
                ORDER BY pj.nama_pjlp
            """, (bulan, tahun))
            return [dict(row) for row in cursor.fetchall()]

    def delete_pembayaran_pjlp(self, payment_id: int) -> bool:
        """Delete PJLP payment"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pembayaran_pjlp WHERE id = ?", (payment_id,))
            conn.commit()
            return cursor.rowcount > 0

    def get_pjlp_summary(self, pjlp_id: int) -> Dict:
        """Get payment summary for a PJLP contract"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    COUNT(*) as total_bulan_dibayar,
                    SUM(nilai_bruto) as total_bruto,
                    SUM(potongan_pajak) as total_pajak,
                    SUM(potongan_lain) as total_potongan_lain,
                    SUM(nilai_netto) as total_netto
                FROM pembayaran_pjlp
                WHERE pjlp_id = ? AND status = 'dibayar'
            """, (pjlp_id,))
            row = cursor.fetchone()
            if row:
                return dict(row)
            return {
                'total_bulan_dibayar': 0,
                'total_bruto': 0,
                'total_pajak': 0,
                'total_potongan_lain': 0,
                'total_netto': 0
            }

    # ========================================================================
    # SK KPA (Surat Keputusan KPA) METHODS
    # ========================================================================

    def create_sk_kpa(self, data: Dict) -> int:
        """Create new SK KPA"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO sk_kpa (
                    tahun_anggaran, nomor_sk, tanggal_sk, perihal,
                    jenis_pembayaran, referensi_id, kpa_nama, kpa_nip, kpa_jabatan,
                    ppk_nama, ppk_nip, ppk_jabatan, bendahara_nama, bendahara_nip,
                    nilai_pembayaran, dasar_pembayaran, keterangan, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran'),
                data.get('nomor_sk'),
                data.get('tanggal_sk'),
                data.get('perihal'),
                data.get('jenis_pembayaran'),
                data.get('referensi_id'),
                data.get('kpa_nama'),
                data.get('kpa_nip'),
                data.get('kpa_jabatan'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('nilai_pembayaran', 0),
                data.get('dasar_pembayaran'),
                data.get('keterangan'),
                data.get('status', 'draft')
            ))
            conn.commit()
            return cursor.lastrowid

    def get_sk_kpa(self, sk_id: int) -> Optional[Dict]:
        """Get SK KPA by ID"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM sk_kpa WHERE id = ?", (sk_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_sk_kpa(self, tahun: int = None, jenis: str = None) -> List[Dict]:
        """Get all SK KPA"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = "SELECT * FROM sk_kpa WHERE 1=1"
            params = []
            if tahun:
                query += " AND tahun_anggaran = ?"
                params.append(tahun)
            if jenis:
                query += " AND jenis_pembayaran = ?"
                params.append(jenis)
            query += " ORDER BY tanggal_sk DESC"
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]

    def update_sk_kpa(self, sk_id: int, data: Dict) -> bool:
        """Update SK KPA"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE sk_kpa SET
                    nomor_sk = ?, tanggal_sk = ?, perihal = ?,
                    jenis_pembayaran = ?, referensi_id = ?,
                    kpa_nama = ?, kpa_nip = ?, kpa_jabatan = ?,
                    ppk_nama = ?, ppk_nip = ?, ppk_jabatan = ?,
                    bendahara_nama = ?, bendahara_nip = ?,
                    nilai_pembayaran = ?, dasar_pembayaran = ?,
                    keterangan = ?, status = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nomor_sk'),
                data.get('tanggal_sk'),
                data.get('perihal'),
                data.get('jenis_pembayaran'),
                data.get('referensi_id'),
                data.get('kpa_nama'),
                data.get('kpa_nip'),
                data.get('kpa_jabatan'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('ppk_jabatan'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('nilai_pembayaran', 0),
                data.get('dasar_pembayaran'),
                data.get('keterangan'),
                data.get('status', 'draft'),
                sk_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def delete_sk_kpa(self, sk_id: int) -> bool:
        """Delete SK KPA"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM sk_kpa WHERE id = ?", (sk_id,))
            conn.commit()
            return cursor.rowcount > 0

    # ========================================================================
    # HONORARIUM METHODS
    # ========================================================================

    def create_honorarium(self, data: Dict) -> int:
        """Create new honorarium"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO honorarium (
                    tahun_anggaran, nama_kegiatan, jenis_honorarium, kategori,
                    nomor_sk_kpa, tanggal_sk_kpa, nomor_spt, tanggal_spt,
                    nomor_kuitansi, tanggal_kuitansi, bulan, periode_mulai, periode_selesai,
                    sumber_dana, kode_akun, mak,
                    kpa_id, kpa_nama, kpa_nip, ppk_id, ppk_nama, ppk_nip,
                    bendahara_id, bendahara_nama, bendahara_nip,
                    total_bruto, total_pajak, total_netto, status, keterangan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran'),
                data.get('nama_kegiatan'),
                data.get('jenis_honorarium'),
                data.get('kategori', 'reguler'),
                data.get('nomor_sk_kpa'),
                data.get('tanggal_sk_kpa'),
                data.get('nomor_spt'),
                data.get('tanggal_spt'),
                data.get('nomor_kuitansi'),
                data.get('tanggal_kuitansi'),
                data.get('bulan'),
                data.get('periode_mulai'),
                data.get('periode_selesai'),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('mak'),
                data.get('kpa_id'),
                data.get('kpa_nama'),
                data.get('kpa_nip'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('total_bruto', 0),
                data.get('total_pajak', 0),
                data.get('total_netto', 0),
                data.get('status', 'draft'),
                data.get('keterangan')
            ))
            conn.commit()
            return cursor.lastrowid

    def get_honorarium(self, hon_id: int) -> Optional[Dict]:
        """Get honorarium by ID with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT h.*,
                    COALESCE(kpa.nama, h.kpa_nama) as kpa_nama,
                    COALESCE(kpa.nip, h.kpa_nip) as kpa_nip,
                    COALESCE(ppk.nama, h.ppk_nama) as ppk_nama,
                    COALESCE(ppk.nip, h.ppk_nip) as ppk_nip,
                    COALESCE(bend.nama, h.bendahara_nama) as bendahara_nama,
                    COALESCE(bend.nip, h.bendahara_nip) as bendahara_nip
                FROM honorarium h
                LEFT JOIN pegawai kpa ON h.kpa_id = kpa.id
                LEFT JOIN pegawai ppk ON h.ppk_id = ppk.id
                LEFT JOIN pegawai bend ON h.bendahara_id = bend.id
                WHERE h.id = ?
            """, (hon_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_honorarium(self, tahun: int = None, kategori: str = None) -> List[Dict]:
        """Get all honorarium"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = "SELECT * FROM honorarium WHERE 1=1"
            params = []
            if tahun:
                query += " AND tahun_anggaran = ?"
                params.append(tahun)
            if kategori:
                query += " AND kategori = ?"
                params.append(kategori)
            query += " ORDER BY created_at DESC"
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]

    def update_honorarium(self, hon_id: int, data: Dict) -> bool:
        """Update honorarium"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE honorarium SET
                    nama_kegiatan = ?, jenis_honorarium = ?, kategori = ?,
                    nomor_sk_kpa = ?, tanggal_sk_kpa = ?, nomor_spt = ?, tanggal_spt = ?,
                    nomor_kuitansi = ?, tanggal_kuitansi = ?,
                    bulan = ?, periode_mulai = ?, periode_selesai = ?,
                    sumber_dana = ?, kode_akun = ?, mak = ?,
                    kpa_id = ?, kpa_nama = ?, kpa_nip = ?,
                    ppk_id = ?, ppk_nama = ?, ppk_nip = ?,
                    bendahara_id = ?, bendahara_nama = ?, bendahara_nip = ?,
                    total_bruto = ?, total_pajak = ?, total_netto = ?,
                    status = ?, keterangan = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nama_kegiatan'),
                data.get('jenis_honorarium'),
                data.get('kategori', 'reguler'),
                data.get('nomor_sk_kpa'),
                data.get('tanggal_sk_kpa'),
                data.get('nomor_spt'),
                data.get('tanggal_spt'),
                data.get('nomor_kuitansi'),
                data.get('tanggal_kuitansi'),
                data.get('bulan'),
                data.get('periode_mulai'),
                data.get('periode_selesai'),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('mak'),
                data.get('kpa_id'),
                data.get('kpa_nama'),
                data.get('kpa_nip'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('total_bruto', 0),
                data.get('total_pajak', 0),
                data.get('total_netto', 0),
                data.get('status', 'draft'),
                data.get('keterangan'),
                hon_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def delete_honorarium(self, hon_id: int) -> bool:
        """Delete honorarium"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM honorarium WHERE id = ?", (hon_id,))
            conn.commit()
            return cursor.rowcount > 0

    # Honorarium Detail
    def add_honorarium_detail(self, data: Dict) -> int:
        """Add honorarium detail"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO honorarium_detail (
                    honorarium_id, nama, nip, jabatan, pangkat_golongan,
                    npwp, no_rekening, nama_bank, peran,
                    jumlah_jp, tarif_per_jp, nilai_bruto, pph21,
                    potongan_lain, nilai_netto, keterangan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('honorarium_id'),
                data.get('nama'),
                data.get('nip'),
                data.get('jabatan'),
                data.get('pangkat_golongan'),
                data.get('npwp'),
                data.get('no_rekening'),
                data.get('nama_bank'),
                data.get('peran'),
                data.get('jumlah_jp', 1),
                data.get('tarif_per_jp', 0),
                data.get('nilai_bruto', 0),
                data.get('pph21', 0),
                data.get('potongan_lain', 0),
                data.get('nilai_netto', 0),
                data.get('keterangan')
            ))
            conn.commit()
            return cursor.lastrowid

    def get_honorarium_details(self, honorarium_id: int) -> List[Dict]:
        """Get all details for honorarium"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM honorarium_detail
                WHERE honorarium_id = ?
                ORDER BY id
            """, (honorarium_id,))
            return [dict(row) for row in cursor.fetchall()]

    def update_honorarium_detail(self, detail_id: int, data: Dict) -> bool:
        """Update honorarium detail"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE honorarium_detail SET
                    nama = ?, nip = ?, jabatan = ?, pangkat_golongan = ?,
                    npwp = ?, no_rekening = ?, nama_bank = ?, peran = ?,
                    jumlah_jp = ?, tarif_per_jp = ?, nilai_bruto = ?,
                    pph21 = ?, potongan_lain = ?, nilai_netto = ?, keterangan = ?
                WHERE id = ?
            """, (
                data.get('nama'),
                data.get('nip'),
                data.get('jabatan'),
                data.get('pangkat_golongan'),
                data.get('npwp'),
                data.get('no_rekening'),
                data.get('nama_bank'),
                data.get('peran'),
                data.get('jumlah_jp', 1),
                data.get('tarif_per_jp', 0),
                data.get('nilai_bruto', 0),
                data.get('pph21', 0),
                data.get('potongan_lain', 0),
                data.get('nilai_netto', 0),
                data.get('keterangan'),
                detail_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def delete_honorarium_detail(self, detail_id: int) -> bool:
        """Delete honorarium detail"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM honorarium_detail WHERE id = ?", (detail_id,))
            conn.commit()
            return cursor.rowcount > 0

    # ========================================================================
    # JAMUAN TAMU METHODS
    # ========================================================================

    def create_jamuan_tamu(self, data: Dict) -> int:
        """Create new jamuan tamu"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO jamuan_tamu (
                    tahun_anggaran, nama_kegiatan, tanggal_kegiatan, tempat, kategori,
                    nama_tamu, instansi_tamu, jabatan_tamu, jumlah_tamu,
                    nomor_sk_kpa, tanggal_sk_kpa, nomor_nota_dinas, tanggal_nota_dinas,
                    nomor_kuitansi, tanggal_kuitansi,
                    sumber_dana, kode_akun, mak,
                    kpa_id, kpa_nama, kpa_nip, ppk_id, ppk_nama, ppk_nip,
                    bendahara_id, bendahara_nama, bendahara_nip,
                    biaya_konsumsi, biaya_akomodasi, biaya_transportasi, biaya_lainnya, total_biaya,
                    status, keterangan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran'),
                data.get('nama_kegiatan'),
                data.get('tanggal_kegiatan'),
                data.get('tempat'),
                data.get('kategori', 'reguler'),
                data.get('nama_tamu'),
                data.get('instansi_tamu'),
                data.get('jabatan_tamu'),
                data.get('jumlah_tamu', 1),
                data.get('nomor_sk_kpa'),
                data.get('tanggal_sk_kpa'),
                data.get('nomor_nota_dinas'),
                data.get('tanggal_nota_dinas'),
                data.get('nomor_kuitansi'),
                data.get('tanggal_kuitansi'),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('mak'),
                data.get('kpa_id'),
                data.get('kpa_nama'),
                data.get('kpa_nip'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('biaya_konsumsi', 0),
                data.get('biaya_akomodasi', 0),
                data.get('biaya_transportasi', 0),
                data.get('biaya_lainnya', 0),
                data.get('total_biaya', 0),
                data.get('status', 'draft'),
                data.get('keterangan')
            ))
            conn.commit()
            return cursor.lastrowid

    def get_jamuan_tamu(self, jt_id: int) -> Optional[Dict]:
        """Get jamuan tamu by ID with pegawai data"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT jt.*,
                    COALESCE(kpa.nama, jt.kpa_nama) as kpa_nama,
                    COALESCE(kpa.nip, jt.kpa_nip) as kpa_nip,
                    COALESCE(ppk.nama, jt.ppk_nama) as ppk_nama,
                    COALESCE(ppk.nip, jt.ppk_nip) as ppk_nip,
                    COALESCE(bend.nama, jt.bendahara_nama) as bendahara_nama,
                    COALESCE(bend.nip, jt.bendahara_nip) as bendahara_nip
                FROM jamuan_tamu jt
                LEFT JOIN pegawai kpa ON jt.kpa_id = kpa.id
                LEFT JOIN pegawai ppk ON jt.ppk_id = ppk.id
                LEFT JOIN pegawai bend ON jt.bendahara_id = bend.id
                WHERE jt.id = ?
            """, (jt_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_jamuan_tamu(self, tahun: int = None, kategori: str = None) -> List[Dict]:
        """Get all jamuan tamu"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = "SELECT * FROM jamuan_tamu WHERE 1=1"
            params = []
            if tahun:
                query += " AND tahun_anggaran = ?"
                params.append(tahun)
            if kategori:
                query += " AND kategori = ?"
                params.append(kategori)
            query += " ORDER BY tanggal_kegiatan DESC"
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]

    def update_jamuan_tamu(self, jt_id: int, data: Dict) -> bool:
        """Update jamuan tamu"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE jamuan_tamu SET
                    nama_kegiatan = ?, tanggal_kegiatan = ?, tempat = ?, kategori = ?,
                    nama_tamu = ?, instansi_tamu = ?, jabatan_tamu = ?, jumlah_tamu = ?,
                    nomor_sk_kpa = ?, tanggal_sk_kpa = ?,
                    nomor_nota_dinas = ?, tanggal_nota_dinas = ?,
                    nomor_kuitansi = ?, tanggal_kuitansi = ?,
                    sumber_dana = ?, kode_akun = ?, mak = ?,
                    kpa_id = ?, kpa_nama = ?, kpa_nip = ?,
                    ppk_id = ?, ppk_nama = ?, ppk_nip = ?,
                    bendahara_id = ?, bendahara_nama = ?, bendahara_nip = ?,
                    biaya_konsumsi = ?, biaya_akomodasi = ?,
                    biaya_transportasi = ?, biaya_lainnya = ?, total_biaya = ?,
                    status = ?, keterangan = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('nama_kegiatan'),
                data.get('tanggal_kegiatan'),
                data.get('tempat'),
                data.get('kategori', 'reguler'),
                data.get('nama_tamu'),
                data.get('instansi_tamu'),
                data.get('jabatan_tamu'),
                data.get('jumlah_tamu', 1),
                data.get('nomor_sk_kpa'),
                data.get('tanggal_sk_kpa'),
                data.get('nomor_nota_dinas'),
                data.get('tanggal_nota_dinas'),
                data.get('nomor_kuitansi'),
                data.get('tanggal_kuitansi'),
                data.get('sumber_dana', 'DIPA'),
                data.get('kode_akun'),
                data.get('mak'),
                data.get('kpa_id'),
                data.get('kpa_nama'),
                data.get('kpa_nip'),
                data.get('ppk_id'),
                data.get('ppk_nama'),
                data.get('ppk_nip'),
                data.get('bendahara_id'),
                data.get('bendahara_nama'),
                data.get('bendahara_nip'),
                data.get('biaya_konsumsi', 0),
                data.get('biaya_akomodasi', 0),
                data.get('biaya_transportasi', 0),
                data.get('biaya_lainnya', 0),
                data.get('total_biaya', 0),
                data.get('status', 'draft'),
                data.get('keterangan'),
                jt_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def delete_jamuan_tamu(self, jt_id: int) -> bool:
        """Delete jamuan tamu"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM jamuan_tamu WHERE id = ?", (jt_id,))
            conn.commit()
            return cursor.rowcount > 0

    # ========================================================================
    # PAGU ANGGARAN / FA DETAIL METHODS
    # ========================================================================

    def create_pagu_anggaran(self, data: Dict) -> int:
        """Create new pagu anggaran entry"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # Calculate sisa
            jumlah = data.get('jumlah', 0) or 0
            realisasi = data.get('realisasi', 0) or 0
            sisa = jumlah - realisasi
            persen = (realisasi / jumlah * 100) if jumlah > 0 else 0

            cursor.execute("""
                INSERT INTO pagu_anggaran (
                    tahun_anggaran, kode_program, kode_kegiatan, kode_kro, kode_ro,
                    kode_komponen, kode_sub_komponen, kode_akun, kode_detail, kode_full,
                    level_kode, parent_id, uraian, volume, satuan,
                    harga_satuan, jumlah, realisasi, sisa, persen_realisasi,
                    sumber_dana, is_locked, is_blokir, keterangan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun_anggaran'),
                data.get('kode_program'),
                data.get('kode_kegiatan'),
                data.get('kode_kro'),
                data.get('kode_ro'),
                data.get('kode_komponen'),
                data.get('kode_sub_komponen'),
                data.get('kode_akun'),
                data.get('kode_detail'),
                data.get('kode_full'),
                data.get('level_kode', 8),
                data.get('parent_id'),
                data.get('uraian'),
                data.get('volume', 0),
                data.get('satuan'),
                data.get('harga_satuan', 0),
                jumlah,
                realisasi,
                sisa,
                persen,
                data.get('sumber_dana', 'RM'),
                data.get('is_locked', 0),
                data.get('is_blokir', 0),
                data.get('keterangan')
            ))
            conn.commit()
            return cursor.lastrowid

    def get_pagu_anggaran(self, pagu_id: int) -> Optional[Dict]:
        """Get pagu anggaran by ID"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM pagu_anggaran WHERE id = ?", (pagu_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_pagu_anggaran(self, tahun: int = None, kode_akun: str = None,
                              level: int = None, parent_id: int = None) -> List[Dict]:
        """Get all pagu anggaran with filters"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = "SELECT * FROM pagu_anggaran WHERE 1=1"
            params = []

            if tahun:
                query += " AND tahun_anggaran = ?"
                params.append(tahun)
            if kode_akun:
                query += " AND kode_akun LIKE ?"
                params.append(f"{kode_akun}%")
            if level:
                query += " AND level_kode = ?"
                params.append(level)
            if parent_id is not None:
                query += " AND parent_id = ?"
                params.append(parent_id)

            query += " ORDER BY kode_full, id"
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]

    def get_pagu_by_kode(self, tahun: int, kode_full: str) -> Optional[Dict]:
        """Get pagu by full code"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM pagu_anggaran
                WHERE tahun_anggaran = ? AND kode_full = ?
            """, (tahun, kode_full))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_pagu_summary(self, tahun: int) -> Dict:
        """Get summary of pagu anggaran"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    COUNT(*) as total_item,
                    SUM(jumlah) as total_pagu,
                    SUM(realisasi) as total_realisasi,
                    SUM(sisa) as total_sisa
                FROM pagu_anggaran
                WHERE tahun_anggaran = ? AND level_kode = 8
            """, (tahun,))
            row = cursor.fetchone()
            if row:
                result = dict(row)
                total_pagu = result.get('total_pagu', 0) or 0
                total_real = result.get('total_realisasi', 0) or 0
                result['persen_realisasi'] = (total_real / total_pagu * 100) if total_pagu > 0 else 0
                return result
            return {
                'total_item': 0,
                'total_pagu': 0,
                'total_realisasi': 0,
                'total_sisa': 0,
                'persen_realisasi': 0
            }

    def get_pagu_by_akun_group(self, tahun: int) -> List[Dict]:
        """Get pagu grouped by kode akun"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    SUBSTR(kode_akun, 1, 2) as grup_akun,
                    SUM(jumlah) as total_pagu,
                    SUM(realisasi) as total_realisasi,
                    SUM(sisa) as total_sisa,
                    COUNT(*) as jumlah_item
                FROM pagu_anggaran
                WHERE tahun_anggaran = ? AND level_kode = 8
                GROUP BY SUBSTR(kode_akun, 1, 2)
                ORDER BY grup_akun
            """, (tahun,))
            return [dict(row) for row in cursor.fetchall()]

    def update_pagu_anggaran(self, pagu_id: int, data: Dict) -> bool:
        """Update pagu anggaran"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            jumlah = data.get('jumlah', 0) or 0
            realisasi = data.get('realisasi', 0) or 0
            sisa = jumlah - realisasi
            persen = (realisasi / jumlah * 100) if jumlah > 0 else 0

            cursor.execute("""
                UPDATE pagu_anggaran SET
                    kode_program = ?, kode_kegiatan = ?, kode_kro = ?, kode_ro = ?,
                    kode_komponen = ?, kode_sub_komponen = ?, kode_akun = ?,
                    kode_detail = ?, kode_full = ?, level_kode = ?, parent_id = ?,
                    uraian = ?, volume = ?, satuan = ?, harga_satuan = ?,
                    jumlah = ?, realisasi = ?, sisa = ?, persen_realisasi = ?,
                    sumber_dana = ?, is_locked = ?, is_blokir = ?, keterangan = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('kode_program'),
                data.get('kode_kegiatan'),
                data.get('kode_kro'),
                data.get('kode_ro'),
                data.get('kode_komponen'),
                data.get('kode_sub_komponen'),
                data.get('kode_akun'),
                data.get('kode_detail'),
                data.get('kode_full'),
                data.get('level_kode', 8),
                data.get('parent_id'),
                data.get('uraian'),
                data.get('volume', 0),
                data.get('satuan'),
                data.get('harga_satuan', 0),
                jumlah,
                realisasi,
                sisa,
                persen,
                data.get('sumber_dana', 'RM'),
                data.get('is_locked', 0),
                data.get('is_blokir', 0),
                data.get('keterangan'),
                pagu_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def update_pagu_realisasi(self, pagu_id: int, realisasi: float) -> bool:
        """Update realisasi pagu"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE pagu_anggaran SET
                    realisasi = ?,
                    sisa = jumlah - ?,
                    persen_realisasi = CASE WHEN jumlah > 0 THEN (? / jumlah * 100) ELSE 0 END,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (realisasi, realisasi, realisasi, pagu_id))
            conn.commit()
            return cursor.rowcount > 0

    def delete_pagu_anggaran(self, pagu_id: int) -> bool:
        """Delete pagu anggaran"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pagu_anggaran WHERE id = ?", (pagu_id,))
            conn.commit()
            return cursor.rowcount > 0

    def delete_all_pagu_tahun(self, tahun: int) -> int:
        """Delete all pagu for a year (for reimport)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pagu_anggaran WHERE tahun_anggaran = ?", (tahun,))
            conn.commit()
            return cursor.rowcount

    def bulk_insert_pagu(self, data_list: List[Dict], upsert: bool = False) -> int:
        """Bulk insert atau update pagu anggaran

        Args:
            data_list: List of pagu data dictionaries
            upsert: Jika True, update data yang sudah ada berdasarkan kode_full
        """
        with self.get_connection() as conn:
            cursor = conn.cursor()
            count = 0
            updated = 0

            for data in data_list:
                jumlah = data.get('jumlah', 0) or 0
                realisasi = data.get('realisasi', 0) or 0
                sisa = jumlah - realisasi
                persen = (realisasi / jumlah * 100) if jumlah > 0 else 0

                # Generate nomor_mak dari kode_akun + kode_detail
                kode_akun = data.get('kode_akun', '')
                kode_detail = data.get('kode_detail', '')
                nomor_mak = f"{kode_akun}.{kode_detail}" if kode_akun and kode_detail else ''

                if upsert:
                    # Cek apakah data sudah ada
                    cursor.execute("""
                        SELECT id, realisasi FROM pagu_anggaran
                        WHERE tahun_anggaran = ? AND kode_full = ?
                    """, (data.get('tahun_anggaran'), data.get('kode_full')))
                    existing = cursor.fetchone()

                    if existing:
                        # Update data yang ada, pertahankan realisasi jika tidak di-set
                        existing_realisasi = existing[1] or 0
                        if realisasi == 0 and existing_realisasi > 0:
                            realisasi = existing_realisasi
                            sisa = jumlah - realisasi
                            persen = (realisasi / jumlah * 100) if jumlah > 0 else 0

                        cursor.execute("""
                            UPDATE pagu_anggaran SET
                                kode_program = ?, kode_kegiatan = ?, kode_kro = ?, kode_ro = ?,
                                kode_komponen = ?, kode_sub_komponen = ?, kode_akun = ?, kode_detail = ?,
                                level_kode = ?, uraian = ?, volume = ?, satuan = ?,
                                harga_satuan = ?, jumlah = ?, realisasi = ?, sisa = ?, persen_realisasi = ?,
                                sumber_dana = ?, nomor_mak = ?, updated_at = CURRENT_TIMESTAMP
                            WHERE id = ?
                        """, (
                            data.get('kode_program'),
                            data.get('kode_kegiatan'),
                            data.get('kode_kro'),
                            data.get('kode_ro'),
                            data.get('kode_komponen'),
                            data.get('kode_sub_komponen'),
                            kode_akun,
                            kode_detail,
                            data.get('level_kode', 8),
                            data.get('uraian'),
                            data.get('volume', 0),
                            data.get('satuan'),
                            data.get('harga_satuan', 0),
                            jumlah,
                            realisasi,
                            sisa,
                            persen,
                            data.get('sumber_dana', 'RM'),
                            nomor_mak,
                            existing[0]
                        ))
                        updated += 1
                        continue

                # Insert baru
                cursor.execute("""
                    INSERT INTO pagu_anggaran (
                        tahun_anggaran, kode_program, kode_kegiatan, kode_kro, kode_ro,
                        kode_komponen, kode_sub_komponen, kode_akun, kode_detail, kode_full,
                        level_kode, parent_id, uraian, volume, satuan,
                        harga_satuan, jumlah, realisasi, sisa, persen_realisasi,
                        sumber_dana, is_locked, is_blokir, nomor_mak, keterangan
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('tahun_anggaran'),
                    data.get('kode_program'),
                    data.get('kode_kegiatan'),
                    data.get('kode_kro'),
                    data.get('kode_ro'),
                    data.get('kode_komponen'),
                    data.get('kode_sub_komponen'),
                    kode_akun,
                    kode_detail,
                    data.get('kode_full'),
                    data.get('level_kode', 8),
                    data.get('parent_id'),
                    data.get('uraian'),
                    data.get('volume', 0),
                    data.get('satuan'),
                    data.get('harga_satuan', 0),
                    jumlah,
                    realisasi,
                    sisa,
                    persen,
                    data.get('sumber_dana', 'RM'),
                    data.get('is_locked', 0),
                    data.get('is_blokir', 0),
                    nomor_mak,
                    data.get('keterangan')
                ))
                count += 1
            conn.commit()
            return count + updated

    # Realisasi tracking
    def add_realisasi(self, data: Dict) -> int:
        """Add realisasi record"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO realisasi_anggaran (
                    pagu_id, bulan, tahun, nilai_realisasi,
                    nomor_sp2d, tanggal_sp2d, keterangan
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('pagu_id'),
                data.get('bulan'),
                data.get('tahun'),
                data.get('nilai_realisasi', 0),
                data.get('nomor_sp2d'),
                data.get('tanggal_sp2d'),
                data.get('keterangan')
            ))
            conn.commit()

            # Update total realisasi di pagu
            self._update_total_realisasi(data.get('pagu_id'))
            return cursor.lastrowid

    def _update_total_realisasi(self, pagu_id: int):
        """Update total realisasi dari semua bulan"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT SUM(nilai_realisasi) as total
                FROM realisasi_anggaran WHERE pagu_id = ?
            """, (pagu_id,))
            row = cursor.fetchone()
            total = row['total'] if row and row['total'] else 0
            self.update_pagu_realisasi(pagu_id, total)

    def get_realisasi_by_pagu(self, pagu_id: int) -> List[Dict]:
        """Get all realisasi for a pagu"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM realisasi_anggaran
                WHERE pagu_id = ?
                ORDER BY tahun, bulan
            """, (pagu_id,))
            return [dict(row) for row in cursor.fetchall()]

    # =========================================================================
    # HONORARIUM PENGELOLA KEUANGAN
    # =========================================================================

    def create_honorarium_pengelola(self, data: Dict) -> int:
        """Create honorarium pengelola keuangan"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO honorarium_pengelola (
                    tahun, bulan, jabatan, pegawai_id,
                    jumlah, pajak, netto, pagu_id, keterangan
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('tahun'),
                data.get('bulan'),
                data.get('jabatan'),
                data.get('pegawai_id'),
                data.get('jumlah', 0),
                data.get('pajak', 0),
                data.get('netto', 0),
                data.get('pagu_id'),
                data.get('keterangan')
            ))
            conn.commit()
            return cursor.lastrowid

    def update_honorarium_pengelola(self, hp_id: int, data: Dict) -> bool:
        """Update honorarium pengelola keuangan"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE honorarium_pengelola SET
                    tahun = ?, bulan = ?, jabatan = ?, pegawai_id = ?,
                    jumlah = ?, pajak = ?, netto = ?, pagu_id = ?, keterangan = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                data.get('tahun'),
                data.get('bulan'),
                data.get('jabatan'),
                data.get('pegawai_id'),
                data.get('jumlah', 0),
                data.get('pajak', 0),
                data.get('netto', 0),
                data.get('pagu_id'),
                data.get('keterangan'),
                hp_id
            ))
            conn.commit()
            return cursor.rowcount > 0

    def delete_honorarium_pengelola(self, hp_id: int) -> bool:
        """Delete honorarium pengelola keuangan"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM honorarium_pengelola WHERE id = ?", (hp_id,))
            conn.commit()
            return cursor.rowcount > 0

    def get_honorarium_pengelola(self, hp_id: int) -> Optional[Dict]:
        """Get single honorarium pengelola by ID"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT hp.*, p.nama as pegawai_nama, p.nip as pegawai_nip
                FROM honorarium_pengelola hp
                LEFT JOIN pegawai p ON hp.pegawai_id = p.id
                WHERE hp.id = ?
            """, (hp_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_honorarium_pengelola(self, tahun: int = None, bulan: int = None) -> List[Dict]:
        """Get all honorarium pengelola with filters"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            sql = """
                SELECT hp.*, p.nama as pegawai_nama, p.nip as pegawai_nip
                FROM honorarium_pengelola hp
                LEFT JOIN pegawai p ON hp.pegawai_id = p.id
            """
            conditions = []
            params = []
            if tahun:
                conditions.append("hp.tahun = ?")
                params.append(tahun)
            if bulan:
                conditions.append("hp.bulan = ?")
                params.append(bulan)
            if conditions:
                sql += " WHERE " + " AND ".join(conditions)
            sql += " ORDER BY hp.tahun DESC, hp.bulan DESC, hp.jabatan"
            cursor.execute(sql, params)
            return [dict(row) for row in cursor.fetchall()]

    def get_pagu_for_honorarium_pengelola(self, tahun: int) -> List[Dict]:
        """Get pagu anggaran items for honorarium pengelola keuangan"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            # Akun 52xxxx adalah Belanja Barang (termasuk honor pengelola keuangan)
            cursor.execute("""
                SELECT id, kode_akun, uraian, jumlah, realisasi, sisa
                FROM pagu_anggaran
                WHERE tahun_anggaran = ?
                    AND kode_akun LIKE '52%'
                    AND (uraian LIKE '%honor%' OR uraian LIKE '%pengelola%'
                         OR uraian LIKE '%keuangan%' OR uraian LIKE '%bendahara%'
                         OR uraian LIKE '%ppk%' OR uraian LIKE '%ppspm%'
                         OR uraian LIKE '%operator%' OR uraian LIKE '%KPA%'
                         OR uraian LIKE '%PNBP%')
                ORDER BY kode_akun
            """, (tahun,))
            return [dict(row) for row in cursor.fetchall()]

    # ========================================================================
    # DIPA (Daftar Isian Pelaksanaan Anggaran) Methods
    # ========================================================================

    def import_dipa_csv(self, filepath: str, tahun_anggaran: int = None) -> Tuple[int, int, List[str]]:
        """
        Import DIPA data from CSV file.

        Expected columns: KDSATKER, KODE_PROGRAM, KODE_KEGIATAN, KODE_OUTPUT,
                         KODE_AKUN, URAIAN_ITEM, VOLKEG, SATKEG, HARGASAT, TOTAL,
                         POK_NILAI_1 to POK_NILAI_12

        Returns: (success_count, error_count, error_messages)
        """
        import csv

        if tahun_anggaran is None:
            tahun_anggaran = TAHUN_ANGGARAN

        success = 0
        errors_count = 0
        errors = []

        try:
            # Try to detect encoding
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
            content = None

            for enc in encodings:
                try:
                    with open(filepath, 'r', encoding=enc) as f:
                        content = f.read()
                    break
                except UnicodeDecodeError:
                    continue

            if content is None:
                return 0, 1, ["Tidak dapat membaca file CSV - format encoding tidak dikenali"]

            # Parse CSV
            lines = content.strip().split('\n')
            if len(lines) < 2:
                return 0, 1, ["File CSV kosong atau hanya berisi header"]

            # Get headers
            reader = csv.reader(lines)
            headers = next(reader)
            headers = [h.strip().upper() for h in headers]

            # Map column indices
            col_map = {}
            required_cols = ['KODE_AKUN', 'URAIAN_ITEM']

            for idx, h in enumerate(headers):
                h_clean = h.replace(' ', '_').replace('-', '_')
                if h_clean in ['KDSATKER', 'KODE_SATKER']:
                    col_map['kode_satker'] = idx
                elif h_clean in ['KODE_PROGRAM']:
                    col_map['kode_program'] = idx
                elif h_clean in ['KODE_KEGIATAN']:
                    col_map['kode_kegiatan'] = idx
                elif h_clean in ['KODE_OUTPUT']:
                    col_map['kode_output'] = idx
                elif h_clean in ['KODE_AKUN']:
                    col_map['kode_akun'] = idx
                elif h_clean in ['URAIAN_ITEM', 'URAIAN']:
                    col_map['uraian_item'] = idx
                elif h_clean in ['VOLKEG', 'VOLUME']:
                    col_map['volume'] = idx
                elif h_clean in ['SATKEG', 'SATUAN']:
                    col_map['satuan'] = idx
                elif h_clean in ['HARGASAT', 'HARGA_SATUAN']:
                    col_map['harga_satuan'] = idx
                elif h_clean in ['TOTAL', 'JUMLAH']:
                    col_map['total'] = idx
                # POK monthly values
                for i in range(1, 13):
                    if h_clean == f'POK_NILAI_{i}':
                        col_map[f'pok_nilai_{i}'] = idx

            # Check required columns
            missing = [c for c in required_cols if c.lower() not in col_map]
            if missing:
                return 0, 1, [f"Kolom wajib tidak ditemukan: {', '.join(missing)}"]

            # Process data rows
            with self.get_connection() as conn:
                cursor = conn.cursor()

                for row_idx, row_text in enumerate(lines[1:], start=2):
                    if not row_text.strip():
                        continue

                    try:
                        # Parse row
                        row = list(csv.reader([row_text]))[0]

                        def get_val(key, default=''):
                            if key in col_map and col_map[key] < len(row):
                                val = row[col_map[key]]
                                return val.strip() if val else default
                            return default

                        def get_num(key, default=0):
                            val = get_val(key, '')
                            if not val:
                                return default
                            try:
                                # Remove thousand separators
                                val = val.replace('.', '').replace(',', '.')
                                return float(val)
                            except:
                                return default

                        kode_akun = get_val('kode_akun')
                        uraian_item = get_val('uraian_item')

                        # Skip if required fields are empty
                        if not kode_akun or not uraian_item:
                            continue

                        kode_satker = get_val('kode_satker')
                        kode_program = get_val('kode_program')
                        kode_kegiatan = get_val('kode_kegiatan')
                        kode_output = get_val('kode_output')

                        # Build kode_full
                        kode_parts = [p for p in [kode_program, kode_kegiatan, kode_output, kode_akun] if p]
                        kode_full = '.'.join(kode_parts)

                        # Check if exists
                        cursor.execute("""
                            SELECT id FROM dipa
                            WHERE tahun_anggaran = ? AND kode_satker = ?
                                  AND kode_akun = ? AND uraian_item = ?
                        """, (tahun_anggaran, kode_satker, kode_akun, uraian_item))
                        existing = cursor.fetchone()

                        data_values = (
                            kode_satker,
                            kode_program,
                            kode_kegiatan,
                            kode_output,
                            kode_akun,
                            uraian_item,
                            get_num('volume'),
                            get_val('satuan'),
                            get_num('harga_satuan'),
                            get_num('total'),
                            get_num('pok_nilai_1'),
                            get_num('pok_nilai_2'),
                            get_num('pok_nilai_3'),
                            get_num('pok_nilai_4'),
                            get_num('pok_nilai_5'),
                            get_num('pok_nilai_6'),
                            get_num('pok_nilai_7'),
                            get_num('pok_nilai_8'),
                            get_num('pok_nilai_9'),
                            get_num('pok_nilai_10'),
                            get_num('pok_nilai_11'),
                            get_num('pok_nilai_12'),
                            kode_full,
                        )

                        if existing:
                            # Update existing
                            cursor.execute("""
                                UPDATE dipa SET
                                    kode_satker = ?, kode_program = ?, kode_kegiatan = ?,
                                    kode_output = ?, kode_akun = ?, uraian_item = ?,
                                    volume = ?, satuan = ?, harga_satuan = ?, total = ?,
                                    pok_nilai_1 = ?, pok_nilai_2 = ?, pok_nilai_3 = ?,
                                    pok_nilai_4 = ?, pok_nilai_5 = ?, pok_nilai_6 = ?,
                                    pok_nilai_7 = ?, pok_nilai_8 = ?, pok_nilai_9 = ?,
                                    pok_nilai_10 = ?, pok_nilai_11 = ?, pok_nilai_12 = ?,
                                    kode_full = ?, updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                            """, data_values + (existing[0],))
                        else:
                            # Insert new
                            cursor.execute("""
                                INSERT INTO dipa (
                                    tahun_anggaran, kode_satker, kode_program, kode_kegiatan,
                                    kode_output, kode_akun, uraian_item,
                                    volume, satuan, harga_satuan, total,
                                    pok_nilai_1, pok_nilai_2, pok_nilai_3,
                                    pok_nilai_4, pok_nilai_5, pok_nilai_6,
                                    pok_nilai_7, pok_nilai_8, pok_nilai_9,
                                    pok_nilai_10, pok_nilai_11, pok_nilai_12,
                                    kode_full
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (tahun_anggaran,) + data_values)

                        success += 1

                    except Exception as e:
                        errors_count += 1
                        errors.append(f"Baris {row_idx}: {str(e)}")
                        if errors_count >= 10:
                            errors.append("... (error lainnya tidak ditampilkan)")
                            break

                conn.commit()

            return success, errors_count, errors

        except Exception as e:
            return 0, 1, [f"Error membaca file: {str(e)}"]

    def get_all_dipa(self, tahun: int = None, kode_akun: str = None) -> List[Dict]:
        """Get all DIPA records with optional filters."""
        if tahun is None:
            tahun = TAHUN_ANGGARAN

        with self.get_connection() as conn:
            cursor = conn.cursor()

            sql = "SELECT * FROM dipa WHERE tahun_anggaran = ?"
            params = [tahun]

            if kode_akun:
                sql += " AND kode_akun LIKE ?"
                params.append(f"{kode_akun}%")

            sql += " ORDER BY kode_akun, uraian_item"

            cursor.execute(sql, params)
            return [dict(row) for row in cursor.fetchall()]

    def get_dipa(self, dipa_id: int) -> Optional[Dict]:
        """Get single DIPA record by ID."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM dipa WHERE id = ?", (dipa_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_dipa_kode_akun_list(self, tahun: int = None) -> List[Dict]:
        """
        Get list of unique kode_akun from DIPA for dropdown.
        Returns: [{'kode_akun': '521111', 'uraian': 'Belanja Barang...', 'total': 1000000}, ...]
        """
        if tahun is None:
            tahun = TAHUN_ANGGARAN

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT kode_akun,
                       MIN(uraian_item) as uraian_sample,
                       SUM(total) as total_pagu,
                       COUNT(*) as jumlah_item
                FROM dipa
                WHERE tahun_anggaran = ?
                GROUP BY kode_akun
                ORDER BY kode_akun
            """, (tahun,))
            return [dict(row) for row in cursor.fetchall()]

    def get_dipa_by_kode_akun(self, kode_akun: str, tahun: int = None) -> List[Dict]:
        """Get all DIPA items for a specific kode_akun."""
        if tahun is None:
            tahun = TAHUN_ANGGARAN

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM dipa
                WHERE tahun_anggaran = ? AND kode_akun = ?
                ORDER BY uraian_item
            """, (tahun, kode_akun))
            return [dict(row) for row in cursor.fetchall()]

    def search_dipa(self, keyword: str, tahun: int = None) -> List[Dict]:
        """Search DIPA by uraian or kode_akun."""
        if tahun is None:
            tahun = TAHUN_ANGGARAN

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM dipa
                WHERE tahun_anggaran = ?
                      AND (uraian_item LIKE ? OR kode_akun LIKE ?)
                ORDER BY kode_akun, uraian_item
                LIMIT 50
            """, (tahun, f"%{keyword}%", f"%{keyword}%"))
            return [dict(row) for row in cursor.fetchall()]

    def get_dipa_summary(self, tahun: int = None) -> Dict:
        """Get DIPA summary statistics."""
        if tahun is None:
            tahun = TAHUN_ANGGARAN

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    COUNT(*) as total_item,
                    COUNT(DISTINCT kode_akun) as total_akun,
                    SUM(total) as total_pagu,
                    SUM(realisasi) as total_realisasi
                FROM dipa
                WHERE tahun_anggaran = ?
            """, (tahun,))
            row = cursor.fetchone()
            return dict(row) if row else {}

    def delete_all_dipa(self, tahun: int) -> int:
        """Delete all DIPA records for a specific year."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM dipa WHERE tahun_anggaran = ?", (tahun,))
            deleted = cursor.rowcount
            conn.commit()
            return deleted


# ============================================================================
# SINGLETON INSTANCE
# ============================================================================

_db_manager_v4 = None

def get_db_manager_v4() -> DatabaseManagerV4:
    """Get singleton database manager instance"""
    global _db_manager_v4
    if _db_manager_v4 is None:
        _db_manager_v4 = DatabaseManagerV4()
    return _db_manager_v4
