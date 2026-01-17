"""
Script untuk membuat template RAB Swakelola (Excel)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates", "excel")
os.makedirs(TEMPLATES_DIR, exist_ok=True)


def create_rab_swakelola():
    """Create RAB Swakelola Excel template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "RAB Swakelola"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    currency_format = '#,##0'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Header
    ws.merge_cells('A1:G1')
    ws['A1'] = "RENCANA ANGGARAN BIAYA (RAB)"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = "KEGIATAN SWAKELOLA"
    ws['A2'].font = header_font
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:G3')
    ws['A3'] = "{{nama_kegiatan}}"
    ws['A3'].font = subheader_font
    ws['A3'].alignment = Alignment(horizontal='center')

    # Info kegiatan
    ws['A5'] = "Satuan Kerja"
    ws['B5'] = ": {{satker_nama}}"
    ws['A6'] = "Tipe Swakelola"
    ws['B6'] = ": {{tipe_swakelola}}"
    ws['A7'] = "Tahun Anggaran"
    ws['B7'] = ": {{tahun_anggaran}}"
    ws['A8'] = "Sumber Dana"
    ws['B8'] = ": {{sumber_dana}}"
    ws['A9'] = "Kode Akun"
    ws['B9'] = ": {{kode_akun}}"

    # Table header
    headers = ["No", "Uraian Kegiatan", "Volume", "Satuan", "Harga Satuan (Rp)", "Jumlah (Rp)", "Keterangan"]
    row_num = 11

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = header_fill

    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20

    # Sample data rows
    sample_data = [
        ("A.", "HONORARIUM", "", "", "", "", ""),
        ("1", "Honorarium Ketua Tim", "1", "OK", "", "", ""),
        ("2", "Honorarium Sekretaris", "1", "OK", "", "", ""),
        ("3", "Honorarium Anggota", "3", "OK", "", "", ""),
        ("", "Sub Total A", "", "", "", "", ""),
        ("B.", "BELANJA BAHAN", "", "", "", "", ""),
        ("1", "ATK", "1", "Paket", "", "", ""),
        ("2", "Bahan Habis Pakai", "1", "Paket", "", "", ""),
        ("", "Sub Total B", "", "", "", "", ""),
        ("C.", "PERJALANAN DINAS", "", "", "", "", ""),
        ("1", "Transport Lokal", "", "OK", "", "", ""),
        ("2", "Perjalanan Dinas Dalam Kota", "", "OH", "", "", ""),
        ("", "Sub Total C", "", "", "", "", ""),
        ("D.", "KONSUMSI DAN AKOMODASI", "", "", "", "", ""),
        ("1", "Konsumsi Rapat", "", "Paket", "", "", ""),
        ("", "Sub Total D", "", "", "", "", ""),
        ("", "TOTAL (A+B+C+D)", "", "", "", "{{pagu_swakelola}}", ""),
    ]

    for i, row_data in enumerate(sample_data):
        row = row_num + 1 + i
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col in [5, 6]:  # Currency columns
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right')
            elif col == 1:  # No column
                cell.alignment = Alignment(horizontal='center')

    # Terbilang
    total_row = row_num + 1 + len(sample_data) + 1
    ws.merge_cells(f'A{total_row}:B{total_row}')
    ws[f'A{total_row}'] = "Terbilang: {{pagu_swakelola_terbilang}}"
    ws[f'A{total_row}'].font = Font(italic=True, size=11)

    # Signature section
    sign_row = total_row + 3
    ws.merge_cells(f'E{sign_row}:G{sign_row}')
    ws[f'E{sign_row}'] = "{{satker_kota}}, {{tanggal_sk_tim}}"

    sign_row += 1
    ws.merge_cells(f'E{sign_row}:G{sign_row}')
    ws[f'E{sign_row}'] = "{{ppk_jabatan}},"

    sign_row += 4
    ws.merge_cells(f'E{sign_row}:G{sign_row}')
    ws[f'E{sign_row}'] = "{{ppk_nama}}"
    ws[f'E{sign_row}'].font = Font(bold=True, underline='single')

    sign_row += 1
    ws.merge_cells(f'E{sign_row}:G{sign_row}')
    ws[f'E{sign_row}'] = "NIP. {{ppk_nip}}"

    # Save
    filepath = os.path.join(TEMPLATES_DIR, "rab_swakelola.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


if __name__ == "__main__":
    create_rab_swakelola()
    print("\nRAB Swakelola template created successfully!")
