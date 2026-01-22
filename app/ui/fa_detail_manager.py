"""
PPK DOCUMENT FACTORY - FA Detail Manager
=========================================
Manager untuk Pagu Anggaran (POK/DIPA) dengan:
- Import data dari Excel
- View hierarkis pagu anggaran
- Tracking realisasi
- Summary dan rekap per akun
"""

import os
from datetime import datetime, date
from typing import List, Dict, Optional

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QWidget, QPushButton, QLabel, QLineEdit, QTextEdit, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QMessageBox, QCheckBox, QFrame, QTabWidget, QTreeWidget, QTreeWidgetItem,
    QAbstractItemView, QDateEdit, QSpinBox, QDoubleSpinBox, QScrollArea,
    QFileDialog, QProgressBar, QSplitter
)
from PySide6.QtCore import Qt, Signal, QDate, QLocale
from PySide6.QtGui import QColor, QFont, QBrush

from app.core.database_v4 import get_db_manager_v4
from app.core.config import SATKER_DEFAULT, TAHUN_ANGGARAN, OUTPUT_DIR


# ============================================================================
# CUSTOM CURRENCY SPINBOX
# ============================================================================

class CurrencySpinBox(QDoubleSpinBox):
    """Custom SpinBox for Indonesian currency"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setRange(0, 999999999999)
        self.setDecimals(0)
        self.setSingleStep(100000)
        locale = QLocale(QLocale.Indonesian, QLocale.Indonesia)
        self.setLocale(locale)

    def textFromValue(self, value: float) -> str:
        return f"Rp {value:,.0f}".replace(",", ".")

    def valueFromText(self, text: str) -> float:
        clean = text.replace("Rp ", "").replace(".", "").replace(",", "").strip()
        try:
            return float(clean) if clean else 0
        except ValueError:
            return 0


# ============================================================================
# PAGU ANGGARAN DIALOG (Add/Edit)
# ============================================================================

class PaguAnggaranDialog(QDialog):
    """Dialog for adding/editing pagu anggaran"""

    def __init__(self, pagu_data: dict = None, parent=None):
        super().__init__(parent)
        self.pagu_data = pagu_data
        self.db = get_db_manager_v4()

        self.setWindowTitle("Tambah Pagu Anggaran" if not pagu_data else "Edit Pagu Anggaran")
        self.setMinimumWidth(700)
        self.setup_ui()

        if pagu_data:
            self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Kode Group
        kode_group = QGroupBox("Kode Anggaran")
        kode_form = QFormLayout()

        self.spn_tahun = QSpinBox()
        self.spn_tahun.setRange(2020, 2050)
        self.spn_tahun.setValue(TAHUN_ANGGARAN)
        kode_form.addRow("Tahun Anggaran:", self.spn_tahun)

        # Kode fields
        kode_layout = QGridLayout()
        self.txt_kode_program = QLineEdit()
        self.txt_kode_program.setPlaceholderText("054.01.GG")
        kode_layout.addWidget(QLabel("Program:"), 0, 0)
        kode_layout.addWidget(self.txt_kode_program, 0, 1)

        self.txt_kode_kegiatan = QLineEdit()
        self.txt_kode_kegiatan.setPlaceholderText("2896")
        kode_layout.addWidget(QLabel("Kegiatan:"), 0, 2)
        kode_layout.addWidget(self.txt_kode_kegiatan, 0, 3)

        self.txt_kode_kro = QLineEdit()
        self.txt_kode_kro.setPlaceholderText("2896.BMA")
        kode_layout.addWidget(QLabel("KRO:"), 1, 0)
        kode_layout.addWidget(self.txt_kode_kro, 1, 1)

        self.txt_kode_ro = QLineEdit()
        self.txt_kode_ro.setPlaceholderText("2896.BMA.004")
        kode_layout.addWidget(QLabel("RO:"), 1, 2)
        kode_layout.addWidget(self.txt_kode_ro, 1, 3)

        self.txt_kode_komponen = QLineEdit()
        self.txt_kode_komponen.setPlaceholderText("051")
        kode_layout.addWidget(QLabel("Komponen:"), 2, 0)
        kode_layout.addWidget(self.txt_kode_komponen, 2, 1)

        self.txt_kode_sub_komponen = QLineEdit()
        self.txt_kode_sub_komponen.setPlaceholderText("A")
        kode_layout.addWidget(QLabel("Sub Komponen:"), 2, 2)
        kode_layout.addWidget(self.txt_kode_sub_komponen, 2, 3)

        self.txt_kode_akun = QLineEdit()
        self.txt_kode_akun.setPlaceholderText("521211")
        kode_layout.addWidget(QLabel("Akun:"), 3, 0)
        kode_layout.addWidget(self.txt_kode_akun, 3, 1)

        self.txt_kode_detail = QLineEdit()
        kode_layout.addWidget(QLabel("Detail:"), 3, 2)
        kode_layout.addWidget(self.txt_kode_detail, 3, 3)

        kode_form.addRow("", kode_layout)

        self.txt_kode_full = QLineEdit()
        self.txt_kode_full.setReadOnly(True)
        self.txt_kode_full.setStyleSheet("background: #ecf0f1;")
        kode_form.addRow("Kode Lengkap:", self.txt_kode_full)

        kode_group.setLayout(kode_form)
        layout.addWidget(kode_group)

        # Uraian
        uraian_group = QGroupBox("Uraian")
        uraian_form = QFormLayout()

        self.txt_uraian = QTextEdit()
        self.txt_uraian.setMaximumHeight(80)
        uraian_form.addRow("Uraian:", self.txt_uraian)

        uraian_group.setLayout(uraian_form)
        layout.addWidget(uraian_group)

        # Volume dan Harga
        nilai_group = QGroupBox("Volume dan Nilai")
        nilai_form = QFormLayout()

        vol_layout = QHBoxLayout()
        self.spn_volume = QDoubleSpinBox()
        self.spn_volume.setRange(0, 999999)
        self.spn_volume.setDecimals(2)
        vol_layout.addWidget(self.spn_volume)
        self.txt_satuan = QLineEdit()
        self.txt_satuan.setPlaceholderText("Paket/Orang/Bulan")
        vol_layout.addWidget(QLabel("Satuan:"))
        vol_layout.addWidget(self.txt_satuan)
        nilai_form.addRow("Volume:", vol_layout)

        self.spn_harga_satuan = CurrencySpinBox()
        self.spn_harga_satuan.valueChanged.connect(self.calc_jumlah)
        nilai_form.addRow("Harga Satuan:", self.spn_harga_satuan)

        self.spn_volume.valueChanged.connect(self.calc_jumlah)

        self.spn_jumlah = CurrencySpinBox()
        self.spn_jumlah.setStyleSheet("background: #d4edda; font-weight: bold;")
        nilai_form.addRow("Jumlah:", self.spn_jumlah)

        nilai_group.setLayout(nilai_form)
        layout.addWidget(nilai_group)

        # Sumber Dana
        dana_layout = QHBoxLayout()
        self.cmb_sumber_dana = QComboBox()
        self.cmb_sumber_dana.addItems(["RM", "PNBP", "BLU", "PLN", "HLN"])
        dana_layout.addWidget(QLabel("Sumber Dana:"))
        dana_layout.addWidget(self.cmb_sumber_dana)

        self.chk_blokir = QCheckBox("Blokir")
        dana_layout.addWidget(self.chk_blokir)
        dana_layout.addStretch()
        layout.addLayout(dana_layout)

        # Keterangan
        self.txt_keterangan = QLineEdit()
        layout.addWidget(QLabel("Keterangan:"))
        layout.addWidget(self.txt_keterangan)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_cancel = QPushButton("Batal")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)

        btn_save = QPushButton("Simpan")
        btn_save.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 20px;")
        btn_save.clicked.connect(self.save_data)
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

        # Connect kode fields to update kode_full
        for txt in [self.txt_kode_program, self.txt_kode_kegiatan, self.txt_kode_kro,
                    self.txt_kode_ro, self.txt_kode_komponen, self.txt_kode_sub_komponen,
                    self.txt_kode_akun, self.txt_kode_detail]:
            txt.textChanged.connect(self.update_kode_full)

    def update_kode_full(self):
        parts = [
            self.txt_kode_program.text(),
            self.txt_kode_kegiatan.text(),
            self.txt_kode_kro.text(),
            self.txt_kode_ro.text(),
            self.txt_kode_komponen.text(),
            self.txt_kode_sub_komponen.text(),
            self.txt_kode_akun.text(),
            self.txt_kode_detail.text()
        ]
        self.txt_kode_full.setText(".".join([p for p in parts if p]))

    def calc_jumlah(self):
        volume = self.spn_volume.value()
        harga = self.spn_harga_satuan.value()
        self.spn_jumlah.setValue(volume * harga)

    def load_data(self):
        d = self.pagu_data
        self.spn_tahun.setValue(d.get('tahun_anggaran', TAHUN_ANGGARAN))
        self.txt_kode_program.setText(d.get('kode_program', ''))
        self.txt_kode_kegiatan.setText(d.get('kode_kegiatan', ''))
        self.txt_kode_kro.setText(d.get('kode_kro', ''))
        self.txt_kode_ro.setText(d.get('kode_ro', ''))
        self.txt_kode_komponen.setText(d.get('kode_komponen', ''))
        self.txt_kode_sub_komponen.setText(d.get('kode_sub_komponen', ''))
        self.txt_kode_akun.setText(d.get('kode_akun', ''))
        self.txt_kode_detail.setText(d.get('kode_detail', ''))
        self.txt_kode_full.setText(d.get('kode_full', ''))
        self.txt_uraian.setPlainText(d.get('uraian', ''))
        self.spn_volume.setValue(d.get('volume', 0) or 0)
        self.txt_satuan.setText(d.get('satuan', ''))
        self.spn_harga_satuan.setValue(d.get('harga_satuan', 0) or 0)
        self.spn_jumlah.setValue(d.get('jumlah', 0) or 0)
        self.cmb_sumber_dana.setCurrentText(d.get('sumber_dana', 'RM'))
        self.chk_blokir.setChecked(d.get('is_blokir', 0) == 1)
        self.txt_keterangan.setText(d.get('keterangan', ''))

    def save_data(self):
        if not self.txt_uraian.toPlainText().strip():
            QMessageBox.warning(self, "Peringatan", "Uraian harus diisi!")
            return

        data = {
            'tahun_anggaran': self.spn_tahun.value(),
            'kode_program': self.txt_kode_program.text(),
            'kode_kegiatan': self.txt_kode_kegiatan.text(),
            'kode_kro': self.txt_kode_kro.text(),
            'kode_ro': self.txt_kode_ro.text(),
            'kode_komponen': self.txt_kode_komponen.text(),
            'kode_sub_komponen': self.txt_kode_sub_komponen.text(),
            'kode_akun': self.txt_kode_akun.text(),
            'kode_detail': self.txt_kode_detail.text(),
            'kode_full': self.txt_kode_full.text(),
            'level_kode': 8,
            'uraian': self.txt_uraian.toPlainText(),
            'volume': self.spn_volume.value(),
            'satuan': self.txt_satuan.text(),
            'harga_satuan': self.spn_harga_satuan.value(),
            'jumlah': self.spn_jumlah.value(),
            'sumber_dana': self.cmb_sumber_dana.currentText(),
            'is_blokir': 1 if self.chk_blokir.isChecked() else 0,
            'keterangan': self.txt_keterangan.text()
        }

        try:
            if self.pagu_data:
                self.db.update_pagu_anggaran(self.pagu_data['id'], data)
            else:
                self.db.create_pagu_anggaran(data)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan: {str(e)}")


# ============================================================================
# FA DETAIL MANAGER (Main Widget)
# ============================================================================

class FADetailManager(QWidget):
    """Main FA Detail / Pagu Anggaran manager"""

    AKUN_GROUPS = {
        '51': 'Belanja Pegawai',
        '52': 'Belanja Barang',
        '53': 'Belanja Modal',
        '54': 'Belanja Pembayaran Kewajiban Utang',
        '55': 'Belanja Subsidi',
        '56': 'Belanja Hibah',
        '57': 'Belanja Bantuan Sosial',
        '58': 'Belanja Lain-lain'
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = get_db_manager_v4()
        self.setup_ui()
        self.refresh_all()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Header
        header_layout = QHBoxLayout()
        header = QLabel("FA Detail - Pagu Anggaran")
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        header_layout.addWidget(header)

        # Year selector
        header_layout.addStretch()
        header_layout.addWidget(QLabel("Tahun:"))
        self.cmb_tahun = QComboBox()
        for y in range(TAHUN_ANGGARAN - 2, TAHUN_ANGGARAN + 3):
            self.cmb_tahun.addItem(str(y), y)
        self.cmb_tahun.setCurrentText(str(TAHUN_ANGGARAN))
        self.cmb_tahun.currentIndexChanged.connect(self.refresh_all)
        header_layout.addWidget(self.cmb_tahun)

        layout.addLayout(header_layout)

        # Summary cards
        self.summary_widget = QWidget()
        summary_layout = QHBoxLayout(self.summary_widget)
        summary_layout.setContentsMargins(0, 0, 0, 0)

        self.card_total_pagu = self._create_summary_card("Total Pagu", "Rp 0", "#3498db")
        self.card_realisasi = self._create_summary_card("Realisasi", "Rp 0", "#27ae60")
        self.card_sisa = self._create_summary_card("Sisa Anggaran", "Rp 0", "#e67e22")
        self.card_persen = self._create_summary_card("% Realisasi", "0%", "#9b59b6")

        summary_layout.addWidget(self.card_total_pagu)
        summary_layout.addWidget(self.card_realisasi)
        summary_layout.addWidget(self.card_sisa)
        summary_layout.addWidget(self.card_persen)

        layout.addWidget(self.summary_widget)

        # Tabs
        self.tabs = QTabWidget()

        # ========== TAB 1: DAFTAR PAGU ==========
        tab_list = QWidget()
        list_layout = QVBoxLayout(tab_list)

        # Toolbar
        toolbar = QHBoxLayout()

        btn_add = QPushButton("+ Tambah Pagu")
        btn_add.setStyleSheet("background-color: #3498db; color: white; padding: 8px 16px;")
        btn_add.clicked.connect(self.add_pagu)
        toolbar.addWidget(btn_add)

        btn_import = QPushButton("📥 Import Excel")
        btn_import.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 16px;")
        btn_import.clicked.connect(self.import_excel)
        toolbar.addWidget(btn_import)

        btn_export = QPushButton("📤 Export Excel")
        btn_export.setStyleSheet("background-color: #e67e22; color: white; padding: 8px 16px;")
        btn_export.clicked.connect(self.export_excel)
        toolbar.addWidget(btn_export)

        toolbar.addStretch()

        # Filter by akun
        toolbar.addWidget(QLabel("Filter Akun:"))
        self.cmb_filter_akun = QComboBox()
        self.cmb_filter_akun.addItem("Semua", None)
        self.cmb_filter_akun.addItem("51 - Belanja Pegawai", "51")
        self.cmb_filter_akun.addItem("52 - Belanja Barang", "52")
        self.cmb_filter_akun.addItem("53 - Belanja Modal", "53")
        self.cmb_filter_akun.currentIndexChanged.connect(self.refresh_pagu_list)
        toolbar.addWidget(self.cmb_filter_akun)

        # Search
        toolbar.addWidget(QLabel("Cari:"))
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Cari uraian...")
        self.txt_search.textChanged.connect(self.filter_table)
        toolbar.addWidget(self.txt_search)

        btn_refresh = QPushButton("🔄")
        btn_refresh.clicked.connect(self.refresh_all)
        toolbar.addWidget(btn_refresh)

        list_layout.addLayout(toolbar)

        # Table
        self.tbl_pagu = QTableWidget()
        self.tbl_pagu.setColumnCount(11)
        self.tbl_pagu.setHorizontalHeaderLabels([
            "ID", "Nomor MAK", "Kode Akun", "Uraian", "Volume", "Satuan",
            "Harga Satuan", "Jumlah", "Realisasi", "Sisa", "Aksi"
        ])
        self.tbl_pagu.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_pagu.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_pagu.horizontalHeader().setStretchLastSection(True)
        self.tbl_pagu.setColumnWidth(0, 50)   # ID
        self.tbl_pagu.setColumnWidth(1, 100)  # Nomor MAK
        self.tbl_pagu.setColumnWidth(2, 100)  # Kode Akun
        self.tbl_pagu.setColumnWidth(3, 280)  # Uraian
        self.tbl_pagu.setColumnWidth(4, 60)   # Volume
        self.tbl_pagu.setColumnWidth(5, 80)   # Satuan
        self.tbl_pagu.setColumnWidth(6, 120)  # Harga Satuan
        self.tbl_pagu.setColumnWidth(7, 120)  # Jumlah
        self.tbl_pagu.setColumnWidth(8, 120)  # Realisasi
        self.tbl_pagu.setColumnWidth(9, 120)  # Sisa
        self.tbl_pagu.setAlternatingRowColors(True)
        list_layout.addWidget(self.tbl_pagu)

        self.tabs.addTab(tab_list, "Daftar Pagu")

        # ========== TAB 2: REKAP PER AKUN ==========
        tab_rekap = QWidget()
        rekap_layout = QVBoxLayout(tab_rekap)

        self.tbl_rekap = QTableWidget()
        self.tbl_rekap.setColumnCount(6)
        self.tbl_rekap.setHorizontalHeaderLabels([
            "Kode Akun", "Nama Akun", "Jumlah Item", "Total Pagu",
            "Total Realisasi", "Total Sisa"
        ])
        self.tbl_rekap.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_rekap.horizontalHeader().setStretchLastSection(True)
        self.tbl_rekap.setAlternatingRowColors(True)
        rekap_layout.addWidget(self.tbl_rekap)

        self.tabs.addTab(tab_rekap, "Rekap per Akun")

        # ========== TAB 3: PROGRESS BULANAN ==========
        tab_progress = QWidget()
        progress_layout = QVBoxLayout(tab_progress)

        info_label = QLabel("Progress realisasi per bulan akan ditampilkan di sini.")
        info_label.setStyleSheet("color: #7f8c8d; padding: 20px;")
        progress_layout.addWidget(info_label)

        # Monthly progress bars
        self.progress_widget = QWidget()
        progress_grid = QGridLayout(self.progress_widget)

        self.monthly_progress = []
        bulan_names = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun",
                       "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
        for i, bulan in enumerate(bulan_names):
            row = i // 4
            col = i % 4
            prog = QProgressBar()
            prog.setMinimum(0)
            prog.setMaximum(100)
            prog.setValue(0)
            prog.setFormat(f"{bulan}: %p%")
            self.monthly_progress.append(prog)
            progress_grid.addWidget(prog, row, col)

        progress_layout.addWidget(self.progress_widget)
        progress_layout.addStretch()

        self.tabs.addTab(tab_progress, "Progress Bulanan")

        layout.addWidget(self.tabs)

    def _create_summary_card(self, title, value, color):
        """Create a summary card widget"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 8px;
                padding: 10px;
            }}
            QLabel {{
                color: white;
            }}
        """)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(15, 10, 15, 10)

        lbl_title = QLabel(title)
        lbl_title.setStyleSheet("font-size: 12px; color: rgba(255,255,255,0.8);")
        card_layout.addWidget(lbl_title)

        lbl_value = QLabel(value)
        lbl_value.setStyleSheet("font-size: 16px; font-weight: bold;")
        lbl_value.setObjectName("value")
        card_layout.addWidget(lbl_value)

        return card

    def _update_card(self, card, value):
        """Update card value"""
        lbl = card.findChild(QLabel, "value")
        if lbl:
            lbl.setText(value)

    def format_currency(self, value):
        return f"Rp {value:,.0f}".replace(",", ".")

    def refresh_all(self):
        self.refresh_summary()
        self.refresh_pagu_list()
        self.refresh_rekap()

    def refresh_summary(self):
        tahun = self.cmb_tahun.currentData()
        summary = self.db.get_pagu_summary(tahun)

        total_pagu = summary.get('total_pagu', 0) or 0
        total_real = summary.get('total_realisasi', 0) or 0
        total_sisa = summary.get('total_sisa', 0) or 0
        persen = summary.get('persen_realisasi', 0) or 0

        self._update_card(self.card_total_pagu, self.format_currency(total_pagu))
        self._update_card(self.card_realisasi, self.format_currency(total_real))
        self._update_card(self.card_sisa, self.format_currency(total_sisa))
        self._update_card(self.card_persen, f"{persen:.1f}%")

    def refresh_pagu_list(self):
        tahun = self.cmb_tahun.currentData()
        kode_akun = self.cmb_filter_akun.currentData()

        data = self.db.get_all_pagu_anggaran(tahun=tahun, kode_akun=kode_akun)
        self.tbl_pagu.setRowCount(len(data))

        for row, d in enumerate(data):
            self.tbl_pagu.setItem(row, 0, QTableWidgetItem(str(d['id'])))

            # Nomor MAK - dari field nomor_mak atau generate dari kode_akun.kode_detail
            nomor_mak = d.get('nomor_mak', '')
            if not nomor_mak:
                kode_akun_val = d.get('kode_akun', '')
                kode_detail_val = d.get('kode_detail', '')
                if kode_akun_val and kode_detail_val:
                    nomor_mak = f"{kode_akun_val}.{kode_detail_val}"
            self.tbl_pagu.setItem(row, 1, QTableWidgetItem(nomor_mak))

            self.tbl_pagu.setItem(row, 2, QTableWidgetItem(d.get('kode_akun', '')))

            # Uraian (truncate if too long)
            uraian = d.get('uraian', '')[:100]
            self.tbl_pagu.setItem(row, 3, QTableWidgetItem(uraian))

            self.tbl_pagu.setItem(row, 4, QTableWidgetItem(str(d.get('volume', 0) or 0)))
            self.tbl_pagu.setItem(row, 5, QTableWidgetItem(d.get('satuan', '')))
            self.tbl_pagu.setItem(row, 6, QTableWidgetItem(self.format_currency(d.get('harga_satuan', 0) or 0)))
            self.tbl_pagu.setItem(row, 7, QTableWidgetItem(self.format_currency(d.get('jumlah', 0) or 0)))
            self.tbl_pagu.setItem(row, 8, QTableWidgetItem(self.format_currency(d.get('realisasi', 0) or 0)))

            # Sisa with color
            sisa = d.get('sisa', 0) or 0
            sisa_item = QTableWidgetItem(self.format_currency(sisa))
            if sisa < 0:
                sisa_item.setBackground(QColor('#f8d7da'))
            elif sisa == 0:
                sisa_item.setBackground(QColor('#d4edda'))
            self.tbl_pagu.setItem(row, 9, sisa_item)

            # Actions
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)

            btn_edit = QPushButton("Edit")
            btn_edit.setStyleSheet("background: #3498db; color: white; padding: 2px 8px;")
            btn_edit.clicked.connect(lambda checked, r=row: self.edit_pagu(r))
            btn_layout.addWidget(btn_edit)

            btn_del = QPushButton("Hapus")
            btn_del.setStyleSheet("background: #e74c3c; color: white; padding: 2px 8px;")
            btn_del.clicked.connect(lambda checked, r=row: self.delete_pagu(r))
            btn_layout.addWidget(btn_del)

            self.tbl_pagu.setCellWidget(row, 10, btn_widget)

    def refresh_rekap(self):
        tahun = self.cmb_tahun.currentData()
        data = self.db.get_pagu_by_akun_group(tahun)

        self.tbl_rekap.setRowCount(len(data))

        for row, d in enumerate(data):
            grup = d.get('grup_akun', '')
            nama = self.AKUN_GROUPS.get(grup, 'Lainnya')

            self.tbl_rekap.setItem(row, 0, QTableWidgetItem(grup))
            self.tbl_rekap.setItem(row, 1, QTableWidgetItem(nama))
            self.tbl_rekap.setItem(row, 2, QTableWidgetItem(str(d.get('jumlah_item', 0))))
            self.tbl_rekap.setItem(row, 3, QTableWidgetItem(self.format_currency(d.get('total_pagu', 0) or 0)))
            self.tbl_rekap.setItem(row, 4, QTableWidgetItem(self.format_currency(d.get('total_realisasi', 0) or 0)))
            self.tbl_rekap.setItem(row, 5, QTableWidgetItem(self.format_currency(d.get('total_sisa', 0) or 0)))

    def filter_table(self):
        search = self.txt_search.text().lower()
        for row in range(self.tbl_pagu.rowCount()):
            item = self.tbl_pagu.item(row, 3)  # Uraian column (now at index 3)
            if item:
                self.tbl_pagu.setRowHidden(row, search not in item.text().lower())

    def add_pagu(self):
        dialog = PaguAnggaranDialog(parent=self)
        if dialog.exec() == QDialog.Accepted:
            self.refresh_all()
            QMessageBox.information(self, "Sukses", "Pagu anggaran berhasil ditambahkan!")

    def edit_pagu(self, row):
        pagu_id = int(self.tbl_pagu.item(row, 0).text())
        pagu_data = self.db.get_pagu_anggaran(pagu_id)
        if pagu_data:
            dialog = PaguAnggaranDialog(pagu_data, parent=self)
            if dialog.exec() == QDialog.Accepted:
                self.refresh_all()

    def delete_pagu(self, row):
        pagu_id = int(self.tbl_pagu.item(row, 0).text())
        uraian = self.tbl_pagu.item(row, 2).text()

        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Yakin ingin menghapus pagu:\n{uraian[:50]}...?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.db.delete_pagu_anggaran(pagu_id):
                self.refresh_all()
                QMessageBox.information(self, "Sukses", "Pagu berhasil dihapus!")

    def import_excel(self):
        """Import pagu from Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Pilih File Excel",
            "", "Excel Files (*.xlsx *.xls)"
        )

        if not file_path:
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            ws = wb.active

            tahun = self.cmb_tahun.currentData()

            # Ask import mode - with update option
            msg = QMessageBox(self)
            msg.setWindowTitle("Konfirmasi Import")
            msg.setText(f"Pilih mode import data pagu tahun {tahun}:")
            msg.setInformativeText(
                "• Hapus & Import Baru = Hapus semua data lama, import data baru\n"
                "• Update = Update data yang ada, tambah data baru (RECOMMENDED)\n"
                "• Tambah Saja = Tambahkan tanpa update data yang ada"
            )
            btn_replace = msg.addButton("Hapus && Import Baru", QMessageBox.DestructiveRole)
            btn_update = msg.addButton("Update", QMessageBox.AcceptRole)
            btn_append = msg.addButton("Tambah Saja", QMessageBox.ActionRole)
            btn_cancel = msg.addButton("Batal", QMessageBox.RejectRole)
            msg.setDefaultButton(btn_update)
            msg.exec()

            clicked = msg.clickedButton()
            if clicked == btn_cancel:
                return

            use_upsert = (clicked == btn_update)

            if clicked == btn_replace:
                deleted = self.db.delete_all_pagu_tahun(tahun)
                QMessageBox.information(self, "Info", f"Deleted {deleted} existing records")

            # Parse Excel
            data_list = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue

                # Try to parse columns - adjust based on your Excel format
                # Expected: Kode, Uraian, Volume, Satuan, Harga Satuan, Jumlah, [Realisasi]
                try:
                    kode = str(row[0]).strip() if row[0] else ''
                    uraian = str(row[1]).strip() if len(row) > 1 and row[1] else ''

                    # Parse kode hierarchically
                    # Format: 054.01.WA.4621.QEB.001.001.A.521211.1
                    # Parts: Program.Kegiatan.KRO.RO.Komponen.SubKomponen.Akun.Detail
                    kode_parts = kode.split('.')
                    kode_akun = ''
                    kode_detail = ''
                    kode_program = ''
                    kode_kegiatan = ''
                    kode_kro = ''
                    kode_ro = ''
                    kode_komponen = ''
                    kode_sub_komponen = ''

                    # Find kode_akun (6-digit number) and kode_detail (after akun)
                    akun_idx = -1
                    for idx, part in enumerate(kode_parts):
                        if part.isdigit() and len(part) == 6:
                            kode_akun = part
                            akun_idx = idx
                            break

                    # Get detail code (after akun)
                    if akun_idx >= 0 and akun_idx < len(kode_parts) - 1:
                        kode_detail = kode_parts[akun_idx + 1]

                    # Parse other hierarchy codes if available
                    if len(kode_parts) >= 2:
                        kode_program = '.'.join(kode_parts[:2]) if len(kode_parts) > 1 else kode_parts[0]
                    if len(kode_parts) >= 3:
                        kode_kegiatan = kode_parts[2] if len(kode_parts) > 2 else ''
                    if len(kode_parts) >= 4:
                        kode_kro = kode_parts[3] if len(kode_parts) > 3 else ''
                    if len(kode_parts) >= 5:
                        kode_ro = kode_parts[4] if len(kode_parts) > 4 else ''

                    volume = float(row[2]) if len(row) > 2 and row[2] else 0
                    satuan = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    harga_satuan = float(row[4]) if len(row) > 4 and row[4] else 0
                    jumlah = float(row[5]) if len(row) > 5 and row[5] else volume * harga_satuan

                    # Parse realisasi if available (column 7 or 8)
                    realisasi = 0
                    if len(row) > 6 and row[6]:
                        try:
                            realisasi = float(row[6])
                        except (ValueError, TypeError):
                            pass

                    # Determine level_kode based on kode structure
                    level_kode = 8  # Default to detail level
                    if kode_detail:
                        level_kode = 8  # Detail level
                    elif kode_akun:
                        level_kode = 7  # Akun level

                    if uraian:  # Only add if has uraian
                        data_list.append({
                            'tahun_anggaran': tahun,
                            'kode_full': kode,
                            'kode_program': kode_program,
                            'kode_kegiatan': kode_kegiatan,
                            'kode_kro': kode_kro,
                            'kode_ro': kode_ro,
                            'kode_komponen': kode_komponen,
                            'kode_sub_komponen': kode_sub_komponen,
                            'kode_akun': kode_akun,
                            'kode_detail': kode_detail,
                            'uraian': uraian,
                            'volume': volume,
                            'satuan': satuan,
                            'harga_satuan': harga_satuan,
                            'jumlah': jumlah,
                            'realisasi': realisasi,
                            'level_kode': level_kode,
                            'sumber_dana': 'RM'
                        })
                except Exception as e:
                    print(f"Error parsing row {row_idx}: {e}")
                    continue

            if data_list:
                count = self.db.bulk_insert_pagu(data_list, upsert=use_upsert)
                self.refresh_all()
                mode_text = "diupdate/ditambah" if use_upsert else "ditambahkan"
                QMessageBox.information(
                    self, "Sukses",
                    f"Berhasil {mode_text} {count} item pagu anggaran!"
                )
            else:
                QMessageBox.warning(self, "Peringatan", "Tidak ada data yang bisa diimport!")

        except ImportError:
            QMessageBox.warning(
                self, "Error",
                "Module openpyxl tidak tersedia.\n"
                "Install dengan: pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal import: {str(e)}")

    def export_excel(self):
        """Export pagu to Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Simpan File Excel",
            f"pagu_anggaran_{self.cmb_tahun.currentData()}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            tahun = self.cmb_tahun.currentData()
            data = self.db.get_all_pagu_anggaran(tahun=tahun)

            wb = Workbook()
            ws = wb.active
            ws.title = f"Pagu {tahun}"

            # Headers
            headers = ["Kode", "Uraian", "Volume", "Satuan", "Harga Satuan",
                       "Jumlah", "Realisasi", "Sisa", "Sumber Dana"]
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Data
            for row_idx, d in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=d.get('kode_full', ''))
                ws.cell(row=row_idx, column=2, value=d.get('uraian', ''))
                ws.cell(row=row_idx, column=3, value=d.get('volume', 0))
                ws.cell(row=row_idx, column=4, value=d.get('satuan', ''))
                ws.cell(row=row_idx, column=5, value=d.get('harga_satuan', 0))
                ws.cell(row=row_idx, column=6, value=d.get('jumlah', 0))
                ws.cell(row=row_idx, column=7, value=d.get('realisasi', 0))
                ws.cell(row=row_idx, column=8, value=d.get('sisa', 0))
                ws.cell(row=row_idx, column=9, value=d.get('sumber_dana', ''))

            # Auto-width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(
                self, "Sukses",
                f"Berhasil export {len(data)} item ke:\n{file_path}"
            )

            # Open file
            import sys
            if sys.platform == 'win32':
                os.startfile(file_path)
            elif sys.platform == 'darwin':
                os.system(f'open "{file_path}"')
            else:
                os.system(f'xdg-open "{file_path}"')

        except ImportError:
            QMessageBox.warning(
                self, "Error",
                "Module openpyxl tidak tersedia.\n"
                "Install dengan: pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export: {str(e)}")
