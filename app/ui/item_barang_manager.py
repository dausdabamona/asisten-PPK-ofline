"""
PPK DOCUMENT FACTORY v3.0 - Item Barang Manager
================================================
UI untuk mengelola Daftar Kebutuhan Barang / Bill of Quantity
Dengan fitur Upload Excel
"""

import os
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QWidget, QPushButton, QLabel, QLineEdit, QTextEdit, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QMessageBox, QSpinBox, QDoubleSpinBox, QFrame, QSplitter,
    QAbstractItemView, QMenu, QFileDialog, QProgressDialog, QApplication
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QAction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.database import get_db_manager, KATEGORI_ITEM, KELOMPOK_ITEM


def format_rupiah(value):
    """Format number as Rupiah"""
    if value is None:
        return "Rp 0"
    return f"Rp {value:,.0f}".replace(",", ".")


class ExcelTemplateGenerator:
    """Generate Excel template for item upload"""
    
    @staticmethod
    def create_upload_template(filepath: str, paket_nama: str = ""):
        """Create Excel template for uploading items"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Daftar Item"
        
        # Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        locked = Protection(locked=True)
        unlocked = Protection(locked=False)
        
        # Title
        ws.merge_cells('A1:L1')
        ws['A1'] = 'TEMPLATE UPLOAD DAFTAR KEBUTUHAN BARANG'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center
        
        ws.merge_cells('A2:L2')
        ws['A2'] = f'Paket: {paket_nama}' if paket_nama else 'Paket: [Nama Paket]'
        ws['A2'].alignment = center
        
        # Instructions
        ws.merge_cells('A4:L4')
        ws['A4'] = '📌 PETUNJUK: Isi data mulai baris 8. Kolom dengan * wajib diisi. Jangan ubah header.'
        ws['A4'].font = Font(italic=True, color="C00000")
        
        # Headers (row 6)
        headers = [
            ('No', 5),
            ('Kategori *', 12),
            ('Kelompok', 15),
            ('Uraian Barang/Jasa *', 35),
            ('Spesifikasi', 30),
            ('Satuan *', 10),
            ('Volume *', 12),
            ('Harga Survey 1', 15),
            ('Harga Survey 2', 15),
            ('Harga Survey 3', 15),
            ('Harga Satuan *', 15),
            ('Keterangan', 20)
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
            cell.protection = locked
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Sub-header with codes (row 7)
        sub_headers = ['', 'A/B/C/D', 'Pilih', '', '', '', 'Angka', 'Angka', 'Angka', 'Angka', 'Angka', '']
        for col, sub in enumerate(sub_headers, 1):
            cell = ws.cell(row=7, column=col, value=sub)
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = center
            cell.border = border
        
        # Data validation for Kategori
        kategori_list = '"A - Bahan/Material,B - Peralatan,C - Jasa/Tenaga,D - Lain-lain"'
        dv_kategori = DataValidation(type="list", formula1=kategori_list, allow_blank=True)
        dv_kategori.error = 'Pilih kategori dari daftar'
        dv_kategori.errorTitle = 'Kategori tidak valid'
        ws.add_data_validation(dv_kategori)
        dv_kategori.add('B8:B1000')
        
        # Data validation for Kelompok
        kelompok_list = '"{}"'.format(','.join(KELOMPOK_ITEM))
        dv_kelompok = DataValidation(type="list", formula1=kelompok_list, allow_blank=True)
        dv_kelompok.error = 'Pilih kelompok dari daftar'
        dv_kelompok.errorTitle = 'Kelompok tidak valid'
        ws.add_data_validation(dv_kelompok)
        dv_kelompok.add('C8:C1000')
        
        # Data validation for Satuan
        satuan_list = '"Unit,Set,Buah,Paket,Kg,Liter,Meter,M2,M3,Lembar,Rim,Dus,Ekor,Orang,OH,OB,OJ,LS"'
        dv_satuan = DataValidation(type="list", formula1=satuan_list, allow_blank=True)
        ws.add_data_validation(dv_satuan)
        dv_satuan.add('F8:F1000')
        
        # Sample data rows (row 8-10)
        sample_data = [
            [1, 'A - Bahan/Material', 'Bahan Habis Pakai', 'Pakan Ikan Floating', 'Protein 30%, ukuran 3mm', 'Kg', 100, 25000, 26000, 24500, 25000, 'Untuk budidaya'],
            [2, 'B - Peralatan', 'Peralatan Utama', 'Jaring Ikan', 'PE, mesh 2 inch, 10x5m', 'Lembar', 5, 350000, 375000, 360000, 360000, ''],
            [3, 'C - Jasa/Tenaga', 'Jasa Teknis', 'Tenaga Ahli Perikanan', 'S1 Perikanan, pengalaman 3 tahun', 'OB', 2, 8000000, 8500000, 8000000, 8000000, 'Full time'],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 8):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.protection = unlocked
                if col_idx in [7, 8, 9, 10, 11]:  # Number columns
                    cell.number_format = '#,##0'
        
        # Add empty rows with borders (row 11 to 510 = 500 rows)
        for row_idx in range(11, 511):
            ws.cell(row=row_idx, column=1, value=row_idx - 7)  # Auto number
            for col_idx in range(1, 13):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.protection = unlocked
        
        # Legend sheet
        ws_legend = wb.create_sheet("Keterangan")
        
        ws_legend['A1'] = 'KETERANGAN PENGISIAN'
        ws_legend['A1'].font = Font(bold=True, size=14)
        
        ws_legend['A3'] = 'KATEGORI:'
        ws_legend['A3'].font = Font(bold=True)
        for i, (code, name) in enumerate(KATEGORI_ITEM.items(), 4):
            ws_legend[f'A{i}'] = f'{code} = {name}'
        
        ws_legend['A9'] = 'KELOMPOK:'
        ws_legend['A9'].font = Font(bold=True)
        for i, kelompok in enumerate(KELOMPOK_ITEM, 10):
            ws_legend[f'A{i}'] = kelompok
        
        ws_legend['A22'] = 'SATUAN UMUM:'
        ws_legend['A22'].font = Font(bold=True)
        satuan_list = ['Unit', 'Set', 'Buah', 'Paket', 'Kg', 'Liter', 'Meter', 'M2', 'M3', 
                       'Lembar', 'Rim', 'Dus', 'Ekor', 'Orang', 'OH (Orang Hari)', 
                       'OB (Orang Bulan)', 'OJ (Orang Jam)', 'LS (Lump Sum)']
        for i, satuan in enumerate(satuan_list, 23):
            ws_legend[f'A{i}'] = satuan
        
        ws_legend.column_dimensions['A'].width = 40
        
        # Protect sheet but allow data entry
        ws.protection.sheet = True
        ws.protection.password = 'ppk123'
        ws.protection.enable()
        
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def parse_upload_file(filepath: str) -> tuple:
        """
        Parse uploaded Excel file
        Returns: (items: list, errors: list)
        """
        items = []
        errors = []
        
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Find header row (looking for "Uraian" in column D)
            header_row = None
            for row in range(1, 10):
                cell_value = ws.cell(row=row, column=4).value
                if cell_value and 'Uraian' in str(cell_value):
                    header_row = row
                    break
            
            if not header_row:
                errors.append("Format file tidak valid. Header tidak ditemukan.")
                return items, errors
            
            # Read data starting from header_row + 2 (skip sub-header)
            data_start = header_row + 2
            
            for row_idx in range(data_start, ws.max_row + 1):
                # Skip empty rows
                uraian = ws.cell(row=row_idx, column=4).value
                if not uraian or str(uraian).strip() == '':
                    continue
                
                try:
                    # Parse kategori (extract code from "A - Bahan/Material")
                    kategori_raw = ws.cell(row=row_idx, column=2).value or ''
                    kategori = str(kategori_raw).split(' - ')[0].strip().upper() if kategori_raw else ''
                    if kategori and kategori not in KATEGORI_ITEM:
                        kategori = ''
                    
                    # Parse other fields
                    kelompok = ws.cell(row=row_idx, column=3).value or ''
                    spesifikasi = ws.cell(row=row_idx, column=5).value or ''
                    satuan = ws.cell(row=row_idx, column=6).value or 'Unit'
                    
                    # Parse numbers
                    def parse_number(val):
                        if val is None:
                            return 0
                        if isinstance(val, (int, float)):
                            return float(val)
                        # Remove formatting
                        val_str = str(val).replace('.', '').replace(',', '.').replace('Rp', '').strip()
                        try:
                            return float(val_str) if val_str else 0
                        except:
                            return 0
                    
                    volume = parse_number(ws.cell(row=row_idx, column=7).value)
                    harga_survey1 = parse_number(ws.cell(row=row_idx, column=8).value)
                    harga_survey2 = parse_number(ws.cell(row=row_idx, column=9).value)
                    harga_survey3 = parse_number(ws.cell(row=row_idx, column=10).value)
                    harga_dasar = parse_number(ws.cell(row=row_idx, column=11).value)
                    keterangan = ws.cell(row=row_idx, column=12).value or ''
                    
                    # Validation
                    row_errors = []
                    if not uraian:
                        row_errors.append("Uraian wajib diisi")
                    if volume <= 0:
                        row_errors.append("Volume harus > 0")
                    if harga_dasar <= 0:
                        # Try to use average of survey prices
                        survey_prices = [p for p in [harga_survey1, harga_survey2, harga_survey3] if p > 0]
                        if survey_prices:
                            harga_dasar = sum(survey_prices) / len(survey_prices)
                        else:
                            row_errors.append("Harga Satuan harus > 0")
                    
                    if row_errors:
                        errors.append(f"Baris {row_idx}: {', '.join(row_errors)}")
                        continue
                    
                    items.append({
                        'kategori': kategori,
                        'kelompok': str(kelompok).strip(),
                        'uraian': str(uraian).strip(),
                        'spesifikasi': str(spesifikasi).strip(),
                        'satuan': str(satuan).strip(),
                        'volume': volume,
                        'harga_survey1': harga_survey1 if harga_survey1 > 0 else None,
                        'harga_survey2': harga_survey2 if harga_survey2 > 0 else None,
                        'harga_survey3': harga_survey3 if harga_survey3 > 0 else None,
                        'harga_dasar': harga_dasar,
                        'keterangan': str(keterangan).strip()
                    })
                    
                except Exception as e:
                    errors.append(f"Baris {row_idx}: Error parsing - {str(e)}")
            
            if not items and not errors:
                errors.append("Tidak ada data yang dapat diimport")
                
        except Exception as e:
            errors.append(f"Error membaca file: {str(e)}")
        
        return items, errors


class ItemBarangDialog(QDialog):
    """Dialog untuk add/edit single item"""
    
    def __init__(self, paket_id: int, item_data: dict = None, parent=None):
        super().__init__(parent)
        self.paket_id = paket_id
        self.item_data = item_data  # None for new item
        self.db = get_db_manager()
        
        # Get paket metode HPS setting
        paket = self.db.get_paket(paket_id)
        self.metode_hps = paket.get('metode_hps', 'RATA') if paket else 'RATA'
        
        self.setWindowTitle("Tambah Item" if not item_data else "Edit Item")
        self.setMinimumWidth(600)
        self.setup_ui()
        
        if item_data:
            self.load_data()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Form
        form_layout = QFormLayout()
        
        # Kategori
        self.cmb_kategori = QComboBox()
        for code, name in KATEGORI_ITEM.items():
            self.cmb_kategori.addItem(f"{code} - {name}", code)
        form_layout.addRow("Kategori:", self.cmb_kategori)
        
        # Kelompok
        self.cmb_kelompok = QComboBox()
        self.cmb_kelompok.addItems(KELOMPOK_ITEM)
        form_layout.addRow("Kelompok:", self.cmb_kelompok)
        
        # Uraian
        self.txt_uraian = QLineEdit()
        self.txt_uraian.setPlaceholderText("Nama barang/jasa...")
        form_layout.addRow("Uraian *:", self.txt_uraian)
        
        # Spesifikasi
        self.txt_spesifikasi = QTextEdit()
        self.txt_spesifikasi.setMaximumHeight(80)
        self.txt_spesifikasi.setPlaceholderText("Merk, tipe, ukuran, warna, dll...")
        form_layout.addRow("Spesifikasi:", self.txt_spesifikasi)
        
        # Satuan
        self.txt_satuan = QComboBox()
        self.txt_satuan.setEditable(True)
        self.txt_satuan.addItems([
            "Unit", "Set", "Buah", "Paket", "Kg", "Liter", "Meter", 
            "M2", "M3", "Lembar", "Rim", "Dus", "Ekor", "Orang", 
            "OH", "OB", "OJ", "LS"
        ])
        form_layout.addRow("Satuan:", self.txt_satuan)
        
        # Volume
        self.spn_volume = QDoubleSpinBox()
        self.spn_volume.setRange(0, 999999999)
        self.spn_volume.setDecimals(2)
        self.spn_volume.valueChanged.connect(self.calculate_total)
        form_layout.addRow("Volume:", self.spn_volume)
        
        layout.addLayout(form_layout)
        
        # Harga Section
        harga_group = QGroupBox("Harga Survey & HPS")
        harga_layout = QGridLayout()
        
        # Harga Survey
        harga_layout.addWidget(QLabel("Survey Harga:"), 0, 0, 1, 2)
        
        # Metode HPS
        harga_layout.addWidget(QLabel("Metode HPS:"), 0, 2)
        self.cmb_metode = QComboBox()
        self.cmb_metode.addItem("📊 Rata-rata", "RATA")
        self.cmb_metode.addItem("📈 Harga Tertinggi", "TERTINGGI")
        self.cmb_metode.addItem("📉 Harga Terendah", "TERENDAH")
        
        # Set current metode from paket setting
        idx = self.cmb_metode.findData(self.metode_hps)
        if idx >= 0:
            self.cmb_metode.setCurrentIndex(idx)
        self.cmb_metode.currentIndexChanged.connect(self.on_metode_changed)
        harga_layout.addWidget(self.cmb_metode, 0, 3)
        
        harga_layout.addWidget(QLabel("Toko 1:"), 1, 0)
        self.spn_harga1 = QDoubleSpinBox()
        self.spn_harga1.setRange(0, 999999999999)
        self.spn_harga1.setPrefix("Rp ")
        self.spn_harga1.valueChanged.connect(self.calculate_hps)
        harga_layout.addWidget(self.spn_harga1, 1, 1)
        
        harga_layout.addWidget(QLabel("Toko 2:"), 1, 2)
        self.spn_harga2 = QDoubleSpinBox()
        self.spn_harga2.setRange(0, 999999999999)
        self.spn_harga2.setPrefix("Rp ")
        self.spn_harga2.valueChanged.connect(self.calculate_hps)
        harga_layout.addWidget(self.spn_harga2, 1, 3)
        
        harga_layout.addWidget(QLabel("Toko 3:"), 2, 0)
        self.spn_harga3 = QDoubleSpinBox()
        self.spn_harga3.setRange(0, 999999999999)
        self.spn_harga3.setPrefix("Rp ")
        self.spn_harga3.valueChanged.connect(self.calculate_hps)
        harga_layout.addWidget(self.spn_harga3, 2, 1)
        
        # HPS Result
        harga_layout.addWidget(QLabel("Harga HPS:"), 2, 2)
        self.lbl_harga_hps = QLabel("Rp 0")
        self.lbl_harga_hps.setStyleSheet("font-weight: bold; color: #2980b9;")
        harga_layout.addWidget(self.lbl_harga_hps, 2, 3)
        
        # Apply HPS button
        btn_apply_hps = QPushButton("⬇️ Gunakan Harga HPS")
        btn_apply_hps.setToolTip("Terapkan harga HPS ke Harga Satuan")
        btn_apply_hps.clicked.connect(self.apply_hps_price)
        harga_layout.addWidget(btn_apply_hps, 3, 0, 1, 4)
        
        # Separator
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        harga_layout.addWidget(line, 4, 0, 1, 4)
        
        # Harga Dasar
        harga_layout.addWidget(QLabel("Harga Satuan:"), 5, 0)
        self.spn_harga_dasar = QDoubleSpinBox()
        self.spn_harga_dasar.setRange(0, 999999999999)
        self.spn_harga_dasar.setPrefix("Rp ")
        self.spn_harga_dasar.valueChanged.connect(self.calculate_total)
        harga_layout.addWidget(self.spn_harga_dasar, 5, 1)
        
        harga_layout.addWidget(QLabel("Total:"), 5, 2)
        self.lbl_total = QLabel("Rp 0")
        self.lbl_total.setStyleSheet("font-weight: bold; font-size: 14px; color: #27ae60;")
        harga_layout.addWidget(self.lbl_total, 5, 3)
        
        harga_group.setLayout(harga_layout)
        layout.addWidget(harga_group)
        
        # Keterangan
        form2 = QFormLayout()
        self.txt_keterangan = QLineEdit()
        self.txt_keterangan.setPlaceholderText("Catatan tambahan...")
        form2.addRow("Keterangan:", self.txt_keterangan)
        layout.addLayout(form2)
        
        layout.addStretch()
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_save = QPushButton("💾 Simpan")
        btn_save.setObjectName("btnSuccess")
        btn_save.clicked.connect(self.save)
        btn_layout.addWidget(btn_save)
        
        btn_cancel = QPushButton("Batal")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)
        
        layout.addLayout(btn_layout)
    
    def load_data(self):
        """Load existing item data"""
        if not self.item_data:
            return
        
        # Kategori
        idx = self.cmb_kategori.findData(self.item_data.get('kategori'))
        if idx >= 0:
            self.cmb_kategori.setCurrentIndex(idx)
        
        # Kelompok
        idx = self.cmb_kelompok.findText(self.item_data.get('kelompok', ''))
        if idx >= 0:
            self.cmb_kelompok.setCurrentIndex(idx)
        
        self.txt_uraian.setText(self.item_data.get('uraian', ''))
        self.txt_spesifikasi.setPlainText(self.item_data.get('spesifikasi', ''))
        self.txt_satuan.setCurrentText(self.item_data.get('satuan', ''))
        self.spn_volume.setValue(self.item_data.get('volume', 0) or 0)
        self.spn_harga1.setValue(self.item_data.get('harga_survey1', 0) or 0)
        self.spn_harga2.setValue(self.item_data.get('harga_survey2', 0) or 0)
        self.spn_harga3.setValue(self.item_data.get('harga_survey3', 0) or 0)
        self.spn_harga_dasar.setValue(self.item_data.get('harga_dasar', 0) or 0)
        self.txt_keterangan.setText(self.item_data.get('keterangan', ''))
        
        self.calculate_hps()
        self.calculate_total()
    
    def on_metode_changed(self):
        """Handle metode HPS change"""
        self.calculate_hps()
    
    def calculate_hps(self):
        """Calculate HPS price based on selected method"""
        from app.core.config import hitung_harga_hps
        
        survey1 = self.spn_harga1.value()
        survey2 = self.spn_harga2.value()
        survey3 = self.spn_harga3.value()
        metode = self.cmb_metode.currentData()
        
        harga_hps = hitung_harga_hps(survey1, survey2, survey3, metode)
        self.lbl_harga_hps.setText(format_rupiah(harga_hps))
        
        # Store for later use
        self._harga_hps = harga_hps
        
        # Auto-fill harga dasar if empty
        if self.spn_harga_dasar.value() == 0 and harga_hps > 0:
            self.spn_harga_dasar.setValue(harga_hps)
    
    def apply_hps_price(self):
        """Apply calculated HPS price to harga dasar"""
        if hasattr(self, '_harga_hps') and self._harga_hps > 0:
            self.spn_harga_dasar.setValue(self._harga_hps)
    
    def calculate_total(self):
        """Calculate total"""
        total = self.spn_volume.value() * self.spn_harga_dasar.value()
        self.lbl_total.setText(format_rupiah(total))
    
    def get_data(self) -> dict:
        """Get form data"""
        return {
            'kategori': self.cmb_kategori.currentData(),
            'kelompok': self.cmb_kelompok.currentText(),
            'uraian': self.txt_uraian.text().strip(),
            'spesifikasi': self.txt_spesifikasi.toPlainText().strip(),
            'satuan': self.txt_satuan.currentText(),
            'volume': self.spn_volume.value(),
            'harga_survey1': self.spn_harga1.value() or None,
            'harga_survey2': self.spn_harga2.value() or None,
            'harga_survey3': self.spn_harga3.value() or None,
            'harga_dasar': self.spn_harga_dasar.value(),
            'keterangan': self.txt_keterangan.text().strip()
        }
    
    def save(self):
        """Validate and save"""
        data = self.get_data()
        
        if not data['uraian']:
            QMessageBox.warning(self, "Validasi", "Uraian barang/jasa wajib diisi!")
            self.txt_uraian.setFocus()
            return
        
        if data['volume'] <= 0:
            QMessageBox.warning(self, "Validasi", "Volume harus lebih dari 0!")
            self.spn_volume.setFocus()
            return
        
        if data['harga_dasar'] <= 0:
            QMessageBox.warning(self, "Validasi", "Harga satuan harus lebih dari 0!")
            self.spn_harga_dasar.setFocus()
            return
        
        try:
            if self.item_data:
                # Update
                self.db.update_item_barang(self.item_data['id'], data)
            else:
                # Insert
                self.db.add_item_barang(self.paket_id, data)
            
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan: {str(e)}")


class ItemBarangManager(QDialog):
    """Main dialog untuk mengelola semua item barang dalam paket"""
    
    items_changed = Signal()
    
    def __init__(self, paket_id: int, parent=None):
        super().__init__(parent)
        self.paket_id = paket_id
        self.db = get_db_manager()
        
        # Get paket info
        self.paket = self.db.get_paket(paket_id)
        
        self.setWindowTitle(f"Daftar Kebutuhan Barang - {self.paket.get('nama', '')}")
        self.setMinimumSize(1100, 700)
        
        self.setup_ui()
        self.load_items()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Header
        header = QHBoxLayout()
        
        title = QLabel(f"📦 Daftar Kebutuhan Barang / Bill of Quantity")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        header.addWidget(title)
        
        header.addStretch()
        
        # Download template button
        btn_download = QPushButton("📥 Download Template")
        btn_download.setToolTip("Download template Excel untuk input item")
        btn_download.clicked.connect(self.download_template)
        header.addWidget(btn_download)
        
        # Upload button
        btn_upload = QPushButton("📤 Upload Excel")
        btn_upload.setToolTip("Import item dari file Excel")
        btn_upload.clicked.connect(self.upload_items)
        header.addWidget(btn_upload)
        
        # Survey Toko button
        btn_survey = QPushButton("🏪 Data Toko Survey")
        btn_survey.setToolTip("Kelola data toko/sumber harga survey")
        btn_survey.clicked.connect(self.open_survey_toko)
        header.addWidget(btn_survey)
        
        # Add button
        btn_add = QPushButton("➕ Tambah Item")
        btn_add.setObjectName("btnPrimary")
        btn_add.clicked.connect(self.add_item)
        header.addWidget(btn_add)
        
        layout.addLayout(header)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "No", "Kategori", "Kelompok", "Uraian", "Spesifikasi",
            "Satuan", "Volume", "Harga Satuan", "Total", "Keterangan", "Aksi", "ID"
        ])
        
        # Hide ID column
        self.table.setColumnHidden(11, True)
        
        # Column widths
        self.table.setColumnWidth(0, 40)    # No
        self.table.setColumnWidth(1, 60)    # Kategori
        self.table.setColumnWidth(2, 100)   # Kelompok
        self.table.setColumnWidth(3, 200)   # Uraian
        self.table.setColumnWidth(4, 150)   # Spesifikasi
        self.table.setColumnWidth(5, 60)    # Satuan
        self.table.setColumnWidth(6, 70)    # Volume
        self.table.setColumnWidth(7, 120)   # Harga Satuan
        self.table.setColumnWidth(8, 130)   # Total
        self.table.setColumnWidth(9, 100)   # Keterangan
        self.table.setColumnWidth(10, 80)   # Aksi
        
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self.edit_item)
        
        # Context menu
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        layout.addWidget(self.table)
        
        # Summary
        summary_group = QGroupBox("Ringkasan")
        summary_layout = QGridLayout()
        
        summary_layout.addWidget(QLabel("Jumlah Item:"), 0, 0)
        self.lbl_count = QLabel("0")
        self.lbl_count.setStyleSheet("font-weight: bold;")
        summary_layout.addWidget(self.lbl_count, 0, 1)
        
        summary_layout.addWidget(QLabel("Subtotal:"), 0, 2)
        self.lbl_subtotal = QLabel("Rp 0")
        self.lbl_subtotal.setStyleSheet("font-weight: bold;")
        summary_layout.addWidget(self.lbl_subtotal, 0, 3)
        
        summary_layout.addWidget(QLabel("PPN 11%:"), 1, 0)
        self.lbl_ppn = QLabel("Rp 0")
        self.lbl_ppn.setStyleSheet("font-weight: bold;")
        summary_layout.addWidget(self.lbl_ppn, 1, 1)
        
        summary_layout.addWidget(QLabel("Grand Total:"), 1, 2)
        self.lbl_grand_total = QLabel("Rp 0")
        self.lbl_grand_total.setStyleSheet("font-weight: bold; font-size: 14px; color: #27ae60;")
        summary_layout.addWidget(self.lbl_grand_total, 1, 3)
        
        summary_layout.addWidget(QLabel("PPh:"), 2, 0)
        self.lbl_pph = QLabel("Rp 0")
        self.lbl_pph.setStyleSheet("font-weight: bold;")
        summary_layout.addWidget(self.lbl_pph, 2, 1)
        
        summary_layout.addWidget(QLabel("Nilai Bersih:"), 2, 2)
        self.lbl_netto = QLabel("Rp 0")
        self.lbl_netto.setStyleSheet("font-weight: bold; font-size: 14px; color: #2980b9;")
        summary_layout.addWidget(self.lbl_netto, 2, 3)
        
        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)
        
        # HPS Calculation Section
        hps_group = QGroupBox("Perhitungan Harga HPS")
        hps_layout = QHBoxLayout()
        
        hps_layout.addWidget(QLabel("Metode:"))
        
        self.cmb_metode_bulk = QComboBox()
        self.cmb_metode_bulk.addItem("📊 Rata-rata", "RATA")
        self.cmb_metode_bulk.addItem("📈 Harga Tertinggi", "TERTINGGI")
        self.cmb_metode_bulk.addItem("📉 Harga Terendah", "TERENDAH")
        
        # Set from paket setting
        idx = self.cmb_metode_bulk.findData(self.paket.get('metode_hps', 'RATA'))
        if idx >= 0:
            self.cmb_metode_bulk.setCurrentIndex(idx)
        hps_layout.addWidget(self.cmb_metode_bulk)
        
        btn_recalc = QPushButton("🔄 Hitung Ulang Semua Harga HPS")
        btn_recalc.setToolTip("Hitung ulang harga satuan semua item berdasarkan metode yang dipilih")
        btn_recalc.clicked.connect(self.recalculate_all_hps)
        hps_layout.addWidget(btn_recalc)
        
        btn_save_metode = QPushButton("💾 Simpan Metode")
        btn_save_metode.setToolTip("Simpan metode HPS sebagai default paket ini")
        btn_save_metode.clicked.connect(self.save_metode_hps)
        hps_layout.addWidget(btn_save_metode)
        
        hps_layout.addStretch()
        
        hps_group.setLayout(hps_layout)
        layout.addWidget(hps_group)
        
        # Bottom buttons
        btn_layout = QHBoxLayout()
        
        btn_sync = QPushButton("🔄 Sinkron ke Paket")
        btn_sync.setToolTip("Update nilai HPS/Kontrak paket dari total item")
        btn_sync.clicked.connect(self.sync_to_paket)
        btn_layout.addWidget(btn_sync)
        
        btn_export = QPushButton("📊 Export Excel")
        btn_export.setToolTip("Export daftar item ke Excel")
        btn_export.clicked.connect(self.export_items)
        btn_layout.addWidget(btn_export)
        
        btn_clear = QPushButton("🗑️ Hapus Semua")
        btn_clear.setToolTip("Hapus semua item")
        btn_clear.setObjectName("btnDanger")
        btn_clear.clicked.connect(self.clear_all_items)
        btn_layout.addWidget(btn_clear)
        
        btn_layout.addStretch()
        
        btn_close = QPushButton("Tutup")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
        
        # Style
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton#btnPrimary {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton#btnPrimary:hover {
                background-color: #2980b9;
            }
            QPushButton#btnSuccess {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton#btnDanger {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 4px 8px;
                border-radius: 4px;
            }
        """)
    
    def load_items(self):
        """Load items from database"""
        summary = self.db.get_item_barang_summary(self.paket_id)
        items = summary['items']
        
        self.table.setRowCount(len(items))
        
        for row, item in enumerate(items):
            # No
            self.table.setItem(row, 0, QTableWidgetItem(str(item['nomor_urut'])))
            
            # Kategori
            kat = item.get('kategori', '')
            kat_name = KATEGORI_ITEM.get(kat, kat)
            self.table.setItem(row, 1, QTableWidgetItem(f"{kat}"))
            
            # Kelompok
            self.table.setItem(row, 2, QTableWidgetItem(item.get('kelompok', '')))
            
            # Uraian
            self.table.setItem(row, 3, QTableWidgetItem(item.get('uraian', '')))
            
            # Spesifikasi
            self.table.setItem(row, 4, QTableWidgetItem(item.get('spesifikasi', '')))
            
            # Satuan
            self.table.setItem(row, 5, QTableWidgetItem(item.get('satuan', '')))
            
            # Volume
            vol_item = QTableWidgetItem(f"{item.get('volume', 0):,.2f}")
            vol_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 6, vol_item)
            
            # Harga Satuan
            harga_item = QTableWidgetItem(format_rupiah(item.get('harga_dasar', 0)))
            harga_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 7, harga_item)
            
            # Total
            total_item = QTableWidgetItem(format_rupiah(item.get('total', 0)))
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 8, total_item)
            
            # Keterangan
            self.table.setItem(row, 9, QTableWidgetItem(item.get('keterangan', '')))
            
            # Aksi buttons
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            btn_layout.setSpacing(2)
            
            btn_edit = QPushButton("✏️")
            btn_edit.setFixedSize(30, 25)
            btn_edit.setToolTip("Edit")
            btn_edit.clicked.connect(lambda checked, r=row: self.edit_item_at_row(r))
            btn_layout.addWidget(btn_edit)
            
            btn_delete = QPushButton("🗑️")
            btn_delete.setFixedSize(30, 25)
            btn_delete.setToolTip("Hapus")
            btn_delete.clicked.connect(lambda checked, r=row: self.delete_item_at_row(r))
            btn_layout.addWidget(btn_delete)
            
            self.table.setCellWidget(row, 10, btn_widget)
            
            # Hidden ID
            self.table.setItem(row, 11, QTableWidgetItem(str(item['id'])))
        
        # Update summary
        self.lbl_count.setText(str(summary['count']))
        self.lbl_subtotal.setText(format_rupiah(summary['subtotal']))
        self.lbl_ppn.setText(format_rupiah(summary['ppn']))
        self.lbl_grand_total.setText(format_rupiah(summary['grand_total']))
        self.lbl_pph.setText(format_rupiah(summary['pph']))
        self.lbl_netto.setText(format_rupiah(summary['nilai_bersih']))
    
    def open_survey_toko(self):
        """Open Survey Toko Manager"""
        from app.ui.survey_toko_manager import SurveyTokoManager
        dialog = SurveyTokoManager(self.paket_id, parent=self)
        dialog.exec()
    
    def add_item(self):
        """Add new item"""
        dialog = ItemBarangDialog(self.paket_id, parent=self)
        if dialog.exec() == QDialog.Accepted:
            self.load_items()
            self.items_changed.emit()
    
    def edit_item(self):
        """Edit selected item (double-click)"""
        row = self.table.currentRow()
        if row >= 0:
            self.edit_item_at_row(row)
    
    def edit_item_at_row(self, row: int):
        """Edit item at specific row"""
        item_id = int(self.table.item(row, 11).text())
        items = self.db.get_item_barang(self.paket_id)
        item_data = next((i for i in items if i['id'] == item_id), None)
        
        if item_data:
            dialog = ItemBarangDialog(self.paket_id, item_data, parent=self)
            if dialog.exec() == QDialog.Accepted:
                self.load_items()
                self.items_changed.emit()
    
    def delete_item_at_row(self, row: int):
        """Delete item at specific row"""
        item_id = int(self.table.item(row, 11).text())
        uraian = self.table.item(row, 3).text()
        
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus item '{uraian}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.db.delete_item_barang(item_id)
            self.load_items()
            self.items_changed.emit()
    
    def show_context_menu(self, position):
        """Show context menu on right-click"""
        menu = QMenu()
        
        edit_action = menu.addAction("✏️ Edit")
        delete_action = menu.addAction("🗑️ Hapus")
        menu.addSeparator()
        duplicate_action = menu.addAction("📋 Duplikat")
        
        action = menu.exec(self.table.mapToGlobal(position))
        
        row = self.table.currentRow()
        if row < 0:
            return
        
        if action == edit_action:
            self.edit_item_at_row(row)
        elif action == delete_action:
            self.delete_item_at_row(row)
        elif action == duplicate_action:
            self.duplicate_item_at_row(row)
    
    def duplicate_item_at_row(self, row: int):
        """Duplicate item"""
        item_id = int(self.table.item(row, 11).text())
        items = self.db.get_item_barang(self.paket_id)
        item_data = next((i for i in items if i['id'] == item_id), None)
        
        if item_data:
            # Remove id and timestamps
            new_data = {k: v for k, v in item_data.items() 
                       if k not in ['id', 'nomor_urut', 'created_at', 'updated_at', 'is_active', 'paket_id']}
            new_data['uraian'] = f"(Copy) {new_data.get('uraian', '')}"
            
            self.db.add_item_barang(self.paket_id, new_data)
            self.load_items()
            self.items_changed.emit()
    
    def sync_to_paket(self):
        """Sync totals to paket nilai_hps/nilai_kontrak"""
        summary = self.db.get_item_barang_summary(self.paket_id)
        
        reply = QMessageBox.question(
            self, "Sinkronisasi",
            f"Update nilai paket dengan:\n\n"
            f"• Nilai HPS: {format_rupiah(summary['grand_total'])}\n"
            f"• Nilai Kontrak: {format_rupiah(summary['subtotal'])}\n"
            f"• PPN: {format_rupiah(summary['ppn'])}\n"
            f"• PPh: {format_rupiah(summary['pph'])}\n\n"
            f"Lanjutkan?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.db.update_paket(self.paket_id, {
                'nilai_hps': summary['grand_total'],
                'nilai_kontrak': summary['subtotal'],
                'nilai_ppn': summary['ppn'],
                'nilai_pph': summary['pph']
            })
            QMessageBox.information(self, "Sukses", "Nilai paket berhasil diupdate!")
            self.items_changed.emit()
    
    def download_template(self):
        """Download Excel template for item input"""
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan Template Excel",
            f"Template_Item_{self.paket.get('kode', 'Paket')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filepath:
            return
        
        try:
            ExcelTemplateGenerator.create_upload_template(
                filepath, 
                self.paket.get('nama', '')
            )
            
            QMessageBox.information(
                self, "Sukses",
                f"Template berhasil disimpan!\n\n"
                f"File: {filepath}\n\n"
                f"Petunjuk:\n"
                f"1. Buka file Excel\n"
                f"2. Isi data mulai baris 8\n"
                f"3. Kolom dengan * wajib diisi\n"
                f"4. Simpan dan upload kembali"
            )
            
            # Open file location
            import subprocess
            import sys
            if sys.platform == 'win32':
                subprocess.Popen(f'explorer /select,"{filepath}"')
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', '-R', filepath])
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuat template:\n{str(e)}")
    
    def upload_items(self):
        """Upload items from Excel file"""
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Pilih File Excel",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if not filepath:
            return
        
        # Parse file
        items, errors = ExcelTemplateGenerator.parse_upload_file(filepath)
        
        if errors and not items:
            QMessageBox.warning(
                self, "Error",
                "Tidak dapat mengimport data:\n\n" + "\n".join(errors[:10])
            )
            return
        
        # Confirm import
        msg = f"Ditemukan {len(items)} item valid untuk diimport."
        if errors:
            msg += f"\n\n⚠️ {len(errors)} baris bermasalah:\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                msg += f"\n... dan {len(errors) - 5} error lainnya"
        
        # Ask mode: append or replace
        reply = QMessageBox.question(
            self, "Konfirmasi Import",
            f"{msg}\n\nPilih mode import:\n"
            f"• Yes = Tambahkan ke item yang ada\n"
            f"• No = Hapus item lama, ganti dengan yang baru",
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Cancel:
            return
        
        try:
            # Show progress
            progress = QProgressDialog("Mengimport item...", None, 0, 0, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.show()
            QApplication.processEvents()
            
            # If replace mode, clear existing items (bulk delete)
            if reply == QMessageBox.No:
                progress.setLabelText("Menghapus item lama...")
                QApplication.processEvents()
                self.db.bulk_delete_item_barang(self.paket_id)
            
            # Bulk import items (much faster than one-by-one)
            progress.setLabelText(f"Mengimport {len(items)} item...")
            QApplication.processEvents()
            
            imported = self.db.bulk_add_item_barang(self.paket_id, items)
            
            progress.close()
            
            self.load_items()
            self.items_changed.emit()
            
            QMessageBox.information(
                self, "Sukses",
                f"Berhasil mengimport {imported} item!"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengimport:\n{str(e)}")
    
    def export_items(self):
        """Export items to Excel file"""
        items = self.db.get_item_barang(self.paket_id)
        
        if not items:
            QMessageBox.warning(self, "Peringatan", "Tidak ada item untuk diexport!")
            return
        
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan Export Excel",
            f"Export_Item_{self.paket.get('kode', 'Paket')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filepath:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Daftar Item"
            
            # Styles
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center = Alignment(horizontal='center', vertical='center')
            
            # Title
            ws.merge_cells('A1:L1')
            ws['A1'] = f"DAFTAR KEBUTUHAN BARANG - {self.paket.get('nama', '')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center
            
            # Headers
            headers = ['No', 'Kategori', 'Kelompok', 'Uraian', 'Spesifikasi', 'Satuan',
                      'Volume', 'Harga Survey 1', 'Harga Survey 2', 'Harga Survey 3', 
                      'Harga Satuan', 'Total']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center
            
            # Data
            summary = self.db.get_item_barang_summary(self.paket_id)
            
            for row_idx, item in enumerate(items, 4):
                data = [
                    item['nomor_urut'],
                    item.get('kategori', ''),
                    item.get('kelompok', ''),
                    item.get('uraian', ''),
                    item.get('spesifikasi', ''),
                    item.get('satuan', ''),
                    item.get('volume', 0),
                    item.get('harga_survey1') or '',
                    item.get('harga_survey2') or '',
                    item.get('harga_survey3') or '',
                    item.get('harga_dasar', 0),
                    item.get('total', 0)
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx in [7, 8, 9, 10, 11, 12]:  # Number columns
                        cell.number_format = '#,##0'
            
            # Summary row
            last_row = len(items) + 4
            ws.merge_cells(f'A{last_row}:K{last_row}')
            ws[f'A{last_row}'] = 'SUBTOTAL'
            ws[f'A{last_row}'].font = Font(bold=True)
            ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'A{last_row}'].border = border
            ws[f'L{last_row}'] = summary['subtotal']
            ws[f'L{last_row}'].font = Font(bold=True)
            ws[f'L{last_row}'].number_format = '#,##0'
            ws[f'L{last_row}'].border = border
            
            # PPN row
            last_row += 1
            ws.merge_cells(f'A{last_row}:K{last_row}')
            ws[f'A{last_row}'] = 'PPN 11%'
            ws[f'A{last_row}'].font = Font(bold=True)
            ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'A{last_row}'].border = border
            ws[f'L{last_row}'] = summary['ppn']
            ws[f'L{last_row}'].font = Font(bold=True)
            ws[f'L{last_row}'].number_format = '#,##0'
            ws[f'L{last_row}'].border = border
            
            # Grand total row
            last_row += 1
            ws.merge_cells(f'A{last_row}:K{last_row}')
            ws[f'A{last_row}'] = 'GRAND TOTAL'
            ws[f'A{last_row}'].font = Font(bold=True, size=12)
            ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'A{last_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws[f'A{last_row}'].border = border
            ws[f'L{last_row}'] = summary['grand_total']
            ws[f'L{last_row}'].font = Font(bold=True, size=12)
            ws[f'L{last_row}'].number_format = '#,##0'
            ws[f'L{last_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws[f'L{last_row}'].border = border
            
            # Column widths
            widths = [5, 10, 15, 35, 25, 8, 10, 15, 15, 15, 15, 18]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(filepath)
            
            QMessageBox.information(
                self, "Sukses",
                f"Data berhasil diexport ke:\n{filepath}"
            )
            
            # Open file
            import subprocess
            import sys
            if sys.platform == 'win32':
                os.startfile(filepath)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', filepath])
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengexport:\n{str(e)}")
    
    def recalculate_all_hps(self):
        """Recalculate all item prices based on selected HPS method"""
        from app.core.config import hitung_harga_hps
        
        items = self.db.get_item_barang(self.paket_id)
        
        if not items:
            QMessageBox.information(self, "Info", "Tidak ada item untuk dihitung ulang.")
            return
        
        # Count items with survey prices
        items_with_survey = [i for i in items if any([
            i.get('harga_survey1'), i.get('harga_survey2'), i.get('harga_survey3')
        ])]
        
        if not items_with_survey:
            QMessageBox.warning(
                self, "Peringatan", 
                "Tidak ada item dengan harga survey.\n"
                "Isi harga survey terlebih dahulu untuk dapat menghitung HPS."
            )
            return
        
        metode = self.cmb_metode_bulk.currentData()
        metode_name = self.cmb_metode_bulk.currentText()
        
        reply = QMessageBox.question(
            self, "Konfirmasi Hitung Ulang",
            f"Hitung ulang harga satuan {len(items_with_survey)} item\n"
            f"menggunakan metode: {metode_name}?\n\n"
            f"Harga satuan yang ada akan ditimpa dengan hasil perhitungan.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            return
        
        try:
            updated = 0
            for item in items_with_survey:
                harga_hps = hitung_harga_hps(
                    item.get('harga_survey1', 0),
                    item.get('harga_survey2', 0),
                    item.get('harga_survey3', 0),
                    metode
                )
                
                if harga_hps > 0:
                    # Update item with new harga_dasar
                    item['harga_dasar'] = harga_hps
                    self.db.update_item_barang(item['id'], item)
                    updated += 1
            
            self.load_items()
            self.items_changed.emit()
            
            QMessageBox.information(
                self, "Sukses",
                f"Berhasil menghitung ulang harga {updated} item\n"
                f"dengan metode: {metode_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menghitung ulang:\n{str(e)}")
    
    def save_metode_hps(self):
        """Save HPS method as paket default"""
        metode = self.cmb_metode_bulk.currentData()
        metode_name = self.cmb_metode_bulk.currentText()
        
        try:
            # Update paket metode_hps
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE paket SET metode_hps = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (metode, self.paket_id))
                conn.commit()
            
            # Refresh paket data
            self.paket = self.db.get_paket(self.paket_id)
            
            QMessageBox.information(
                self, "Sukses",
                f"Metode HPS default paket disimpan:\n{metode_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan:\n{str(e)}")
    
    def clear_all_items(self):
        """Clear all items"""
        items = self.db.get_item_barang(self.paket_id)
        
        if not items:
            QMessageBox.information(self, "Info", "Tidak ada item untuk dihapus.")
            return
        
        reply = QMessageBox.warning(
            self, "Konfirmasi Hapus",
            f"Hapus SEMUA {len(items)} item?\n\n"
            f"Tindakan ini tidak dapat dibatalkan!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            deleted = self.db.bulk_delete_item_barang(self.paket_id)
            
            self.load_items()
            self.items_changed.emit()
            QMessageBox.information(self, "Sukses", f"{deleted} item berhasil dihapus.")


# Singleton instance for quick access
def show_item_barang_manager(paket_id: int, parent=None) -> ItemBarangManager:
    """Show item barang manager dialog"""
    dialog = ItemBarangManager(paket_id, parent)
    dialog.exec()
    return dialog
