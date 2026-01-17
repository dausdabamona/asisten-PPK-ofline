"""
PPK DOCUMENT FACTORY v4.0 - Harga Lifecycle Manager
====================================================
Alur Penetapan Harga:
1. Spesifikasi & BOQ
2. Survey Harga (min 3 sumber per item)
3. Penetapan HPS
4. Revisi/Klarifikasi Harga
5. Kontrak Final
"""

import os
from datetime import datetime, date
from typing import List, Dict, Optional, Tuple

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QWidget, QPushButton, QLabel, QLineEdit, QTextEdit, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QMessageBox, QDoubleSpinBox, QFrame, QSplitter, QTabWidget,
    QAbstractItemView, QFileDialog, QProgressBar, QScrollArea,
    QApplication
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont

from app.core.database_v4 import get_db_manager_v4, WORKFLOW_STAGES_V4
from app.core.config import METODE_HPS, hitung_harga_hps


def format_rupiah(value) -> str:
    """Format number as Rupiah"""
    if value is None:
        return "Rp 0"
    try:
        return f"Rp {int(value):,}".replace(",", ".")
    except:
        return "Rp 0"


# ============================================================================
# SURVEY HARGA DIALOG (Per Item)
# ============================================================================

class SurveyHargaItemDialog(QDialog):
    """Dialog for entering survey harga for single item"""
    
    def __init__(self, paket_id: int, item_data: dict, parent=None):
        super().__init__(parent)
        self.paket_id = paket_id
        self.item_data = item_data
        self.db = get_db_manager_v4()
        
        self.setWindowTitle(f"Survey Harga - {item_data.get('uraian', '')[:50]}")
        self.setMinimumWidth(700)
        
        self.setup_ui()
        self.load_existing()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Item info
        info = QGroupBox("Informasi Item")
        info_layout = QFormLayout()
        info_layout.addRow("Uraian:", QLabel(self.item_data.get('uraian', '')))
        info_layout.addRow("Spesifikasi:", QLabel(self.item_data.get('spesifikasi', '') or '-'))
        info_layout.addRow("Volume:", QLabel(f"{self.item_data.get('volume', 0)} {self.item_data.get('satuan', '')}"))
        info.setLayout(info_layout)
        layout.addWidget(info)
        
        # Survey sources
        self.survey_widgets = []
        
        for i in range(1, 4):
            group = QGroupBox(f"Sumber {i}")
            group_layout = QFormLayout()
            
            # Jenis survey
            cmb_jenis = QComboBox()
            cmb_jenis.addItem("🏪 Luring (Toko Fisik)", "LURING")
            cmb_jenis.addItem("🌐 Daring (Marketplace)", "DARING")
            cmb_jenis.addItem("📄 Kontrak Sebelumnya", "KONTRAK")
            cmb_jenis.addItem("📋 Brosur/Katalog", "BROSUR")
            group_layout.addRow("Jenis:", cmb_jenis)
            
            # Nama sumber
            txt_nama = QLineEdit()
            txt_nama.setPlaceholderText("Nama toko/marketplace/kontrak...")
            group_layout.addRow("Sumber:", txt_nama)
            
            # Alamat/Link
            txt_alamat = QLineEdit()
            txt_alamat.setPlaceholderText("Alamat atau link produk...")
            group_layout.addRow("Alamat/Link:", txt_alamat)
            
            # Harga
            spn_harga = QDoubleSpinBox()
            spn_harga.setRange(0, 999999999999)
            spn_harga.setPrefix("Rp ")
            spn_harga.setGroupSeparatorShown(True)
            spn_harga.valueChanged.connect(self.calculate_summary)
            group_layout.addRow("Harga:", spn_harga)
            
            # Bukti upload
            bukti_layout = QHBoxLayout()
            txt_bukti = QLineEdit()
            txt_bukti.setReadOnly(True)
            txt_bukti.setPlaceholderText("Belum ada bukti...")
            bukti_layout.addWidget(txt_bukti)
            
            btn_upload = QPushButton("📎 Upload")
            btn_upload.setMaximumWidth(80)
            bukti_layout.addWidget(btn_upload)
            
            group_layout.addRow("Bukti:", bukti_layout)
            
            group.setLayout(group_layout)
            layout.addWidget(group)
            
            self.survey_widgets.append({
                'jenis': cmb_jenis,
                'nama': txt_nama,
                'alamat': txt_alamat,
                'harga': spn_harga,
                'bukti': txt_bukti,
            })
        
        # Summary
        summary = QGroupBox("Ringkasan Harga")
        summary_layout = QGridLayout()
        
        summary_layout.addWidget(QLabel("Harga Terendah:"), 0, 0)
        self.lbl_min = QLabel("Rp 0")
        self.lbl_min.setStyleSheet("font-weight: bold; color: #27ae60;")
        summary_layout.addWidget(self.lbl_min, 0, 1)
        
        summary_layout.addWidget(QLabel("Harga Rata-rata:"), 0, 2)
        self.lbl_avg = QLabel("Rp 0")
        self.lbl_avg.setStyleSheet("font-weight: bold; color: #2980b9;")
        summary_layout.addWidget(self.lbl_avg, 0, 3)
        
        summary_layout.addWidget(QLabel("Harga Tertinggi:"), 0, 4)
        self.lbl_max = QLabel("Rp 0")
        self.lbl_max.setStyleSheet("font-weight: bold; color: #e74c3c;")
        summary_layout.addWidget(self.lbl_max, 0, 5)
        
        summary.setLayout(summary_layout)
        layout.addWidget(summary)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_save = QPushButton("💾 Simpan Survey")
        btn_save.setObjectName("btnSuccess")
        btn_save.clicked.connect(self.save)
        btn_layout.addWidget(btn_save)
        
        btn_close = QPushButton("Tutup")
        btn_close.clicked.connect(self.reject)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
        
        # Style
        self.setStyleSheet("""
            QGroupBox { font-weight: bold; margin-top: 10px; }
            QPushButton#btnSuccess {
                background-color: #27ae60; color: white;
                border: none; padding: 10px 20px; border-radius: 4px;
            }
            QPushButton#btnSuccess:hover { background-color: #219a52; }
        """)
    
    def load_existing(self):
        """Load existing survey data"""
        surveys = self.db.get_survey_harga_detail(self.paket_id, self.item_data['id'])
        
        for survey in surveys:
            idx = survey.get('sumber_ke', 1) - 1
            if 0 <= idx < 3:
                w = self.survey_widgets[idx]
                
                # Set jenis
                jenis_idx = w['jenis'].findData(survey.get('jenis_survey'))
                if jenis_idx >= 0:
                    w['jenis'].setCurrentIndex(jenis_idx)
                
                w['nama'].setText(survey.get('nama_sumber', '') or '')
                w['alamat'].setText(survey.get('alamat', '') or survey.get('link_produk', '') or '')
                w['harga'].setValue(survey.get('harga', 0) or 0)
                
                if survey.get('bukti_path'):
                    w['bukti'].setText(os.path.basename(survey['bukti_path']))
        
        self.calculate_summary()
    
    def calculate_summary(self):
        """Calculate price summary"""
        prices = []
        for w in self.survey_widgets:
            h = w['harga'].value()
            if h > 0:
                prices.append(h)
        
        if prices:
            self.lbl_min.setText(format_rupiah(min(prices)))
            self.lbl_avg.setText(format_rupiah(sum(prices) / len(prices)))
            self.lbl_max.setText(format_rupiah(max(prices)))
        else:
            self.lbl_min.setText("Rp 0")
            self.lbl_avg.setText("Rp 0")
            self.lbl_max.setText("Rp 0")
    
    def save(self):
        """Save survey data"""
        try:
            # Delete existing
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE survey_harga_detail SET is_active = 0
                    WHERE paket_id = ? AND item_id = ?
                """, (self.paket_id, self.item_data['id']))
                conn.commit()
            
            # Save new
            prices = []
            for i, w in enumerate(self.survey_widgets, 1):
                harga = w['harga'].value()
                if harga > 0:
                    self.db.add_survey_harga_detail(self.paket_id, self.item_data['id'], {
                        'sumber_ke': i,
                        'jenis_survey': w['jenis'].currentData(),
                        'nama_sumber': w['nama'].text().strip(),
                        'alamat': w['alamat'].text().strip(),
                        'harga': harga,
                    })
                    prices.append(harga)
            
            # Update item_barang with survey prices
            if prices:
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE item_barang SET
                            harga_survey1 = ?,
                            harga_survey2 = ?,
                            harga_survey3 = ?,
                            harga_rata = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                    """, (
                        prices[0] if len(prices) > 0 else None,
                        prices[1] if len(prices) > 1 else None,
                        prices[2] if len(prices) > 2 else None,
                        sum(prices) / len(prices) if prices else None,
                        self.item_data['id']
                    ))
                    conn.commit()
            
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan:\n{str(e)}")


# ============================================================================
# HARGA LIFECYCLE MANAGER (Main Dialog)
# ============================================================================

class HargaLifecycleManager(QDialog):
    """Main dialog for managing harga lifecycle"""
    
    data_changed = Signal()
    
    def __init__(self, paket_id: int, parent=None):
        super().__init__(parent)
        self.paket_id = paket_id
        self.db = get_db_manager_v4()
        
        # Get paket info
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM paket WHERE id = ?", (paket_id,))
            row = cursor.fetchone()
            self.paket = dict(row) if row else {}
        
        self.setWindowTitle(f"Lifecycle Harga - {self.paket.get('nama', '')}")
        self.setMinimumSize(1100, 700)
        
        self.setup_ui()
        self.load_data()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Header with workflow progress
        header = QWidget()
        header.setStyleSheet("background: #f8f9fa; padding: 10px;")
        header_layout = QVBoxLayout(header)
        
        title = QLabel(f"📊 Lifecycle Harga: {self.paket.get('nama', '')}")
        title.setStyleSheet("font-size: 14px; font-weight: bold;")
        header_layout.addWidget(title)
        
        # Stage progress
        progress_layout = QHBoxLayout()
        stages = ['SPESIFIKASI', 'SURVEY', 'HPS', 'KONTRAK_DRAFT', 'KONTRAK_FINAL']
        stage_names = ['Spesifikasi', 'Survey', 'HPS', 'Draft Kontrak', 'Kontrak Final']
        
        for i, (code, name) in enumerate(zip(stages, stage_names)):
            lbl = QLabel(f"{'✅' if self.is_stage_complete(code) else '⬜'} {name}")
            lbl.setStyleSheet("padding: 5px;")
            progress_layout.addWidget(lbl)
            
            if i < len(stages) - 1:
                progress_layout.addWidget(QLabel("→"))
        
        progress_layout.addStretch()
        header_layout.addLayout(progress_layout)
        
        layout.addWidget(header)
        
        # Tabs
        tabs = QTabWidget()
        
        # Tab 1: Items & Survey
        tab_survey = self.create_survey_tab()
        tabs.addTab(tab_survey, "📋 Item & Survey Harga")
        
        # Tab 2: HPS
        tab_hps = self.create_hps_tab()
        tabs.addTab(tab_hps, "📊 Penetapan HPS")
        
        # Tab 3: Revisi & Kontrak Final
        tab_kontrak = self.create_kontrak_tab()
        tabs.addTab(tab_kontrak, "📝 Revisi & Kontrak Final")
        
        # Tab 4: History
        tab_history = self.create_history_tab()
        tabs.addTab(tab_history, "📜 Riwayat Harga")
        
        layout.addWidget(tabs)
        
        # Bottom buttons
        btn_layout = QHBoxLayout()
        
        btn_validate = QPushButton("✓ Validasi Workflow")
        btn_validate.clicked.connect(self.validate_workflow)
        btn_layout.addWidget(btn_validate)
        
        btn_layout.addStretch()
        
        btn_close = QPushButton("Tutup")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
    
    def is_stage_complete(self, stage_code: str) -> bool:
        """Check if stage is complete"""
        lock_map = {
            'SPESIFIKASI': 'spesifikasi_locked',
            'SURVEY': 'survey_locked',
            'HPS': 'hps_locked',
            'KONTRAK_DRAFT': 'kontrak_draft_locked',
            'KONTRAK_FINAL': 'kontrak_final_locked',
        }
        col = lock_map.get(stage_code)
        if col:
            return bool(self.paket.get(col))
        return False
    
    def create_survey_tab(self) -> QWidget:
        """Create survey tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Toolbar
        toolbar = QHBoxLayout()
        
        toolbar.addWidget(QLabel("Status Survey:"))
        self.lbl_survey_status = QLabel("Loading...")
        self.lbl_survey_status.setStyleSheet("font-weight: bold;")
        toolbar.addWidget(self.lbl_survey_status)
        
        toolbar.addStretch()
        
        # Import/Export buttons
        btn_export_survey = QPushButton("📥 Export Excel")
        btn_export_survey.setToolTip("Export data survey ke Excel untuk pengisian offline")
        btn_export_survey.clicked.connect(self.export_survey_excel)
        toolbar.addWidget(btn_export_survey)
        
        btn_import_survey = QPushButton("📤 Import Excel")
        btn_import_survey.setToolTip("Import data survey dari Excel")
        btn_import_survey.clicked.connect(self.import_survey_excel)
        toolbar.addWidget(btn_import_survey)
        
        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self.load_data)
        toolbar.addWidget(btn_refresh)
        
        layout.addLayout(toolbar)
        
        # Table
        self.table_survey = QTableWidget()
        self.table_survey.setColumnCount(9)
        self.table_survey.setHorizontalHeaderLabels([
            "No", "Uraian", "Volume", "Survey 1", "Survey 2", "Survey 3",
            "Rata-rata", "Status", "Aksi"
        ])
        
        self.table_survey.setColumnWidth(0, 40)
        self.table_survey.setColumnWidth(1, 250)
        self.table_survey.setColumnWidth(2, 80)
        self.table_survey.setColumnWidth(3, 100)
        self.table_survey.setColumnWidth(4, 100)
        self.table_survey.setColumnWidth(5, 100)
        self.table_survey.setColumnWidth(6, 100)
        self.table_survey.setColumnWidth(7, 80)
        self.table_survey.setColumnWidth(8, 80)
        
        self.table_survey.horizontalHeader().setStretchLastSection(False)
        self.table_survey.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_survey.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_survey.setAlternatingRowColors(True)
        
        layout.addWidget(self.table_survey)
        
        # Lock button
        lock_layout = QHBoxLayout()
        lock_layout.addStretch()
        
        self.btn_lock_survey = QPushButton("🔒 Kunci Survey (Lanjut ke HPS)")
        self.btn_lock_survey.setObjectName("btnWarning")
        self.btn_lock_survey.clicked.connect(self.lock_survey)
        lock_layout.addWidget(self.btn_lock_survey)
        
        layout.addLayout(lock_layout)
        
        return widget
    
    def create_hps_tab(self) -> QWidget:
        """Create HPS tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Settings
        settings = QGroupBox("Pengaturan HPS")
        settings_layout = QHBoxLayout()
        
        settings_layout.addWidget(QLabel("Metode Perhitungan:"))
        self.cmb_metode = QComboBox()
        self.cmb_metode.addItem("📊 Rata-rata", "RATA")
        self.cmb_metode.addItem("📈 Harga Tertinggi", "TERTINGGI")
        self.cmb_metode.addItem("📉 Harga Terendah", "TERENDAH")
        
        idx = self.cmb_metode.findData(self.paket.get('metode_hps', 'RATA'))
        if idx >= 0:
            self.cmb_metode.setCurrentIndex(idx)
        
        settings_layout.addWidget(self.cmb_metode)
        
        settings_layout.addWidget(QLabel("Overhead & Profit:"))
        self.spn_overhead = QDoubleSpinBox()
        self.spn_overhead.setRange(0, 50)
        self.spn_overhead.setSuffix(" %")
        self.spn_overhead.setValue((self.paket.get('overhead_profit', 0.10) or 0.10) * 100)
        settings_layout.addWidget(self.spn_overhead)
        
        btn_calculate = QPushButton("🔄 Hitung HPS")
        btn_calculate.clicked.connect(self.calculate_hps)
        settings_layout.addWidget(btn_calculate)
        
        settings_layout.addStretch()
        settings.setLayout(settings_layout)
        layout.addWidget(settings)
        
        # Table
        self.table_hps = QTableWidget()
        self.table_hps.setColumnCount(8)
        self.table_hps.setHorizontalHeaderLabels([
            "No", "Uraian", "Volume", "Satuan", "Harga Survey", 
            "Harga HPS/Satuan", "Total HPS", "ID"
        ])
        
        self.table_hps.setColumnHidden(7, True)
        self.table_hps.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_hps.setAlternatingRowColors(True)
        
        layout.addWidget(self.table_hps)
        
        # Summary
        summary = QGroupBox("Ringkasan HPS")
        summary_layout = QGridLayout()
        
        summary_layout.addWidget(QLabel("Subtotal:"), 0, 0)
        self.lbl_hps_subtotal = QLabel("Rp 0")
        self.lbl_hps_subtotal.setStyleSheet("font-weight: bold;")
        summary_layout.addWidget(self.lbl_hps_subtotal, 0, 1)
        
        summary_layout.addWidget(QLabel("PPN 11%:"), 0, 2)
        self.lbl_hps_ppn = QLabel("Rp 0")
        summary_layout.addWidget(self.lbl_hps_ppn, 0, 3)
        
        summary_layout.addWidget(QLabel("Total HPS:"), 0, 4)
        self.lbl_hps_total = QLabel("Rp 0")
        self.lbl_hps_total.setStyleSheet("font-weight: bold; font-size: 14px; color: #27ae60;")
        summary_layout.addWidget(self.lbl_hps_total, 0, 5)
        
        summary.setLayout(summary_layout)
        layout.addWidget(summary)
        
        # Bottom buttons
        lock_layout = QHBoxLayout()
        
        btn_export_hps = QPushButton("📥 Export HPS ke Excel")
        btn_export_hps.clicked.connect(self.export_hps_excel)
        lock_layout.addWidget(btn_export_hps)
        
        lock_layout.addStretch()
        
        self.btn_lock_hps = QPushButton("🔒 Tetapkan HPS Final")
        self.btn_lock_hps.setObjectName("btnSuccess")
        self.btn_lock_hps.clicked.connect(self.lock_hps)
        lock_layout.addWidget(self.btn_lock_hps)
        
        layout.addLayout(lock_layout)
        
        return widget
    
    def create_kontrak_tab(self) -> QWidget:
        """Create kontrak tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Info
        info = QLabel(
            "💡 Setelah proses pemilihan penyedia, masukkan harga hasil negosiasi/klarifikasi di sini.\n"
            "Harga Kontrak Final akan digunakan untuk pembayaran (SPP, SSP, Kuitansi)."
        )
        info.setWordWrap(True)
        info.setStyleSheet("padding: 10px; background: #e3f2fd; border-radius: 4px;")
        layout.addWidget(info)
        
        # Toolbar with export/import
        toolbar = QHBoxLayout()
        
        btn_export_kontrak = QPushButton("📥 Export Form Kontrak")
        btn_export_kontrak.setToolTip("Export form untuk pengisian harga kontrak")
        btn_export_kontrak.clicked.connect(self.export_kontrak_excel)
        toolbar.addWidget(btn_export_kontrak)
        
        btn_import_kontrak = QPushButton("📤 Import Harga Kontrak")
        btn_import_kontrak.setToolTip("Import harga kontrak dari Excel")
        btn_import_kontrak.clicked.connect(self.import_kontrak_excel)
        toolbar.addWidget(btn_import_kontrak)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # Table
        self.table_kontrak = QTableWidget()
        self.table_kontrak.setColumnCount(8)
        self.table_kontrak.setHorizontalHeaderLabels([
            "No", "Uraian", "Volume", "Harga HPS", "Harga Kontrak",
            "Selisih", "Total Kontrak", "ID"
        ])
        
        self.table_kontrak.setColumnHidden(7, True)
        self.table_kontrak.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_kontrak.setAlternatingRowColors(True)
        
        layout.addWidget(self.table_kontrak)
        
        # BA Klarifikasi
        ba_group = QGroupBox("Berita Acara Klarifikasi")
        ba_layout = QFormLayout()
        
        self.txt_nomor_ba = QLineEdit()
        self.txt_nomor_ba.setPlaceholderText("Nomor BA Klarifikasi Harga...")
        ba_layout.addRow("Nomor BA:", self.txt_nomor_ba)
        
        self.txt_catatan = QTextEdit()
        self.txt_catatan.setMaximumHeight(80)
        self.txt_catatan.setPlaceholderText("Catatan hasil negosiasi...")
        ba_layout.addRow("Catatan:", self.txt_catatan)
        
        ba_group.setLayout(ba_layout)
        layout.addWidget(ba_group)
        
        # Summary
        summary = QGroupBox("Perbandingan HPS vs Kontrak")
        summary_layout = QGridLayout()
        
        summary_layout.addWidget(QLabel("Total HPS:"), 0, 0)
        self.lbl_kontrak_hps = QLabel("Rp 0")
        self.lbl_kontrak_hps.setStyleSheet("font-weight: bold;")
        summary_layout.addWidget(self.lbl_kontrak_hps, 0, 1)
        
        summary_layout.addWidget(QLabel("Total Kontrak:"), 0, 2)
        self.lbl_kontrak_total = QLabel("Rp 0")
        self.lbl_kontrak_total.setStyleSheet("font-weight: bold; color: #2980b9;")
        summary_layout.addWidget(self.lbl_kontrak_total, 0, 3)
        
        summary_layout.addWidget(QLabel("Selisih:"), 0, 4)
        self.lbl_kontrak_selisih = QLabel("Rp 0")
        self.lbl_kontrak_selisih.setStyleSheet("font-weight: bold; color: #27ae60;")
        summary_layout.addWidget(self.lbl_kontrak_selisih, 0, 5)
        
        summary.setLayout(summary_layout)
        layout.addWidget(summary)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        btn_input = QPushButton("📝 Input Harga Kontrak")
        btn_input.clicked.connect(self.input_harga_kontrak)
        btn_layout.addWidget(btn_input)
        
        btn_layout.addStretch()
        
        self.btn_lock_kontrak = QPushButton("🔒 Tetapkan Kontrak Final")
        self.btn_lock_kontrak.setObjectName("btnSuccess")
        self.btn_lock_kontrak.clicked.connect(self.lock_kontrak)
        btn_layout.addWidget(self.btn_lock_kontrak)
        
        layout.addLayout(btn_layout)
        
        return widget
    
    def create_history_tab(self) -> QWidget:
        """Create history tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        info = QLabel("📜 Riwayat perubahan harga untuk audit trail")
        info.setStyleSheet("padding: 10px; background: #f8f9fa;")
        layout.addWidget(info)
        
        self.table_history = QTableWidget()
        self.table_history.setColumnCount(6)
        self.table_history.setHorizontalHeaderLabels([
            "Tanggal", "Tahap", "Item", "Harga Satuan", "Total", "Keterangan"
        ])
        
        self.table_history.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table_history.setAlternatingRowColors(True)
        
        layout.addWidget(self.table_history)
        
        return widget
    
    def load_data(self):
        """Load all data"""
        self.load_survey_data()
        self.load_hps_data()
        self.load_kontrak_data()
        self.load_history_data()
    
    def load_survey_data(self):
        """Load survey data into table"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM item_barang
                WHERE paket_id = ? AND is_active = 1
                ORDER BY nomor_urut
            """, (self.paket_id,))
            items = [dict(row) for row in cursor.fetchall()]
        
        survey_counts = self.db.count_survey_per_item(self.paket_id)
        
        self.table_survey.setRowCount(len(items))
        
        complete = 0
        for row, item in enumerate(items):
            item_id = item['id']
            survey_count = survey_counts.get(item_id, 0)
            
            self.table_survey.setItem(row, 0, QTableWidgetItem(str(item.get('nomor_urut', row+1))))
            self.table_survey.setItem(row, 1, QTableWidgetItem(item.get('uraian', '')))
            self.table_survey.setItem(row, 2, QTableWidgetItem(f"{item.get('volume', 0)} {item.get('satuan', '')}"))
            self.table_survey.setItem(row, 3, QTableWidgetItem(format_rupiah(item.get('harga_survey1'))))
            self.table_survey.setItem(row, 4, QTableWidgetItem(format_rupiah(item.get('harga_survey2'))))
            self.table_survey.setItem(row, 5, QTableWidgetItem(format_rupiah(item.get('harga_survey3'))))
            self.table_survey.setItem(row, 6, QTableWidgetItem(format_rupiah(item.get('harga_rata'))))
            
            # Status
            if survey_count >= 3:
                status = "✅ Lengkap"
                complete += 1
            elif survey_count > 0:
                status = f"⚠️ {survey_count}/3"
            else:
                status = "❌ Kosong"
            
            status_item = QTableWidgetItem(status)
            self.table_survey.setItem(row, 7, status_item)
            
            # Action button
            btn = QPushButton("📝 Input")
            btn.clicked.connect(lambda checked, i=item: self.open_survey_dialog(i))
            self.table_survey.setCellWidget(row, 8, btn)
        
        # Update status
        if len(items) == 0:
            self.lbl_survey_status.setText("⚠️ Belum ada item")
        elif complete == len(items):
            self.lbl_survey_status.setText(f"✅ Lengkap ({complete}/{len(items)} item)")
            self.lbl_survey_status.setStyleSheet("font-weight: bold; color: #27ae60;")
        else:
            self.lbl_survey_status.setText(f"⏳ Progress: {complete}/{len(items)} item")
            self.lbl_survey_status.setStyleSheet("font-weight: bold; color: #f39c12;")
        
        # Disable lock if not complete
        self.btn_lock_survey.setEnabled(complete == len(items) and len(items) > 0)
    
    def load_hps_data(self):
        """Load HPS data"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM item_barang
                WHERE paket_id = ? AND is_active = 1
                ORDER BY nomor_urut
            """, (self.paket_id,))
            items = [dict(row) for row in cursor.fetchall()]
        
        self.table_hps.setRowCount(len(items))
        
        subtotal = 0
        for row, item in enumerate(items):
            self.table_hps.setItem(row, 0, QTableWidgetItem(str(item.get('nomor_urut', row+1))))
            self.table_hps.setItem(row, 1, QTableWidgetItem(item.get('uraian', '')))
            self.table_hps.setItem(row, 2, QTableWidgetItem(str(item.get('volume', 0))))
            self.table_hps.setItem(row, 3, QTableWidgetItem(item.get('satuan', '')))
            self.table_hps.setItem(row, 4, QTableWidgetItem(format_rupiah(item.get('harga_rata'))))
            self.table_hps.setItem(row, 5, QTableWidgetItem(format_rupiah(item.get('harga_hps_satuan') or item.get('harga_dasar'))))
            
            total_hps = item.get('total_hps') or item.get('total', 0)
            self.table_hps.setItem(row, 6, QTableWidgetItem(format_rupiah(total_hps)))
            self.table_hps.setItem(row, 7, QTableWidgetItem(str(item['id'])))
            
            subtotal += total_hps or 0
        
        ppn = subtotal * 0.11
        total = subtotal + ppn
        
        self.lbl_hps_subtotal.setText(format_rupiah(subtotal))
        self.lbl_hps_ppn.setText(format_rupiah(ppn))
        self.lbl_hps_total.setText(format_rupiah(total))
    
    def load_kontrak_data(self):
        """Load kontrak data"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM item_barang
                WHERE paket_id = ? AND is_active = 1
                ORDER BY nomor_urut
            """, (self.paket_id,))
            items = [dict(row) for row in cursor.fetchall()]
        
        self.table_kontrak.setRowCount(len(items))
        
        total_hps = 0
        total_kontrak = 0
        
        for row, item in enumerate(items):
            self.table_kontrak.setItem(row, 0, QTableWidgetItem(str(item.get('nomor_urut', row+1))))
            self.table_kontrak.setItem(row, 1, QTableWidgetItem(item.get('uraian', '')))
            self.table_kontrak.setItem(row, 2, QTableWidgetItem(str(item.get('volume', 0))))
            
            hps_satuan = item.get('harga_hps_satuan') or item.get('harga_dasar', 0)
            kontrak_satuan = item.get('harga_kontrak_satuan') or 0
            selisih = hps_satuan - kontrak_satuan if kontrak_satuan else 0
            
            self.table_kontrak.setItem(row, 3, QTableWidgetItem(format_rupiah(hps_satuan)))
            self.table_kontrak.setItem(row, 4, QTableWidgetItem(format_rupiah(kontrak_satuan)))
            self.table_kontrak.setItem(row, 5, QTableWidgetItem(format_rupiah(selisih)))
            
            item_total_kontrak = item.get('total_kontrak') or 0
            self.table_kontrak.setItem(row, 6, QTableWidgetItem(format_rupiah(item_total_kontrak)))
            self.table_kontrak.setItem(row, 7, QTableWidgetItem(str(item['id'])))
            
            total_hps += (item.get('total_hps') or item.get('total', 0))
            total_kontrak += item_total_kontrak
        
        self.lbl_kontrak_hps.setText(format_rupiah(total_hps))
        self.lbl_kontrak_total.setText(format_rupiah(total_kontrak))
        self.lbl_kontrak_selisih.setText(format_rupiah(total_hps - total_kontrak))
    
    def load_history_data(self):
        """Load history data"""
        history = self.db.get_harga_history(self.paket_id)
        
        self.table_history.setRowCount(len(history))
        
        for row, h in enumerate(history):
            self.table_history.setItem(row, 0, QTableWidgetItem(str(h.get('tanggal', ''))[:19]))
            self.table_history.setItem(row, 1, QTableWidgetItem(h.get('tahap', '')))
            self.table_history.setItem(row, 2, QTableWidgetItem(str(h.get('item_id', '-'))))
            self.table_history.setItem(row, 3, QTableWidgetItem(format_rupiah(h.get('harga_satuan'))))
            self.table_history.setItem(row, 4, QTableWidgetItem(format_rupiah(h.get('total'))))
            self.table_history.setItem(row, 5, QTableWidgetItem(h.get('keterangan', '') or ''))
    
    def open_survey_dialog(self, item: dict):
        """Open survey dialog for item"""
        if self.paket.get('survey_locked'):
            QMessageBox.warning(self, "Terkunci", "Survey sudah dikunci!")
            return
        
        dialog = SurveyHargaItemDialog(self.paket_id, item, self)
        if dialog.exec():
            self.load_data()
            self.data_changed.emit()
    
    def calculate_hps(self):
        """Calculate HPS for all items"""
        metode = self.cmb_metode.currentData()
        overhead = self.spn_overhead.value() / 100
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT * FROM item_barang
                WHERE paket_id = ? AND is_active = 1
            """, (self.paket_id,))
            items = cursor.fetchall()
            
            for item in items:
                item = dict(item)
                
                # Calculate base price from survey
                harga_survey = hitung_harga_hps(
                    item.get('harga_survey1') or 0,
                    item.get('harga_survey2') or 0,
                    item.get('harga_survey3') or 0,
                    metode
                )
                
                # Add overhead
                harga_hps = harga_survey * (1 + overhead)
                total_hps = harga_hps * (item.get('volume') or 0)
                
                # Update
                cursor.execute("""
                    UPDATE item_barang SET
                        harga_dasar = ?,
                        harga_hps_satuan = ?,
                        total_hps = ?,
                        total = ?,
                        overhead_profit = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (harga_survey, harga_hps, total_hps, total_hps, overhead, item['id']))
                
                # Log to history
                self.db.log_harga_change(
                    self.paket_id, item['id'], 'HPS',
                    harga_hps, total_hps, f"Metode: {metode}, Overhead: {overhead*100}%"
                )
            
            # Update paket
            cursor.execute("""
                UPDATE paket SET metode_hps = ?, overhead_profit = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (metode, overhead, self.paket_id))
            
            conn.commit()
        
        self.load_data()
        self.data_changed.emit()
        QMessageBox.information(self, "Sukses", "HPS berhasil dihitung!")
    
    def lock_survey(self):
        """Lock survey stage"""
        can_proceed, missing = self.db.check_stage_requirements(self.paket_id, 'HPS')
        
        if not can_proceed:
            QMessageBox.warning(
                self, "Tidak Dapat Melanjutkan",
                "Survey belum lengkap:\n\n" + "\n".join(missing)
            )
            return
        
        reply = QMessageBox.question(
            self, "Konfirmasi",
            "Kunci survey harga?\n\n"
            "Setelah dikunci, data survey tidak dapat diubah lagi.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.db.lock_stage(self.paket_id, 'SURVEY')
            self.paket['survey_locked'] = 1
            self.load_data()
            self.data_changed.emit()
            QMessageBox.information(self, "Sukses", "Survey telah dikunci!")
    
    def lock_hps(self):
        """Lock HPS stage"""
        reply = QMessageBox.question(
            self, "Konfirmasi",
            "Tetapkan HPS sebagai final?\n\n"
            "Setelah ditetapkan, HPS tidak dapat diubah lagi.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Calculate and save total HPS to paket
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT SUM(total_hps) as total FROM item_barang
                    WHERE paket_id = ? AND is_active = 1
                """, (self.paket_id,))
                total = cursor.fetchone()['total'] or 0
                
                cursor.execute("""
                    UPDATE paket SET 
                        nilai_hps = ?,
                        nilai_hps_final = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (total, total, self.paket_id))
                conn.commit()
            
            self.db.lock_stage(self.paket_id, 'HPS')
            self.paket['hps_locked'] = 1
            self.load_data()
            self.data_changed.emit()
            QMessageBox.information(self, "Sukses", "HPS telah ditetapkan!")
    
    def input_harga_kontrak(self):
        """Input harga kontrak untuk semua item"""
        if not self.paket.get('hps_locked'):
            QMessageBox.warning(self, "Peringatan", "HPS harus ditetapkan terlebih dahulu!")
            return
        
        # Simple dialog to input kontrak prices
        QMessageBox.information(
            self, "Info",
            "Fitur input harga kontrak per item.\n\n"
            "Double-click pada kolom 'Harga Kontrak' untuk mengisi."
        )
    
    def lock_kontrak(self):
        """Lock kontrak final"""
        reply = QMessageBox.question(
            self, "Konfirmasi",
            "Tetapkan Kontrak Final?\n\n"
            "Setelah ditetapkan:\n"
            "- Harga kontrak tidak dapat diubah\n"
            "- Pembayaran menggunakan harga kontrak final",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT SUM(total_kontrak) as total FROM item_barang
                    WHERE paket_id = ? AND is_active = 1
                """, (self.paket_id,))
                total = cursor.fetchone()['total'] or 0
                
                cursor.execute("""
                    UPDATE paket SET 
                        nilai_kontrak = ?,
                        nilai_kontrak_final = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (total, total, self.paket_id))
                conn.commit()
            
            self.db.lock_stage(self.paket_id, 'KONTRAK_FINAL')
            self.paket['kontrak_final_locked'] = 1
            self.load_data()
            self.data_changed.emit()
            QMessageBox.information(self, "Sukses", "Kontrak Final telah ditetapkan!")
    
    def validate_workflow(self):
        """Validate current workflow state"""
        stages = ['SPESIFIKASI', 'SURVEY', 'HPS', 'KONTRAK_FINAL']
        results = []
        
        for stage in stages:
            can_proceed, missing = self.db.check_stage_requirements(self.paket_id, stage)
            if can_proceed:
                results.append(f"✅ {stage}: Siap")
            else:
                results.append(f"❌ {stage}:\n   " + "\n   ".join(missing))
        
        QMessageBox.information(
            self, "Validasi Workflow",
            "\n\n".join(results)
        )
    
    # =========================================================================
    # IMPORT/EXPORT EXCEL FUNCTIONS
    # =========================================================================
    
    def export_survey_excel(self):
        """Export survey template to Excel for offline filling"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Get items
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, uraian, spesifikasi, satuan, volume,
                           harga_survey1, harga_survey2, harga_survey3
                    FROM item_barang
                    WHERE paket_id = ? AND is_active = 1
                    ORDER BY nomor_urut
                """, (self.paket_id,))
                items = [dict(row) for row in cursor.fetchall()]
            
            if not items:
                QMessageBox.warning(self, "Peringatan", "Tidak ada item untuk diexport!")
                return
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Survey Harga"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:L1')
            ws['A1'] = f"FORM SURVEY HARGA - {self.paket.get('nama', '')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:L2')
            ws['A2'] = f"Kode Paket: {self.paket.get('kode', '')} | Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Instructions
            ws.merge_cells('A3:L3')
            ws['A3'] = "PETUNJUK: Isi kolom Harga Survey 1, 2, 3 dengan harga dari sumber berbeda. Jangan ubah kolom ID dan Uraian."
            ws['A3'].font = Font(italic=True, color="FF0000")
            
            # Headers
            headers = [
                "ID", "No", "Uraian", "Spesifikasi", "Satuan", "Volume",
                "Sumber 1", "Harga Survey 1", 
                "Sumber 2", "Harga Survey 2", 
                "Sumber 3", "Harga Survey 3"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            for row_idx, item in enumerate(items, 6):
                ws.cell(row=row_idx, column=1, value=item['id']).border = border
                ws.cell(row=row_idx, column=2, value=row_idx - 5).border = border
                ws.cell(row=row_idx, column=3, value=item['uraian']).border = border
                ws.cell(row=row_idx, column=4, value=item.get('spesifikasi', '')).border = border
                ws.cell(row=row_idx, column=5, value=item.get('satuan', '')).border = border
                ws.cell(row=row_idx, column=6, value=item.get('volume', 0)).border = border
                
                # Survey columns (editable)
                ws.cell(row=row_idx, column=7, value="").border = border  # Sumber 1
                ws.cell(row=row_idx, column=8, value=item.get('harga_survey1', 0) or 0).border = border
                ws.cell(row=row_idx, column=9, value="").border = border  # Sumber 2
                ws.cell(row=row_idx, column=10, value=item.get('harga_survey2', 0) or 0).border = border
                ws.cell(row=row_idx, column=11, value="").border = border  # Sumber 3
                ws.cell(row=row_idx, column=12, value=item.get('harga_survey3', 0) or 0).border = border
                
                # Format currency columns
                for c in [8, 10, 12]:
                    ws.cell(row=row_idx, column=c).number_format = '#,##0'
            
            # Column widths
            widths = [8, 5, 40, 30, 10, 10, 20, 15, 20, 15, 20, 15]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            
            # Hide ID column
            ws.column_dimensions['A'].hidden = True
            
            # Protect non-editable columns (optional visual)
            for col in [1, 2, 3, 4, 5, 6]:
                for row in range(6, len(items) + 6):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"
                    )
            
            # Save
            filepath, _ = QFileDialog.getSaveFileName(
                self, "Export Survey Excel",
                f"Survey_Harga_{self.paket.get('kode', 'PAKET')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if filepath:
                wb.save(filepath)
                QMessageBox.information(
                    self, "Sukses",
                    f"Data survey berhasil diexport ke:\n{filepath}\n\n"
                    f"Total {len(items)} item."
                )
                
                # Open file
                import subprocess
                import sys
                if sys.platform == 'win32':
                    os.startfile(filepath)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export:\n{str(e)}")
    
    def import_survey_excel(self):
        """Import survey data from Excel"""
        try:
            from openpyxl import load_workbook
            
            # Check if survey is locked
            if self.paket.get('survey_locked'):
                QMessageBox.warning(
                    self, "Peringatan",
                    "Survey sudah dikunci!\nTidak dapat mengimport data baru."
                )
                return
            
            filepath, _ = QFileDialog.getOpenFileName(
                self, "Import Survey Excel",
                "",
                "Excel Files (*.xlsx *.xls)"
            )
            
            if not filepath:
                return
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Find header row
            header_row = None
            for row in range(1, 10):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == "ID":
                    header_row = row
                    break
            
            if not header_row:
                QMessageBox.warning(
                    self, "Format Error",
                    "Format Excel tidak valid!\n"
                    "Pastikan menggunakan template dari Export Excel."
                )
                return
            
            # Read data
            updated = 0
            errors = []
            
            for row in range(header_row + 1, ws.max_row + 1):
                item_id = ws.cell(row=row, column=1).value
                
                if not item_id:
                    continue
                
                try:
                    # Get survey values
                    sumber1 = ws.cell(row=row, column=7).value or ""
                    harga1 = float(ws.cell(row=row, column=8).value or 0)
                    sumber2 = ws.cell(row=row, column=9).value or ""
                    harga2 = float(ws.cell(row=row, column=10).value or 0)
                    sumber3 = ws.cell(row=row, column=11).value or ""
                    harga3 = float(ws.cell(row=row, column=12).value or 0)
                    
                    # Calculate average
                    prices = [p for p in [harga1, harga2, harga3] if p > 0]
                    harga_rata = sum(prices) / len(prices) if prices else 0
                    
                    # Update database
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("""
                            UPDATE item_barang SET
                                harga_survey1 = ?,
                                harga_survey2 = ?,
                                harga_survey3 = ?,
                                harga_rata = ?,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE id = ? AND paket_id = ?
                        """, (
                            harga1 if harga1 > 0 else None,
                            harga2 if harga2 > 0 else None,
                            harga3 if harga3 > 0 else None,
                            harga_rata if harga_rata > 0 else None,
                            item_id, self.paket_id
                        ))
                        
                        if cursor.rowcount > 0:
                            updated += 1
                        
                        conn.commit()
                    
                    # Save survey detail if sumber provided
                    for i, (sumber, harga) in enumerate([(sumber1, harga1), (sumber2, harga2), (sumber3, harga3)], 1):
                        if harga > 0:
                            self.db.add_survey_harga_detail(self.paket_id, item_id, {
                                'sumber_ke': i,
                                'jenis_survey': 'LURING',
                                'nama_sumber': str(sumber) if sumber else f"Sumber {i}",
                                'alamat': '',
                                'harga': harga,
                            })
                    
                except Exception as e:
                    errors.append(f"Row {row}: {str(e)}")
            
            # Refresh data
            self.load_data()
            self.data_changed.emit()
            
            # Show result
            msg = f"Import selesai!\n\n✅ {updated} item berhasil diupdate"
            if errors:
                msg += f"\n\n⚠️ {len(errors)} error:\n" + "\n".join(errors[:5])
            
            QMessageBox.information(self, "Import Survey", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal import:\n{str(e)}")
    
    def export_hps_excel(self):
        """Export HPS data to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Get items with HPS
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, nomor_urut, uraian, spesifikasi, satuan, volume,
                           harga_survey1, harga_survey2, harga_survey3, harga_rata,
                           harga_hps_satuan, total_hps
                    FROM item_barang
                    WHERE paket_id = ? AND is_active = 1
                    ORDER BY nomor_urut
                """, (self.paket_id,))
                items = [dict(row) for row in cursor.fetchall()]
            
            if not items:
                QMessageBox.warning(self, "Peringatan", "Tidak ada item HPS untuk diexport!")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "HPS"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:K1')
            ws['A1'] = f"HARGA PERKIRAAN SENDIRI (HPS) - {self.paket.get('nama', '')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:K2')
            ws['A2'] = f"Metode: {self.paket.get('metode_hps', 'RATA')} | Tanggal: {datetime.now().strftime('%d/%m/%Y')}"
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = [
                "No", "Uraian", "Spesifikasi", "Satuan", "Volume",
                "Survey 1", "Survey 2", "Survey 3", "Rata-rata",
                "Harga HPS/Satuan", "Total HPS"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Data
            subtotal = 0
            for row_idx, item in enumerate(items, 5):
                ws.cell(row=row_idx, column=1, value=row_idx - 4).border = border
                ws.cell(row=row_idx, column=2, value=item['uraian']).border = border
                ws.cell(row=row_idx, column=3, value=item.get('spesifikasi', '')).border = border
                ws.cell(row=row_idx, column=4, value=item.get('satuan', '')).border = border
                ws.cell(row=row_idx, column=5, value=item.get('volume', 0)).border = border
                
                for c, field in enumerate(['harga_survey1', 'harga_survey2', 'harga_survey3', 'harga_rata'], 6):
                    cell = ws.cell(row=row_idx, column=c, value=item.get(field, 0) or 0)
                    cell.number_format = '#,##0'
                    cell.border = border
                
                ws.cell(row=row_idx, column=10, value=item.get('harga_hps_satuan', 0) or 0).number_format = '#,##0'
                ws.cell(row=row_idx, column=10).border = border
                
                total_hps = item.get('total_hps', 0) or 0
                ws.cell(row=row_idx, column=11, value=total_hps).number_format = '#,##0'
                ws.cell(row=row_idx, column=11).border = border
                
                subtotal += total_hps
            
            # Summary
            last_row = len(items) + 5
            ws.merge_cells(f'A{last_row}:J{last_row}')
            ws[f'A{last_row}'] = 'SUBTOTAL'
            ws[f'A{last_row}'].font = Font(bold=True)
            ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'K{last_row}'] = subtotal
            ws[f'K{last_row}'].number_format = '#,##0'
            ws[f'K{last_row}'].font = Font(bold=True)
            
            ppn = subtotal * 0.11
            last_row += 1
            ws.merge_cells(f'A{last_row}:J{last_row}')
            ws[f'A{last_row}'] = 'PPN 11%'
            ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'K{last_row}'] = ppn
            ws[f'K{last_row}'].number_format = '#,##0'
            
            last_row += 1
            ws.merge_cells(f'A{last_row}:J{last_row}')
            ws[f'A{last_row}'] = 'TOTAL HPS'
            ws[f'A{last_row}'].font = Font(bold=True, size=12)
            ws[f'A{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'K{last_row}'] = subtotal + ppn
            ws[f'K{last_row}'].number_format = '#,##0'
            ws[f'K{last_row}'].font = Font(bold=True, size=12)
            
            # Column widths
            widths = [5, 35, 25, 8, 8, 12, 12, 12, 12, 15, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            
            # Save
            filepath, _ = QFileDialog.getSaveFileName(
                self, "Export HPS Excel",
                f"HPS_{self.paket.get('kode', 'PAKET')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if filepath:
                wb.save(filepath)
                QMessageBox.information(self, "Sukses", f"HPS berhasil diexport ke:\n{filepath}")
                
                import subprocess
                import sys
                if sys.platform == 'win32':
                    os.startfile(filepath)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export HPS:\n{str(e)}")
    
    def export_kontrak_excel(self):
        """Export kontrak prices to Excel for filling"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Check if HPS is locked
            if not self.paket.get('hps_locked'):
                QMessageBox.warning(self, "Peringatan", "HPS harus ditetapkan terlebih dahulu!")
                return
            
            # Get items
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, nomor_urut, uraian, spesifikasi, satuan, volume,
                           harga_hps_satuan, total_hps, harga_kontrak_satuan, total_kontrak
                    FROM item_barang
                    WHERE paket_id = ? AND is_active = 1
                    ORDER BY nomor_urut
                """, (self.paket_id,))
                items = [dict(row) for row in cursor.fetchall()]
            
            if not items:
                QMessageBox.warning(self, "Peringatan", "Tidak ada item!")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Harga Kontrak"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:I1')
            ws['A1'] = f"FORM HARGA KONTRAK - {self.paket.get('nama', '')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:I2')
            ws['A2'] = "PETUNJUK: Isi kolom 'Harga Kontrak/Satuan' hasil negosiasi. Harga tidak boleh melebihi HPS."
            ws['A2'].font = Font(italic=True, color="FF0000")
            
            # Headers
            headers = [
                "ID", "No", "Uraian", "Satuan", "Volume",
                "Harga HPS/Satuan", "Total HPS", 
                "Harga Kontrak/Satuan", "Total Kontrak"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Data
            for row_idx, item in enumerate(items, 5):
                ws.cell(row=row_idx, column=1, value=item['id']).border = border
                ws.cell(row=row_idx, column=2, value=row_idx - 4).border = border
                ws.cell(row=row_idx, column=3, value=item['uraian']).border = border
                ws.cell(row=row_idx, column=4, value=item.get('satuan', '')).border = border
                ws.cell(row=row_idx, column=5, value=item.get('volume', 0)).border = border
                
                hps_satuan = item.get('harga_hps_satuan', 0) or 0
                ws.cell(row=row_idx, column=6, value=hps_satuan).number_format = '#,##0'
                ws.cell(row=row_idx, column=6).border = border
                ws.cell(row=row_idx, column=6).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
                
                ws.cell(row=row_idx, column=7, value=item.get('total_hps', 0) or 0).number_format = '#,##0'
                ws.cell(row=row_idx, column=7).border = border
                ws.cell(row=row_idx, column=7).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
                
                # Editable kontrak columns
                kontrak_satuan = item.get('harga_kontrak_satuan', 0) or hps_satuan
                ws.cell(row=row_idx, column=8, value=kontrak_satuan).number_format = '#,##0'
                ws.cell(row=row_idx, column=8).border = border
                ws.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                
                # Formula for total kontrak
                ws.cell(row=row_idx, column=9, value=f"=E{row_idx}*H{row_idx}")
                ws.cell(row=row_idx, column=9).number_format = '#,##0'
                ws.cell(row=row_idx, column=9).border = border
            
            # Hide ID column
            ws.column_dimensions['A'].hidden = True
            
            # Column widths
            widths = [8, 5, 40, 10, 10, 15, 18, 15, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            
            # Save
            filepath, _ = QFileDialog.getSaveFileName(
                self, "Export Form Harga Kontrak",
                f"Form_Kontrak_{self.paket.get('kode', 'PAKET')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if filepath:
                wb.save(filepath)
                QMessageBox.information(self, "Sukses", f"Form kontrak berhasil diexport ke:\n{filepath}")
                
                import subprocess
                import sys
                if sys.platform == 'win32':
                    os.startfile(filepath)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export form kontrak:\n{str(e)}")
    
    def import_kontrak_excel(self):
        """Import kontrak prices from Excel"""
        try:
            from openpyxl import load_workbook
            
            # Check if HPS is locked
            if not self.paket.get('hps_locked'):
                QMessageBox.warning(self, "Peringatan", "HPS harus ditetapkan terlebih dahulu!")
                return
            
            # Check if kontrak is already locked
            if self.paket.get('kontrak_final_locked'):
                QMessageBox.warning(
                    self, "Peringatan",
                    "Kontrak Final sudah dikunci!\nNilai kontrak tidak dapat diubah lagi."
                )
                return
            
            filepath, _ = QFileDialog.getOpenFileName(
                self, "Import Harga Kontrak",
                "",
                "Excel Files (*.xlsx *.xls)"
            )
            
            if not filepath:
                return
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Find header row
            header_row = None
            for row in range(1, 10):
                if ws.cell(row=row, column=1).value == "ID":
                    header_row = row
                    break
            
            if not header_row:
                QMessageBox.warning(
                    self, "Format Error",
                    "Format Excel tidak valid!\n"
                    "Pastikan menggunakan template dari Export Excel."
                )
                return
            
            updated = 0
            errors = []
            
            for row in range(header_row + 1, ws.max_row + 1):
                item_id = ws.cell(row=row, column=1).value
                
                if not item_id:
                    continue
                
                try:
                    harga_kontrak = float(ws.cell(row=row, column=8).value or 0)
                    volume = float(ws.cell(row=row, column=5).value or 0)
                    
                    if harga_kontrak > 0:
                        total_kontrak = harga_kontrak * volume
                        
                        with self.db.get_connection() as conn:
                            cursor = conn.cursor()
                            
                            # Get HPS for validation
                            cursor.execute("""
                                SELECT harga_hps_satuan FROM item_barang
                                WHERE id = ? AND paket_id = ?
                            """, (item_id, self.paket_id))
                            
                            item = cursor.fetchone()
                            if item:
                                hps = item['harga_hps_satuan'] or 0
                                
                                # Validate: kontrak must not exceed HPS
                                if harga_kontrak > hps:
                                    errors.append(f"ID {item_id}: Harga kontrak melebihi HPS!")
                                    continue
                                
                                selisih = hps - harga_kontrak
                                
                                cursor.execute("""
                                    UPDATE item_barang SET
                                        harga_kontrak_satuan = ?,
                                        total_kontrak = ?,
                                        selisih_harga = ?,
                                        selisih_total = ?,
                                        updated_at = CURRENT_TIMESTAMP
                                    WHERE id = ? AND paket_id = ?
                                """, (
                                    harga_kontrak, total_kontrak,
                                    selisih, selisih * volume,
                                    item_id, self.paket_id
                                ))
                                
                                if cursor.rowcount > 0:
                                    updated += 1
                            
                            conn.commit()
                    
                except Exception as e:
                    errors.append(f"Row {row}: {str(e)}")
            
            # Refresh
            self.load_data()
            self.data_changed.emit()
            
            msg = f"Import selesai!\n\n✅ {updated} item berhasil diupdate"
            if errors:
                msg += f"\n\n⚠️ {len(errors)} error:\n" + "\n".join(errors[:5])
            
            QMessageBox.information(self, "Import Kontrak", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal import:\n{str(e)}")


# ============================================================================
# HELPER FUNCTION
# ============================================================================

def show_harga_lifecycle_manager(paket_id: int, parent=None) -> HargaLifecycleManager:
    """Show harga lifecycle manager dialog"""
    dialog = HargaLifecycleManager(paket_id, parent)
    dialog.exec()
    return dialog
