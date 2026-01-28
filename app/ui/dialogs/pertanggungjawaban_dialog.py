"""
PPK DOCUMENT FACTORY - Pertanggungjawaban Dialogs
==================================================
Dialog untuk:
1. Rekap Bukti Pengeluaran - input nota/bukti dengan opsi scanner
2. Perhitungan Tambah/Kurang - tabel permintaan vs realisasi
"""

import os
import shutil
from datetime import datetime
from typing import Dict, List, Any, Optional
from decimal import Decimal

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout,
    QPushButton, QLabel, QFrame, QLineEdit, QTextEdit,
    QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit,
    QGroupBox, QScrollArea, QWidget, QMessageBox,
    QFileDialog, QProgressBar, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QSplitter, QTabWidget
)
from PySide6.QtCore import Qt, Signal, QDate
from PySide6.QtGui import QFont, QColor, QPixmap


# =============================================================================
# REKAP BUKTI PENGELUARAN DIALOG
# =============================================================================

class RekapBuktiPengeluaranDialog(QDialog):
    """
    Dialog untuk merekap bukti pengeluaran (nota/faktur).
    Fitur:
    - Input manual bukti pengeluaran
    - Upload scan nota/faktur
    - Koneksi scanner (jika tersedia)
    - Rekap total pengeluaran
    """

    bukti_saved = Signal(list)  # List of bukti items

    def __init__(self, transaksi: Dict[str, Any], parent=None):
        super().__init__(parent)
        self.transaksi = transaksi
        self.bukti_list = []
        self.uploaded_files = []

        self._setup_ui()
        self._load_existing_bukti()

    def _setup_ui(self):
        self.setWindowTitle("Rekap Bukti Pengeluaran")
        self.setMinimumSize(900, 650)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Header
        header = QLabel("REKAP BUKTI PENGELUARAN")
        header.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
            background-color: #ecf0f1;
            border-radius: 5px;
        """)
        layout.addWidget(header)

        # Info transaksi
        info_group = QGroupBox("Informasi Transaksi")
        info_layout = QGridLayout(info_group)
        info_layout.addWidget(QLabel("Kegiatan:"), 0, 0)
        info_layout.addWidget(QLabel(self.transaksi.get('nama_kegiatan', '-')), 0, 1)
        info_layout.addWidget(QLabel("Uang Muka:"), 0, 2)
        uang_muka = self.transaksi.get('uang_muka', 0) or self.transaksi.get('estimasi_biaya', 0)
        info_layout.addWidget(QLabel(f"Rp {uang_muka:,.0f}".replace(",", ".")), 0, 3)
        layout.addWidget(info_group)

        # Splitter for input and list
        splitter = QSplitter(Qt.Horizontal)

        # Left: Input bukti
        input_panel = self._create_input_panel()
        splitter.addWidget(input_panel)

        # Right: List bukti
        list_panel = self._create_list_panel()
        splitter.addWidget(list_panel)

        splitter.setSizes([400, 500])
        layout.addWidget(splitter)

        # Summary
        summary_group = self._create_summary_panel()
        layout.addWidget(summary_group)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_export = QPushButton("Export ke Excel")
        btn_export.setStyleSheet("background-color: #17a2b8;")
        btn_export.clicked.connect(self._export_to_excel)
        btn_layout.addWidget(btn_export)

        btn_save = QPushButton("Simpan")
        btn_save.setStyleSheet("background-color: #27ae60;")
        btn_save.clicked.connect(self._save_and_close)
        btn_layout.addWidget(btn_save)

        btn_close = QPushButton("Tutup")
        btn_close.clicked.connect(self.close)
        btn_layout.addWidget(btn_close)

        layout.addLayout(btn_layout)

    def _create_input_panel(self) -> QWidget:
        group = QGroupBox("Input Bukti Pengeluaran")
        layout = QVBoxLayout(group)

        # Form input
        form_layout = QGridLayout()

        form_layout.addWidget(QLabel("No. Bukti/Nota:"), 0, 0)
        self.nomor_bukti_edit = QLineEdit()
        self.nomor_bukti_edit.setPlaceholderText("Nomor nota/faktur")
        form_layout.addWidget(self.nomor_bukti_edit, 0, 1)

        form_layout.addWidget(QLabel("Tanggal:"), 1, 0)
        self.tanggal_edit = QDateEdit()
        self.tanggal_edit.setDate(QDate.currentDate())
        self.tanggal_edit.setCalendarPopup(True)
        form_layout.addWidget(self.tanggal_edit, 1, 1)

        form_layout.addWidget(QLabel("Uraian:"), 2, 0)
        self.uraian_edit = QLineEdit()
        self.uraian_edit.setPlaceholderText("Uraian barang/jasa")
        form_layout.addWidget(self.uraian_edit, 2, 1)

        form_layout.addWidget(QLabel("Volume:"), 3, 0)
        vol_layout = QHBoxLayout()
        self.volume_spin = QSpinBox()
        self.volume_spin.setRange(1, 9999)
        self.volume_spin.setValue(1)
        vol_layout.addWidget(self.volume_spin)
        self.satuan_combo = QComboBox()
        self.satuan_combo.setEditable(True)
        self.satuan_combo.addItems(["paket", "unit", "buah", "lembar", "orang", "set", "rim", "dus"])
        vol_layout.addWidget(self.satuan_combo)
        form_layout.addLayout(vol_layout, 3, 1)

        form_layout.addWidget(QLabel("Harga Satuan:"), 4, 0)
        self.harga_spin = QDoubleSpinBox()
        self.harga_spin.setRange(0, 999999999)
        self.harga_spin.setDecimals(0)
        self.harga_spin.setPrefix("Rp ")
        self.harga_spin.setGroupSeparatorShown(True)
        form_layout.addWidget(self.harga_spin, 4, 1)

        form_layout.addWidget(QLabel("Jumlah:"), 5, 0)
        self.jumlah_label = QLabel("Rp 0")
        self.jumlah_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        form_layout.addWidget(self.jumlah_label, 5, 1)

        # Connect for auto-calculate
        self.volume_spin.valueChanged.connect(self._calculate_jumlah)
        self.harga_spin.valueChanged.connect(self._calculate_jumlah)

        layout.addLayout(form_layout)

        # File upload section
        file_group = QGroupBox("Upload Scan Bukti")
        file_layout = QVBoxLayout(file_group)

        self.file_label = QLabel("Belum ada file dipilih")
        self.file_label.setStyleSheet("color: #7f8c8d;")
        file_layout.addWidget(self.file_label)

        btn_file_layout = QHBoxLayout()

        btn_browse = QPushButton("Pilih File")
        btn_browse.clicked.connect(self._browse_file)
        btn_file_layout.addWidget(btn_browse)

        btn_scan = QPushButton("Scan Langsung")
        btn_scan.setStyleSheet("background-color: #9b59b6;")
        btn_scan.clicked.connect(self._scan_document)
        btn_file_layout.addWidget(btn_scan)

        file_layout.addLayout(btn_file_layout)
        layout.addWidget(file_group)

        # Preview image
        self.preview_label = QLabel()
        self.preview_label.setFixedHeight(100)
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.preview_label.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dee2e6;")
        layout.addWidget(self.preview_label)

        # Add button
        btn_add = QPushButton("+ Tambah Bukti")
        btn_add.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 10px;
                font-weight: bold;
            }
        """)
        btn_add.clicked.connect(self._add_bukti)
        layout.addWidget(btn_add)

        layout.addStretch()
        return group

    def _create_list_panel(self) -> QWidget:
        group = QGroupBox("Daftar Bukti Pengeluaran")
        layout = QVBoxLayout(group)

        # Table
        self.bukti_table = QTableWidget()
        self.bukti_table.setColumnCount(7)
        self.bukti_table.setHorizontalHeaderLabels([
            "No", "No. Bukti", "Tanggal", "Uraian", "Vol x Harga", "Jumlah", "Aksi"
        ])
        self.bukti_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.bukti_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.bukti_table.setAlternatingRowColors(True)
        layout.addWidget(self.bukti_table)

        return group

    def _create_summary_panel(self) -> QWidget:
        group = QGroupBox("Ringkasan")
        layout = QHBoxLayout(group)

        layout.addWidget(QLabel("Jumlah Bukti:"))
        self.count_label = QLabel("0")
        self.count_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.count_label)

        layout.addStretch()

        layout.addWidget(QLabel("TOTAL PENGELUARAN:"))
        self.total_label = QLabel("Rp 0")
        self.total_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #e74c3c;")
        layout.addWidget(self.total_label)

        return group

    def _calculate_jumlah(self):
        volume = self.volume_spin.value()
        harga = self.harga_spin.value()
        jumlah = volume * harga
        self.jumlah_label.setText(f"Rp {jumlah:,.0f}".replace(",", "."))

    def _browse_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Pilih Scan Bukti",
            "",
            "Images (*.jpg *.jpeg *.png);;PDF (*.pdf);;All Files (*)"
        )
        if filepath:
            self.selected_file = filepath
            self.file_label.setText(os.path.basename(filepath))
            self.file_label.setStyleSheet("color: #27ae60;")

            # Show preview for images
            if filepath.lower().endswith(('.jpg', '.jpeg', '.png')):
                pixmap = QPixmap(filepath)
                if not pixmap.isNull():
                    scaled = pixmap.scaled(200, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.preview_label.setPixmap(scaled)

    def _scan_document(self):
        """
        Scan document using connected scanner.
        Requires python-sane (Linux) or WIA (Windows).
        """
        try:
            import platform

            if platform.system() == 'Linux':
                # Try using SANE
                try:
                    import sane
                    sane.init()
                    devices = sane.get_devices()
                    if devices:
                        scanner = sane.open(devices[0][0])
                        scanner.mode = 'color'
                        scanner.resolution = 200

                        # Scan
                        scanner.start()
                        img = scanner.snap()

                        # Save to temp file
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filepath = f"/tmp/scan_{timestamp}.jpg"
                        img.save(filepath)

                        self.selected_file = filepath
                        self.file_label.setText(f"Scan berhasil: {os.path.basename(filepath)}")
                        self.file_label.setStyleSheet("color: #27ae60;")

                        pixmap = QPixmap(filepath)
                        if not pixmap.isNull():
                            scaled = pixmap.scaled(200, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            self.preview_label.setPixmap(scaled)

                        scanner.close()
                    else:
                        QMessageBox.warning(self, "Scanner", "Tidak ada scanner yang terdeteksi!")
                except ImportError:
                    QMessageBox.information(
                        self, "Scanner",
                        "Fitur scan memerlukan library python-sane.\n"
                        "Install dengan: pip install python-sane\n\n"
                        "Atau gunakan 'Pilih File' untuk upload scan manual."
                    )
            elif platform.system() == 'Windows':
                QMessageBox.information(
                    self, "Scanner",
                    "Untuk Windows, gunakan aplikasi scanner bawaan Windows\n"
                    "kemudian upload hasil scan menggunakan 'Pilih File'.\n\n"
                    "Integrasi scanner Windows sedang dalam pengembangan."
                )
            else:
                QMessageBox.information(
                    self, "Scanner",
                    "Fitur scan belum didukung untuk sistem operasi ini.\n"
                    "Gunakan 'Pilih File' untuk upload scan manual."
                )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal scan: {str(e)}")

    def _add_bukti(self):
        uraian = self.uraian_edit.text().strip()
        if not uraian:
            QMessageBox.warning(self, "Peringatan", "Uraian tidak boleh kosong!")
            return

        nomor = self.nomor_bukti_edit.text().strip() or f"BKT-{len(self.bukti_list)+1:03d}"
        tanggal = self.tanggal_edit.date().toString("yyyy-MM-dd")
        volume = self.volume_spin.value()
        satuan = self.satuan_combo.currentText()
        harga = self.harga_spin.value()
        jumlah = volume * harga

        # Handle file
        file_path = getattr(self, 'selected_file', None)
        saved_file = None
        if file_path and os.path.exists(file_path):
            # Copy file to output folder
            try:
                from app.services.dokumen_generator import get_dokumen_generator
                generator = get_dokumen_generator()
                output_folder = generator.get_output_folder(transaksi=self.transaksi)
                bukti_folder = output_folder / "bukti_pengeluaran"
                bukti_folder.mkdir(parents=True, exist_ok=True)

                ext = os.path.splitext(file_path)[1]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                new_name = f"BUKTI_{len(self.bukti_list)+1:03d}_{timestamp}{ext}"
                dest_path = bukti_folder / new_name
                shutil.copy2(file_path, dest_path)
                saved_file = str(dest_path)
            except Exception as e:
                print(f"Error copying file: {e}")

        bukti = {
            'nomor': nomor,
            'tanggal': tanggal,
            'uraian': uraian,
            'volume': volume,
            'satuan': satuan,
            'harga_satuan': harga,
            'jumlah': jumlah,
            'file_path': saved_file,
        }
        self.bukti_list.append(bukti)

        self._refresh_table()
        self._clear_form()

    def _refresh_table(self):
        self.bukti_table.setRowCount(len(self.bukti_list))

        total = 0
        for i, bukti in enumerate(self.bukti_list):
            self.bukti_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.bukti_table.setItem(i, 1, QTableWidgetItem(bukti['nomor']))
            self.bukti_table.setItem(i, 2, QTableWidgetItem(bukti['tanggal']))
            self.bukti_table.setItem(i, 3, QTableWidgetItem(bukti['uraian']))
            self.bukti_table.setItem(i, 4, QTableWidgetItem(
                f"{bukti['volume']} {bukti['satuan']} x Rp {bukti['harga_satuan']:,.0f}".replace(",", ".")
            ))
            self.bukti_table.setItem(i, 5, QTableWidgetItem(
                f"Rp {bukti['jumlah']:,.0f}".replace(",", ".")
            ))

            # Delete button
            btn_del = QPushButton("Hapus")
            btn_del.setStyleSheet("background-color: #e74c3c; color: white;")
            btn_del.clicked.connect(lambda checked, idx=i: self._delete_bukti(idx))
            self.bukti_table.setCellWidget(i, 6, btn_del)

            total += bukti['jumlah']

        self.count_label.setText(str(len(self.bukti_list)))
        self.total_label.setText(f"Rp {total:,.0f}".replace(",", "."))

    def _delete_bukti(self, index: int):
        if 0 <= index < len(self.bukti_list):
            self.bukti_list.pop(index)
            self._refresh_table()

    def _clear_form(self):
        self.nomor_bukti_edit.clear()
        self.uraian_edit.clear()
        self.volume_spin.setValue(1)
        self.harga_spin.setValue(0)
        self.file_label.setText("Belum ada file dipilih")
        self.file_label.setStyleSheet("color: #7f8c8d;")
        self.preview_label.clear()
        self.selected_file = None

    def _load_existing_bukti(self):
        """Load existing bukti from transaksi if available."""
        existing = self.transaksi.get('bukti_pengeluaran', [])
        if existing:
            self.bukti_list = existing
            self._refresh_table()

    def _export_to_excel(self):
        """Export bukti list to Excel."""
        if not self.bukti_list:
            QMessageBox.warning(self, "Peringatan", "Tidak ada bukti untuk diekspor!")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Export ke Excel",
            f"Rekap_Bukti_{self.transaksi.get('kode', 'transaksi')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if filepath:
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Rekap Bukti"

                # Header
                ws['A1'] = "REKAP BUKTI PENGELUARAN"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:F1')

                ws['A2'] = f"Kegiatan: {self.transaksi.get('nama_kegiatan', '-')}"
                ws.merge_cells('A2:F2')

                # Table headers
                headers = ['No', 'No. Bukti', 'Tanggal', 'Uraian', 'Volume', 'Harga Satuan', 'Jumlah']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)

                # Data
                total = 0
                for row, bukti in enumerate(self.bukti_list, 5):
                    ws.cell(row=row, column=1, value=row - 4)
                    ws.cell(row=row, column=2, value=bukti['nomor'])
                    ws.cell(row=row, column=3, value=bukti['tanggal'])
                    ws.cell(row=row, column=4, value=bukti['uraian'])
                    ws.cell(row=row, column=5, value=f"{bukti['volume']} {bukti['satuan']}")
                    ws.cell(row=row, column=6, value=bukti['harga_satuan'])
                    ws.cell(row=row, column=7, value=bukti['jumlah'])
                    total += bukti['jumlah']

                # Total row
                total_row = 5 + len(self.bukti_list)
                ws.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
                ws.cell(row=total_row, column=7, value=total).font = Font(bold=True)

                wb.save(filepath)
                QMessageBox.information(self, "Sukses", f"Berhasil export ke:\n{filepath}")

            except ImportError:
                QMessageBox.warning(self, "Error", "Library openpyxl tidak tersedia!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal export: {str(e)}")

    def _save_and_close(self):
        self.bukti_saved.emit(self.bukti_list)
        self.accept()

    def get_bukti_list(self) -> List[Dict]:
        return self.bukti_list

    def get_total(self) -> float:
        return sum(b['jumlah'] for b in self.bukti_list)


# =============================================================================
# PERHITUNGAN TAMBAH KURANG DIALOG
# =============================================================================

class PerhitunganTambahKurangDialog(QDialog):
    """
    Dialog untuk menghitung selisih antara:
    - Uang Muka (permintaan awal)
    - Realisasi (bukti pengeluaran)

    Hasil:
    - Kurang Bayar: perlu tambahan pembayaran
    - Lebih Bayar: perlu pengembalian
    - Nihil: pas, tidak ada selisih
    """

    calculation_saved = Signal(dict)  # Result dictionary

    def __init__(self, transaksi: Dict[str, Any], bukti_list: List[Dict] = None, parent=None):
        super().__init__(parent)
        self.transaksi = transaksi
        self.bukti_list = bukti_list or []
        self.permintaan_list = []

        self._setup_ui()
        self._load_data()

    def _setup_ui(self):
        self.setWindowTitle("Perhitungan Tambah/Kurang")
        self.setMinimumSize(1000, 700)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Header
        header = QLabel("PERHITUNGAN TAMBAH / KURANG")
        header.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
            background-color: #ecf0f1;
            border-radius: 5px;
        """)
        layout.addWidget(header)

        # Info
        info_group = QGroupBox("Informasi")
        info_layout = QGridLayout(info_group)
        info_layout.addWidget(QLabel("Kegiatan:"), 0, 0)
        info_layout.addWidget(QLabel(self.transaksi.get('nama_kegiatan', '-')), 0, 1)
        info_layout.addWidget(QLabel("Mekanisme:"), 0, 2)
        info_layout.addWidget(QLabel(self.transaksi.get('mekanisme', '-')), 0, 3)
        layout.addWidget(info_group)

        # Main content - two tables side by side
        splitter = QSplitter(Qt.Horizontal)

        # Left: Permintaan (Uang Muka)
        left_panel = self._create_permintaan_panel()
        splitter.addWidget(left_panel)

        # Right: Realisasi (Bukti Pengeluaran)
        right_panel = self._create_realisasi_panel()
        splitter.addWidget(right_panel)

        splitter.setSizes([500, 500])
        layout.addWidget(splitter)

        # Calculation result
        result_group = self._create_result_panel()
        layout.addWidget(result_group)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_export = QPushButton("Export ke Excel")
        btn_export.setStyleSheet("background-color: #17a2b8;")
        btn_export.clicked.connect(self._export_to_excel)
        btn_layout.addWidget(btn_export)

        btn_save = QPushButton("Simpan Perhitungan")
        btn_save.setStyleSheet("background-color: #27ae60;")
        btn_save.clicked.connect(self._save_and_close)
        btn_layout.addWidget(btn_save)

        btn_close = QPushButton("Tutup")
        btn_close.clicked.connect(self.close)
        btn_layout.addWidget(btn_close)

        layout.addLayout(btn_layout)

    def _create_permintaan_panel(self) -> QWidget:
        group = QGroupBox("PERMINTAAN (Uang Muka)")
        group.setStyleSheet("QGroupBox { font-weight: bold; color: #3498db; }")
        layout = QVBoxLayout(group)

        # Table
        self.permintaan_table = QTableWidget()
        self.permintaan_table.setColumnCount(5)
        self.permintaan_table.setHorizontalHeaderLabels([
            "No", "Uraian", "Volume", "Harga", "Jumlah"
        ])
        self.permintaan_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.permintaan_table.setAlternatingRowColors(True)
        layout.addWidget(self.permintaan_table)

        # Total
        total_layout = QHBoxLayout()
        total_layout.addStretch()
        total_layout.addWidget(QLabel("TOTAL PERMINTAAN:"))
        self.permintaan_total_label = QLabel("Rp 0")
        self.permintaan_total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #3498db;")
        total_layout.addWidget(self.permintaan_total_label)
        layout.addLayout(total_layout)

        return group

    def _create_realisasi_panel(self) -> QWidget:
        group = QGroupBox("REALISASI (Bukti Pengeluaran)")
        group.setStyleSheet("QGroupBox { font-weight: bold; color: #e74c3c; }")
        layout = QVBoxLayout(group)

        # Table
        self.realisasi_table = QTableWidget()
        self.realisasi_table.setColumnCount(5)
        self.realisasi_table.setHorizontalHeaderLabels([
            "No", "Uraian", "Volume", "Harga", "Jumlah"
        ])
        self.realisasi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.realisasi_table.setAlternatingRowColors(True)
        layout.addWidget(self.realisasi_table)

        # Total
        total_layout = QHBoxLayout()
        total_layout.addStretch()
        total_layout.addWidget(QLabel("TOTAL REALISASI:"))
        self.realisasi_total_label = QLabel("Rp 0")
        self.realisasi_total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #e74c3c;")
        total_layout.addWidget(self.realisasi_total_label)
        layout.addLayout(total_layout)

        return group

    def _create_result_panel(self) -> QWidget:
        group = QGroupBox("HASIL PERHITUNGAN")
        layout = QGridLayout(group)

        # Row 1: Values
        layout.addWidget(QLabel("Uang Muka (A):"), 0, 0)
        self.val_uang_muka = QLabel("Rp 0")
        self.val_uang_muka.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.val_uang_muka, 0, 1)

        layout.addWidget(QLabel("Realisasi (B):"), 0, 2)
        self.val_realisasi = QLabel("Rp 0")
        self.val_realisasi.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.val_realisasi, 0, 3)

        layout.addWidget(QLabel("Selisih (B - A):"), 0, 4)
        self.val_selisih = QLabel("Rp 0")
        self.val_selisih.setStyleSheet("font-weight: bold; font-size: 16px;")
        layout.addWidget(self.val_selisih, 0, 5)

        # Row 2: Result status
        self.result_frame = QFrame()
        self.result_frame.setMinimumHeight(60)
        self.result_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 2px solid #dee2e6;
                border-radius: 8px;
            }
        """)
        result_inner = QHBoxLayout(self.result_frame)

        self.result_icon = QLabel()
        self.result_icon.setFixedSize(40, 40)
        result_inner.addWidget(self.result_icon)

        self.result_label = QLabel("Belum dihitung")
        self.result_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        result_inner.addWidget(self.result_label)

        self.result_action = QLabel("")
        self.result_action.setStyleSheet("font-size: 12px; color: #6c757d;")
        result_inner.addWidget(self.result_action)

        result_inner.addStretch()

        layout.addWidget(self.result_frame, 1, 0, 1, 6)

        return group

    def _load_data(self):
        """
        Load permintaan and realisasi data.

        Document Flow:
        - PERMINTAAN = From LBR_REQ (Lembar Permintaan) - uang muka yang diterima
        - REALISASI = From REKAP_BKT (Rekap Bukti Pengeluaran) - pengeluaran aktual
        """
        try:
            transaksi_id = self.transaksi.get('id')

            if transaksi_id:
                from app.models.pencairan_models import PencairanManager
                manager = PencairanManager()

                # Load PERMINTAAN from LBR_REQ (Lembar Permintaan)
                lbr_items = manager.get_rincian_items(transaksi_id, 'LBR_REQ')
                if lbr_items:
                    self.permintaan_list = []
                    for item in lbr_items:
                        self.permintaan_list.append({
                            'uraian': item.get('uraian', ''),
                            'volume': item.get('volume', 1),
                            'satuan': item.get('satuan', 'paket'),
                            'harga_satuan': item.get('harga_satuan', 0),
                            'jumlah': item.get('jumlah', 0),
                        })
                    print(f"Loaded {len(lbr_items)} permintaan items from LBR_REQ")

                # Load REALISASI from REKAP_BKT (Rekap Bukti Pengeluaran)
                # Only if bukti_list is empty (not passed from RekapBuktiDialog)
                if not self.bukti_list:
                    rekap_items = manager.get_rincian_items(transaksi_id, 'REKAP_BKT')
                    if rekap_items:
                        self.bukti_list = []
                        for item in rekap_items:
                            self.bukti_list.append({
                                'uraian': item.get('uraian', ''),
                                'volume': item.get('volume', 1),
                                'satuan': item.get('satuan', 'paket'),
                                'harga_satuan': item.get('harga_satuan', 0),
                                'jumlah': item.get('jumlah', 0),
                            })
                        print(f"Loaded {len(rekap_items)} realisasi items from REKAP_BKT")

        except Exception as e:
            print(f"Error loading data from database: {e}")

        # Fallback: Load permintaan from transaksi rincian or estimasi
        if not self.permintaan_list:
            rincian = self.transaksi.get('rincian_items', [])
            if rincian:
                self.permintaan_list = rincian
            else:
                # Use single item from estimasi
                estimasi = self.transaksi.get('estimasi_biaya', 0) or self.transaksi.get('uang_muka', 0)
                if estimasi:
                    self.permintaan_list = [{
                        'uraian': self.transaksi.get('nama_kegiatan', 'Kegiatan'),
                        'volume': 1,
                        'satuan': 'paket',
                        'harga_satuan': estimasi,
                        'jumlah': estimasi,
                    }]

        self._refresh_tables()
        self._calculate()

    def _refresh_tables(self):
        """Refresh both tables."""
        # Permintaan table
        self.permintaan_table.setRowCount(len(self.permintaan_list))
        total_permintaan = 0
        for i, item in enumerate(self.permintaan_list):
            self.permintaan_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.permintaan_table.setItem(i, 1, QTableWidgetItem(item.get('uraian', '')))
            self.permintaan_table.setItem(i, 2, QTableWidgetItem(
                f"{item.get('volume', 1)} {item.get('satuan', '')}"
            ))
            self.permintaan_table.setItem(i, 3, QTableWidgetItem(
                f"Rp {item.get('harga_satuan', 0):,.0f}".replace(",", ".")
            ))
            jumlah = item.get('jumlah', 0)
            self.permintaan_table.setItem(i, 4, QTableWidgetItem(
                f"Rp {jumlah:,.0f}".replace(",", ".")
            ))
            total_permintaan += jumlah

        self.permintaan_total_label.setText(f"Rp {total_permintaan:,.0f}".replace(",", "."))

        # Realisasi table
        self.realisasi_table.setRowCount(len(self.bukti_list))
        total_realisasi = 0
        for i, item in enumerate(self.bukti_list):
            self.realisasi_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.realisasi_table.setItem(i, 1, QTableWidgetItem(item.get('uraian', '')))
            self.realisasi_table.setItem(i, 2, QTableWidgetItem(
                f"{item.get('volume', 1)} {item.get('satuan', '')}"
            ))
            self.realisasi_table.setItem(i, 3, QTableWidgetItem(
                f"Rp {item.get('harga_satuan', 0):,.0f}".replace(",", ".")
            ))
            jumlah = item.get('jumlah', 0)
            self.realisasi_table.setItem(i, 4, QTableWidgetItem(
                f"Rp {jumlah:,.0f}".replace(",", ".")
            ))
            total_realisasi += jumlah

        self.realisasi_total_label.setText(f"Rp {total_realisasi:,.0f}".replace(",", "."))

    def _calculate(self):
        """Calculate the difference."""
        total_permintaan = sum(item.get('jumlah', 0) for item in self.permintaan_list)
        total_realisasi = sum(item.get('jumlah', 0) for item in self.bukti_list)
        selisih = total_realisasi - total_permintaan

        # Update display
        self.val_uang_muka.setText(f"Rp {total_permintaan:,.0f}".replace(",", "."))
        self.val_realisasi.setText(f"Rp {total_realisasi:,.0f}".replace(",", "."))
        self.val_selisih.setText(f"Rp {selisih:,.0f}".replace(",", "."))

        # Determine result
        if selisih > 0:
            # Kurang bayar - realisasi lebih besar dari uang muka
            self.result_label.setText("KURANG BAYAR")
            self.result_action.setText("Perlu mengajukan pembayaran tambahan")
            self.result_frame.setStyleSheet("""
                QFrame {
                    background-color: #fadbd8;
                    border: 2px solid #e74c3c;
                    border-radius: 8px;
                }
            """)
            self.val_selisih.setStyleSheet("font-weight: bold; font-size: 16px; color: #e74c3c;")
        elif selisih < 0:
            # Lebih bayar - realisasi lebih kecil dari uang muka
            self.result_label.setText("LEBIH BAYAR")
            self.result_action.setText(f"Perlu mengembalikan Rp {abs(selisih):,.0f}".replace(",", "."))
            self.result_frame.setStyleSheet("""
                QFrame {
                    background-color: #fef9e7;
                    border: 2px solid #f39c12;
                    border-radius: 8px;
                }
            """)
            self.val_selisih.setStyleSheet("font-weight: bold; font-size: 16px; color: #f39c12;")
        else:
            # Nihil - pas
            self.result_label.setText("NIHIL / PAS")
            self.result_action.setText("Tidak ada selisih, lanjut ke kuitansi rampung")
            self.result_frame.setStyleSheet("""
                QFrame {
                    background-color: #d5f5e3;
                    border: 2px solid #27ae60;
                    border-radius: 8px;
                }
            """)
            self.val_selisih.setStyleSheet("font-weight: bold; font-size: 16px; color: #27ae60;")

    def set_bukti_list(self, bukti_list: List[Dict]):
        """Set bukti list from external source."""
        self.bukti_list = bukti_list
        self._refresh_tables()
        self._calculate()

    def _export_to_excel(self):
        """Export calculation to Excel."""
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Export ke Excel",
            f"Perhitungan_TK_{self.transaksi.get('kode', 'transaksi')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if filepath:
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Perhitungan TK"

                # Title
                ws['A1'] = "PERHITUNGAN TAMBAH / KURANG"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:F1')

                ws['A2'] = f"Kegiatan: {self.transaksi.get('nama_kegiatan', '-')}"
                ws.merge_cells('A2:F2')

                # Permintaan section
                ws['A4'] = "A. PERMINTAAN (UANG MUKA)"
                ws['A4'].font = Font(bold=True)
                headers = ['No', 'Uraian', 'Volume', 'Harga Satuan', 'Jumlah']
                for col, h in enumerate(headers, 1):
                    ws.cell(row=5, column=col, value=h).font = Font(bold=True)

                total_perm = 0
                for i, item in enumerate(self.permintaan_list):
                    row = 6 + i
                    ws.cell(row=row, column=1, value=i + 1)
                    ws.cell(row=row, column=2, value=item.get('uraian', ''))
                    ws.cell(row=row, column=3, value=f"{item.get('volume', 1)} {item.get('satuan', '')}")
                    ws.cell(row=row, column=4, value=item.get('harga_satuan', 0))
                    ws.cell(row=row, column=5, value=item.get('jumlah', 0))
                    total_perm += item.get('jumlah', 0)

                perm_total_row = 6 + len(self.permintaan_list)
                ws.cell(row=perm_total_row, column=4, value="TOTAL A:").font = Font(bold=True)
                ws.cell(row=perm_total_row, column=5, value=total_perm).font = Font(bold=True)

                # Realisasi section
                real_start = perm_total_row + 2
                ws.cell(row=real_start, column=1, value="B. REALISASI (BUKTI PENGELUARAN)").font = Font(bold=True)
                for col, h in enumerate(headers, 1):
                    ws.cell(row=real_start + 1, column=col, value=h).font = Font(bold=True)

                total_real = 0
                for i, item in enumerate(self.bukti_list):
                    row = real_start + 2 + i
                    ws.cell(row=row, column=1, value=i + 1)
                    ws.cell(row=row, column=2, value=item.get('uraian', ''))
                    ws.cell(row=row, column=3, value=f"{item.get('volume', 1)} {item.get('satuan', '')}")
                    ws.cell(row=row, column=4, value=item.get('harga_satuan', 0))
                    ws.cell(row=row, column=5, value=item.get('jumlah', 0))
                    total_real += item.get('jumlah', 0)

                real_total_row = real_start + 2 + len(self.bukti_list)
                ws.cell(row=real_total_row, column=4, value="TOTAL B:").font = Font(bold=True)
                ws.cell(row=real_total_row, column=5, value=total_real).font = Font(bold=True)

                # Calculation section
                calc_start = real_total_row + 2
                ws.cell(row=calc_start, column=1, value="C. PERHITUNGAN").font = Font(bold=True)
                ws.cell(row=calc_start + 1, column=1, value="Uang Muka (A)")
                ws.cell(row=calc_start + 1, column=2, value=total_perm)
                ws.cell(row=calc_start + 2, column=1, value="Realisasi (B)")
                ws.cell(row=calc_start + 2, column=2, value=total_real)
                ws.cell(row=calc_start + 3, column=1, value="Selisih (B - A)").font = Font(bold=True)
                ws.cell(row=calc_start + 3, column=2, value=total_real - total_perm).font = Font(bold=True)

                selisih = total_real - total_perm
                if selisih > 0:
                    ws.cell(row=calc_start + 4, column=1, value="Status: KURANG BAYAR")
                elif selisih < 0:
                    ws.cell(row=calc_start + 4, column=1, value="Status: LEBIH BAYAR (perlu dikembalikan)")
                else:
                    ws.cell(row=calc_start + 4, column=1, value="Status: NIHIL / PAS")

                wb.save(filepath)
                QMessageBox.information(self, "Sukses", f"Berhasil export ke:\n{filepath}")

            except ImportError:
                QMessageBox.warning(self, "Error", "Library openpyxl tidak tersedia!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal export: {str(e)}")

    def _save_and_close(self):
        """Save calculation result and close."""
        total_permintaan = sum(item.get('jumlah', 0) for item in self.permintaan_list)
        total_realisasi = sum(item.get('jumlah', 0) for item in self.bukti_list)
        selisih = total_realisasi - total_permintaan

        result = {
            'uang_muka': total_permintaan,
            'realisasi': total_realisasi,
            'selisih': selisih,
            'status': 'KURANG_BAYAR' if selisih > 0 else ('LEBIH_BAYAR' if selisih < 0 else 'NIHIL'),
            'permintaan_list': self.permintaan_list,
            'bukti_list': self.bukti_list,
            'tanggal_hitung': datetime.now().strftime("%Y-%m-%d"),
        }

        self.calculation_saved.emit(result)
        self.accept()

    def get_result(self) -> Dict:
        """Get calculation result."""
        total_permintaan = sum(item.get('jumlah', 0) for item in self.permintaan_list)
        total_realisasi = sum(item.get('jumlah', 0) for item in self.bukti_list)
        selisih = total_realisasi - total_permintaan

        return {
            'uang_muka': total_permintaan,
            'realisasi': total_realisasi,
            'selisih': selisih,
            'status': 'KURANG_BAYAR' if selisih > 0 else ('LEBIH_BAYAR' if selisih < 0 else 'NIHIL'),
        }
