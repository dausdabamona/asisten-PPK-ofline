"""
PPK Document Factory - Database Workers
=======================================
Background workers for database operations.

Workers:
- DataLoadWorker: Load data from database
- DataSaveWorker: Save data to database
- ExportWorker: Export data to files
- BulkOperationWorker: Perform bulk database operations
"""

from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Union
from pathlib import Path
from enum import Enum
import logging
import time
import csv
import json

from PySide6.QtCore import QObject

from ..workers_base import BaseWorker

logger = logging.getLogger(__name__)


# =============================================================================
# ENUMS AND DATA CLASSES
# =============================================================================


class ExportFormat(Enum):
    """Supported export formats."""
    CSV = "csv"
    JSON = "json"
    EXCEL = "xlsx"
    PDF = "pdf"


@dataclass
class LoadResult:
    """
    Result of data loading operation.

    Attributes:
        success: Whether loading was successful
        data: Loaded data (list, dict, or other)
        count: Number of records loaded
        error_message: Error message if failed
        load_time: Time taken in seconds
        from_cache: Whether data came from cache
    """
    success: bool
    data: Any = None
    count: int = 0
    error_message: Optional[str] = None
    load_time: float = 0.0
    from_cache: bool = False


@dataclass
class SaveResult:
    """
    Result of data save operation.

    Attributes:
        success: Whether save was successful
        affected_rows: Number of rows affected
        generated_id: ID of new record if insert
        error_message: Error message if failed
        save_time: Time taken in seconds
    """
    success: bool
    affected_rows: int = 0
    generated_id: Optional[int] = None
    error_message: Optional[str] = None
    save_time: float = 0.0


@dataclass
class ExportResult:
    """
    Result of data export operation.

    Attributes:
        success: Whether export was successful
        output_path: Path to exported file
        record_count: Number of records exported
        file_size: Size of exported file in bytes
        error_message: Error message if failed
        export_time: Time taken in seconds
    """
    success: bool
    output_path: Optional[str] = None
    record_count: int = 0
    file_size: int = 0
    error_message: Optional[str] = None
    export_time: float = 0.0


@dataclass
class BulkResult:
    """
    Result of bulk operation.

    Attributes:
        success: Whether all operations succeeded
        total: Total number of operations
        successful: Number of successful operations
        failed: Number of failed operations
        errors: List of error messages
        total_time: Time taken in seconds
    """
    success: bool = True
    total: int = 0
    successful: int = 0
    failed: int = 0
    errors: List[str] = field(default_factory=list)
    total_time: float = 0.0


# =============================================================================
# DATA LOAD WORKER
# =============================================================================


class DataLoadWorker(BaseWorker):
    """
    Worker for loading data from database in background.

    Executes a query function and returns the results.
    Supports caching for repeated queries.

    Example:
    --------
    ```python
    # Load with a function
    worker = DataLoadWorker(
        query_func=db.list_transaksi,
        mekanisme="UP",
        status="Aktif"
    )
    worker.signals.result.connect(self._on_data_loaded)
    worker.start()

    # With caching
    worker = DataLoadWorker(
        query_func=db.list_pegawai,
        use_cache=True,
        cache_key="pegawai:all",
        cache_ttl=1800
    )
    ```
    """

    def __init__(
        self,
        query_func: Callable,
        *args,
        use_cache: bool = False,
        cache_key: Optional[str] = None,
        cache_ttl: int = 300,
        parent: Optional[QObject] = None,
        **kwargs
    ):
        """
        Initialize DataLoadWorker.

        Args:
            query_func: Function to execute for loading data
            *args: Positional arguments for query_func
            use_cache: Whether to use caching
            cache_key: Key for caching (auto-generated if None)
            cache_ttl: Cache time-to-live in seconds
            parent: Parent QObject
            **kwargs: Keyword arguments for query_func
        """
        super().__init__(parent)
        self._query_func = query_func
        self._args = args
        self._kwargs = kwargs
        self._use_cache = use_cache
        self._cache_key = cache_key
        self._cache_ttl = cache_ttl

    def do_work(self) -> LoadResult:
        """Execute the query and return results."""
        start_time = time.time()

        try:
            self.emit_progress(10)
            self.emit_status("Memuat data...")

            from_cache = False
            data = None

            # Try cache first
            if self._use_cache:
                try:
                    from ..cache import DataCache
                    cache = DataCache()
                    cache_key = self._cache_key or self._generate_cache_key()

                    cached = cache.get(cache_key)
                    if cached is not None:
                        data = cached
                        from_cache = True
                        logger.debug(f"Data loaded from cache: {cache_key}")
                except ImportError:
                    logger.warning("Cache not available")

            if self.is_cancelled:
                return LoadResult(success=False, error_message="Dibatalkan")

            self.emit_progress(30)

            # Load from database if not cached
            if data is None:
                data = self._query_func(*self._args, **self._kwargs)

                # Save to cache
                if self._use_cache:
                    try:
                        from ..cache import DataCache
                        cache = DataCache()
                        cache_key = self._cache_key or self._generate_cache_key()
                        cache.set(cache_key, data, ttl=self._cache_ttl)
                    except ImportError:
                        pass

            if self.is_cancelled:
                return LoadResult(success=False, error_message="Dibatalkan")

            self.emit_progress(90)

            # Calculate count
            count = 0
            if isinstance(data, (list, tuple)):
                count = len(data)
            elif isinstance(data, dict):
                count = len(data)
            elif data is not None:
                count = 1

            self.emit_progress(100)
            load_time = time.time() - start_time

            logger.info(
                f"Data loaded: {count} records in {load_time:.2f}s "
                f"(cache: {from_cache})"
            )

            return LoadResult(
                success=True,
                data=data,
                count=count,
                load_time=load_time,
                from_cache=from_cache
            )

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Data load failed: {error_msg}")

            return LoadResult(
                success=False,
                error_message=error_msg,
                load_time=time.time() - start_time
            )

    def _generate_cache_key(self) -> str:
        """Generate cache key from function and arguments."""
        func_name = self._query_func.__name__
        args_hash = hash((
            self._args,
            tuple(sorted(self._kwargs.items()))
        ))
        return f"query:{func_name}:{args_hash}"


# =============================================================================
# DATA SAVE WORKER
# =============================================================================


class DataSaveWorker(BaseWorker):
    """
    Worker for saving data to database in background.

    Example:
    --------
    ```python
    worker = DataSaveWorker(
        save_func=db.create_transaksi,
        data={"nama": "Test", "nilai": 1000000}
    )
    worker.signals.result.connect(self._on_save_complete)
    worker.start()
    ```
    """

    def __init__(
        self,
        save_func: Callable,
        data: Any,
        invalidate_cache_patterns: Optional[List[str]] = None,
        parent: Optional[QObject] = None
    ):
        """
        Initialize DataSaveWorker.

        Args:
            save_func: Function to call for saving
            data: Data to save
            invalidate_cache_patterns: Cache patterns to invalidate after save
            parent: Parent QObject
        """
        super().__init__(parent)
        self._save_func = save_func
        self._data = data
        self._invalidate_patterns = invalidate_cache_patterns or []

    def do_work(self) -> SaveResult:
        """Execute the save operation."""
        start_time = time.time()

        try:
            self.emit_progress(10)
            self.emit_status("Menyimpan data...")

            if self.is_cancelled:
                return SaveResult(success=False, error_message="Dibatalkan")

            self.emit_progress(30)

            # Execute save
            result = self._save_func(self._data)

            if self.is_cancelled:
                return SaveResult(success=False, error_message="Dibatalkan")

            self.emit_progress(70)
            self.emit_status("Memperbarui cache...")

            # Invalidate cache
            if self._invalidate_patterns:
                try:
                    from ..cache import DataCache
                    cache = DataCache()
                    for pattern in self._invalidate_patterns:
                        cache.invalidate_pattern(pattern)
                        logger.debug(f"Cache invalidated: {pattern}")
                except ImportError:
                    pass

            self.emit_progress(100)
            save_time = time.time() - start_time

            # Parse result
            generated_id = None
            affected_rows = 1

            if isinstance(result, int):
                generated_id = result
            elif isinstance(result, dict):
                generated_id = result.get("id")
                affected_rows = result.get("affected_rows", 1)

            logger.info(f"Data saved in {save_time:.2f}s (id: {generated_id})")

            return SaveResult(
                success=True,
                affected_rows=affected_rows,
                generated_id=generated_id,
                save_time=save_time
            )

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Data save failed: {error_msg}")

            return SaveResult(
                success=False,
                error_message=error_msg,
                save_time=time.time() - start_time
            )


# =============================================================================
# EXPORT WORKER
# =============================================================================


class ExportWorker(BaseWorker):
    """
    Worker for exporting data to files.

    Supports CSV, JSON, Excel, and PDF formats.

    Example:
    --------
    ```python
    worker = ExportWorker(
        data=transaksi_list,
        format=ExportFormat.CSV,
        output_path="/exports/transaksi.csv",
        columns=["id", "nama", "nilai", "status"]
    )
    worker.signals.result.connect(self._on_export_complete)
    worker.start()
    ```
    """

    def __init__(
        self,
        data: List[Dict[str, Any]],
        format: Union[ExportFormat, str],
        output_path: str,
        columns: Optional[List[str]] = None,
        column_headers: Optional[Dict[str, str]] = None,
        parent: Optional[QObject] = None
    ):
        """
        Initialize ExportWorker.

        Args:
            data: List of dictionaries to export
            format: Export format (ExportFormat or string)
            output_path: Path for output file
            columns: List of columns to include (None = all)
            column_headers: Mapping of column keys to display headers
            parent: Parent QObject
        """
        super().__init__(parent)
        self._data = data
        self._format = ExportFormat(format) if isinstance(format, str) else format
        self._output_path = Path(output_path)
        self._columns = columns
        self._column_headers = column_headers or {}

    def do_work(self) -> ExportResult:
        """Export the data."""
        start_time = time.time()

        try:
            self.emit_progress(10)
            self.emit_status("Mempersiapkan data...")

            if not self._data:
                return ExportResult(
                    success=True,
                    output_path=str(self._output_path),
                    record_count=0,
                    export_time=time.time() - start_time
                )

            # Determine columns
            columns = self._columns
            if columns is None:
                # Get all columns from first record
                columns = list(self._data[0].keys())

            if self.is_cancelled:
                return ExportResult(success=False, error_message="Dibatalkan")

            self.emit_progress(30)
            self.emit_status("Mengekspor data...")

            # Ensure output directory exists
            self._output_path.parent.mkdir(parents=True, exist_ok=True)

            # Export based on format
            if self._format == ExportFormat.CSV:
                self._export_csv(columns)
            elif self._format == ExportFormat.JSON:
                self._export_json(columns)
            elif self._format == ExportFormat.EXCEL:
                self._export_excel(columns)
            elif self._format == ExportFormat.PDF:
                self._export_pdf(columns)
            else:
                raise ValueError(f"Unsupported format: {self._format}")

            if self.is_cancelled:
                return ExportResult(success=False, error_message="Dibatalkan")

            self.emit_progress(90)

            # Get file size
            file_size = 0
            if self._output_path.exists():
                file_size = self._output_path.stat().st_size

            self.emit_progress(100)
            export_time = time.time() - start_time

            logger.info(
                f"Exported {len(self._data)} records to {self._format.value} "
                f"in {export_time:.2f}s ({file_size} bytes)"
            )

            return ExportResult(
                success=True,
                output_path=str(self._output_path),
                record_count=len(self._data),
                file_size=file_size,
                export_time=export_time
            )

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Export failed: {error_msg}")

            return ExportResult(
                success=False,
                error_message=error_msg,
                export_time=time.time() - start_time
            )

    def _export_csv(self, columns: List[str]) -> None:
        """Export to CSV format."""
        headers = [self._column_headers.get(col, col) for col in columns]

        with open(self._output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            total = len(self._data)
            for i, row in enumerate(self._data):
                if self.is_cancelled:
                    break

                values = [row.get(col, '') for col in columns]
                writer.writerow(values)

                if i % 100 == 0:
                    progress = 30 + (i * 60 // total)
                    self.emit_progress(progress)

    def _export_json(self, columns: List[str]) -> None:
        """Export to JSON format."""
        # Filter columns
        filtered_data = [
            {col: row.get(col) for col in columns}
            for row in self._data
        ]

        with open(self._output_path, 'w', encoding='utf-8') as f:
            json.dump(filtered_data, f, ensure_ascii=False, indent=2, default=str)

    def _export_excel(self, columns: List[str]) -> None:
        """Export to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers
        headers = [self._column_headers.get(col, col) for col in columns]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0",
                end_color="E0E0E0",
                fill_type="solid"
            )

        # Data rows
        total = len(self._data)
        for row_idx, row_data in enumerate(self._data, 2):
            if self.is_cancelled:
                break

            for col_idx, col in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col, ''))

            if row_idx % 100 == 0:
                progress = 30 + ((row_idx - 1) * 60 // total)
                self.emit_progress(progress)

        # Adjust column widths
        for col_idx, col in enumerate(columns, 1):
            max_length = len(headers[col_idx - 1])
            for row in ws.iter_rows(
                min_row=2, max_row=min(100, total + 1),
                min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_length + 2, 50)

        wb.save(self._output_path)

    def _export_pdf(self, columns: List[str]) -> None:
        """Export to PDF format."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib.units import cm
        except ImportError:
            raise RuntimeError("reportlab required for PDF export")

        doc = SimpleDocTemplate(
            str(self._output_path),
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )

        # Prepare data
        headers = [self._column_headers.get(col, col) for col in columns]
        table_data = [headers]

        for row in self._data:
            values = [str(row.get(col, ''))[:50] for col in columns]  # Truncate
            table_data.append(values)

        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        doc.build([table])


# =============================================================================
# BULK OPERATION WORKER
# =============================================================================


class BulkOperationWorker(BaseWorker):
    """
    Worker for performing bulk database operations.

    Example:
    --------
    ```python
    # Bulk update
    worker = BulkOperationWorker(
        operation_func=db.update_transaksi,
        items=[
            {"id": 1, "status": "Selesai"},
            {"id": 2, "status": "Selesai"},
        ]
    )
    worker.start()
    ```
    """

    def __init__(
        self,
        operation_func: Callable[[Any], Any],
        items: List[Any],
        parent: Optional[QObject] = None,
        continue_on_error: bool = True,
        invalidate_cache_patterns: Optional[List[str]] = None
    ):
        """
        Initialize BulkOperationWorker.

        Args:
            operation_func: Function to call for each item
            items: List of items to process
            parent: Parent QObject
            continue_on_error: Continue on individual errors
            invalidate_cache_patterns: Cache patterns to invalidate
        """
        super().__init__(parent)
        self._operation_func = operation_func
        self._items = items
        self._continue_on_error = continue_on_error
        self._invalidate_patterns = invalidate_cache_patterns or []

    def do_work(self) -> BulkResult:
        """Execute bulk operations."""
        start_time = time.time()
        total = len(self._items)
        successful = 0
        failed = 0
        errors = []

        for i, item in enumerate(self._items):
            if self.is_cancelled:
                break

            progress = (i * 100) // total
            self.emit_progress(progress)
            self.emit_status(f"Memproses {i + 1}/{total}...")

            try:
                self._operation_func(item)
                successful += 1

            except Exception as e:
                error_msg = f"Item {i}: {str(e)}"
                errors.append(error_msg)
                failed += 1
                logger.warning(f"Bulk operation error: {error_msg}")

                if not self._continue_on_error:
                    break

        # Invalidate cache
        if self._invalidate_patterns and successful > 0:
            try:
                from ..cache import DataCache
                cache = DataCache()
                for pattern in self._invalidate_patterns:
                    cache.invalidate_pattern(pattern)
            except ImportError:
                pass

        self.emit_progress(100)
        total_time = time.time() - start_time

        logger.info(
            f"Bulk operation complete: {successful}/{total} successful "
            f"in {total_time:.2f}s"
        )

        return BulkResult(
            success=failed == 0,
            total=total,
            successful=successful,
            failed=failed,
            errors=errors,
            total_time=total_time
        )


# =============================================================================
# FACTORY FUNCTIONS
# =============================================================================


def create_load_worker(
    query_func: Callable,
    *args,
    **kwargs
) -> DataLoadWorker:
    """Create a data load worker."""
    return DataLoadWorker(query_func, *args, **kwargs)


def create_save_worker(
    save_func: Callable,
    data: Any,
    **kwargs
) -> DataSaveWorker:
    """Create a data save worker."""
    return DataSaveWorker(save_func, data, **kwargs)


def create_export_worker(
    data: List[Dict],
    format: Union[ExportFormat, str],
    output_path: str,
    **kwargs
) -> ExportWorker:
    """Create an export worker."""
    return ExportWorker(data, format, output_path, **kwargs)


# =============================================================================
# EXPORTS
# =============================================================================

__all__ = [
    # Enums
    'ExportFormat',

    # Data classes
    'LoadResult',
    'SaveResult',
    'ExportResult',
    'BulkResult',

    # Workers
    'DataLoadWorker',
    'DataSaveWorker',
    'ExportWorker',
    'BulkOperationWorker',

    # Factory functions
    'create_load_worker',
    'create_save_worker',
    'create_export_worker',
]
