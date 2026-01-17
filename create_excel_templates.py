"""
Script untuk membuat template Excel Survey Harga dan HPS
dengan placeholder item yang benar
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATES_DIR = "templates/excel"

def create_survey_harga_excel_template():
    """Create Survey Harga Excel template with item placeholders"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey Harga"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "FORM SURVEY HARGA"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Paket info
    ws['A2'] = "Nama Paket:"
    ws['B2'] = "{{nama_paket}}"
    ws.merge_cells('B2:L2')
    
    ws['A3'] = "Kode Paket:"
    ws['B3'] = "{{kode_paket}}"
    
    ws['A4'] = "Tahun Anggaran:"
    ws['B4'] = "{{tahun_anggaran}}"
    
    # Empty row
    ws['A5'] = ""
    
    # Headers (row 6)
    headers = [
        ("A", "No"),
        ("B", "Uraian Barang/Jasa"),
        ("C", "Spesifikasi"),
        ("D", "Satuan"),
        ("E", "Volume"),
        ("F", "Sumber 1"),
        ("G", "Harga Survey 1"),
        ("H", "Sumber 2"),
        ("I", "Harga Survey 2"),
        ("J", "Sumber 3"),
        ("K", "Harga Survey 3"),
        ("L", "Rata-rata"),
    ]
    
    for col, header in headers:
        cell = ws[f'{col}6']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Template row with placeholders (row 7)
    ws['A7'] = "{{no}}"
    ws['A7'].border = border
    ws['A7'].alignment = Alignment(horizontal='center')
    
    ws['B7'] = "{{item.uraian}}"
    ws['B7'].border = border
    
    ws['C7'] = "{{item.spesifikasi}}"
    ws['C7'].border = border
    
    ws['D7'] = "{{item.satuan}}"
    ws['D7'].border = border
    ws['D7'].alignment = Alignment(horizontal='center')
    
    ws['E7'] = "{{item.volume}}"
    ws['E7'].border = border
    ws['E7'].alignment = Alignment(horizontal='right')
    
    ws['F7'] = ""  # Sumber 1 - to be filled
    ws['F7'].border = border
    
    ws['G7'] = "{{item.harga_survey1}}"
    ws['G7'].border = border
    ws['G7'].alignment = Alignment(horizontal='right')
    ws['G7'].number_format = '#,##0'
    
    ws['H7'] = ""  # Sumber 2
    ws['H7'].border = border
    
    ws['I7'] = "{{item.harga_survey2}}"
    ws['I7'].border = border
    ws['I7'].alignment = Alignment(horizontal='right')
    ws['I7'].number_format = '#,##0'
    
    ws['J7'] = ""  # Sumber 3
    ws['J7'].border = border
    
    ws['K7'] = "{{item.harga_survey3}}"
    ws['K7'].border = border
    ws['K7'].alignment = Alignment(horizontal='right')
    ws['K7'].number_format = '#,##0'
    
    ws['L7'] = "{{item.harga_rata}}"
    ws['L7'].border = border
    ws['L7'].alignment = Alignment(horizontal='right')
    ws['L7'].number_format = '#,##0'
    
    # Column widths
    widths = {'A': 5, 'B': 35, 'C': 25, 'D': 8, 'E': 8, 'F': 15, 'G': 12, 
              'H': 15, 'I': 12, 'J': 15, 'K': 12, 'L': 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Row height for header
    ws.row_dimensions[6].height = 30
    
    # Save
    filepath = os.path.join(TEMPLATES_DIR, "survey_harga.xlsx")
    wb.save(filepath)
    print(f"✅ Created: {filepath}")
    return filepath


def create_hps_excel_template():
    """Create HPS Excel template with item placeholders"""
    wb = Workbook()
    ws = wb.active
    ws.title = "HPS"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "HARGA PERKIRAAN SENDIRI (HPS)"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Paket info
    ws['A2'] = "Nama Paket:"
    ws['B2'] = "{{nama_paket}}"
    ws.merge_cells('B2:H2')
    
    ws['A3'] = "Kode Paket:"
    ws['B3'] = "{{kode_paket}}"
    
    ws['A4'] = "Tahun Anggaran:"
    ws['B4'] = "{{tahun_anggaran}}"
    
    ws['A5'] = "Pagu Anggaran:"
    ws['B5'] = "{{nilai_pagu_fmt}}"
    
    ws['A6'] = "Metode HPS:"
    ws['B6'] = "{{metode_hps}}"
    
    # Empty row
    ws['A7'] = ""
    
    # Headers (row 8)
    headers = [
        ("A", "No"),
        ("B", "Uraian Barang/Jasa"),
        ("C", "Spesifikasi"),
        ("D", "Satuan"),
        ("E", "Volume"),
        ("F", "Harga Survey"),
        ("G", "Harga HPS/Satuan"),
        ("H", "Total HPS"),
    ]
    
    for col, header in headers:
        cell = ws[f'{col}8']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Template row with placeholders (row 9)
    ws['A9'] = "{{no}}"
    ws['A9'].border = border
    ws['A9'].alignment = Alignment(horizontal='center')
    
    ws['B9'] = "{{item.uraian}}"
    ws['B9'].border = border
    
    ws['C9'] = "{{item.spesifikasi}}"
    ws['C9'].border = border
    
    ws['D9'] = "{{item.satuan}}"
    ws['D9'].border = border
    ws['D9'].alignment = Alignment(horizontal='center')
    
    ws['E9'] = "{{item.volume}}"
    ws['E9'].border = border
    ws['E9'].alignment = Alignment(horizontal='right')
    
    ws['F9'] = "{{item.harga_rata}}"
    ws['F9'].border = border
    ws['F9'].alignment = Alignment(horizontal='right')
    ws['F9'].number_format = '#,##0'
    
    ws['G9'] = "{{item.harga_hps_satuan}}"
    ws['G9'].border = border
    ws['G9'].alignment = Alignment(horizontal='right')
    ws['G9'].number_format = '#,##0'
    
    ws['H9'] = "{{item.total_hps}}"
    ws['H9'].border = border
    ws['H9'].alignment = Alignment(horizontal='right')
    ws['H9'].number_format = '#,##0'
    
    # Summary rows (these will be after items, adjust dynamically)
    # For now, put them at row 10+ (will be pushed down when items inserted)
    summary_start = 11
    
    ws.merge_cells(f'A{summary_start}:G{summary_start}')
    ws[f'A{summary_start}'] = "SUBTOTAL"
    ws[f'A{summary_start}'].font = bold_font
    ws[f'A{summary_start}'].alignment = Alignment(horizontal='right')
    ws[f'A{summary_start}'].border = border
    ws[f'H{summary_start}'] = "{{subtotal_item_fmt}}"
    ws[f'H{summary_start}'].font = bold_font
    ws[f'H{summary_start}'].border = border
    ws[f'H{summary_start}'].alignment = Alignment(horizontal='right')
    
    ws.merge_cells(f'A{summary_start+1}:G{summary_start+1}')
    ws[f'A{summary_start+1}'] = "PPN 11%"
    ws[f'A{summary_start+1}'].alignment = Alignment(horizontal='right')
    ws[f'A{summary_start+1}'].border = border
    ws[f'H{summary_start+1}'] = "{{ppn_item_fmt}}"
    ws[f'H{summary_start+1}'].border = border
    ws[f'H{summary_start+1}'].alignment = Alignment(horizontal='right')
    
    ws.merge_cells(f'A{summary_start+2}:G{summary_start+2}')
    ws[f'A{summary_start+2}'] = "TOTAL HPS"
    ws[f'A{summary_start+2}'].font = bold_font
    ws[f'A{summary_start+2}'].fill = total_fill
    ws[f'A{summary_start+2}'].alignment = Alignment(horizontal='right')
    ws[f'A{summary_start+2}'].border = border
    ws[f'H{summary_start+2}'] = "{{grand_total_item_fmt}}"
    ws[f'H{summary_start+2}'].font = bold_font
    ws[f'H{summary_start+2}'].fill = total_fill
    ws[f'H{summary_start+2}'].border = border
    ws[f'H{summary_start+2}'].alignment = Alignment(horizontal='right')
    
    # Terbilang
    ws.merge_cells(f'A{summary_start+4}:B{summary_start+4}')
    ws[f'A{summary_start+4}'] = "Terbilang:"
    ws[f'A{summary_start+4}'].font = bold_font
    
    ws.merge_cells(f'C{summary_start+4}:H{summary_start+4}')
    ws[f'C{summary_start+4}'] = "{{grand_total_item_terbilang}}"
    
    # Column widths
    widths = {'A': 5, 'B': 35, 'C': 25, 'D': 8, 'E': 8, 'F': 15, 'G': 15, 'H': 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Row height for header
    ws.row_dimensions[8].height = 30
    
    # Save
    filepath = os.path.join(TEMPLATES_DIR, "hps.xlsx")
    wb.save(filepath)
    print(f"✅ Created: {filepath}")
    return filepath


if __name__ == "__main__":
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    
    print("Creating Excel templates for Survey Harga and HPS...")
    print()
    
    create_survey_harga_excel_template()
    create_hps_excel_template()
    
    print()
    print("=" * 50)
    print("Templates created successfully!")
    print()
    print("Placeholders used:")
    print("  {{no}}              - Nomor urut item")
    print("  {{item.uraian}}     - Uraian barang/jasa")
    print("  {{item.spesifikasi}}- Spesifikasi")
    print("  {{item.satuan}}     - Satuan")
    print("  {{item.volume}}     - Volume")
    print("  {{item.harga_survey1/2/3}} - Harga survey")
    print("  {{item.harga_rata}} - Harga rata-rata")
    print("  {{item.harga_hps_satuan}} - Harga HPS/satuan")
    print("  {{item.total_hps}}  - Total HPS")
