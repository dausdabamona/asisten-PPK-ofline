"""
PPK DOCUMENT FACTORY v4.0 - Pegawai Manager
============================================
Master Data Pegawai dengan:
- Table view (search, filter, sort)
- Add / Edit / Nonaktifkan Pegawai
- Import CSV
- Export Excel
"""

import os
import csv
from datetime import datetime
from typing import List, Dict, Optional

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QWidget, QPushButton, QLabel, QLineEdit, QTextEdit, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QMessageBox, QCheckBox, QFrame, QSplitter, QTabWidget,
    QAbstractItemView, QMenu, QFileDialog, QProgressDialog, QApplication
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QAction, QColor

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.core.database_v4 import get_db_manager_v4, PERAN_PEJABAT


# ============================================================================
# PEGAWAI DIALOG (Add/Edit)
# ============================================================================

class PegawaiDialog(QDialog):
    """Dialog for adding/editing pegawai"""
    
    def __init__(self, pegawai_data: dict = None, parent=None):
        super().__init__(parent)
        self.pegawai_data = pegawai_data
        self.db = get_db_manager_v4()
        
        self.setWindowTitle("Tambah Pegawai" if not pegawai_data else "Edit Pegawai")
        self.setMinimumWidth(600)
        self.setup_ui()
        
        if pegawai_data:
            self.load_data()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Tabs
        tabs = QTabWidget()
        
        # Tab 1: Data Utama
        tab_utama = QWidget()
        form1 = QFormLayout(tab_utama)
        
        self.txt_nip = QLineEdit()
        self.txt_nip.setPlaceholderText("18 digit NIP...")
        form1.addRow("NIP:", self.txt_nip)
        
        # Nama lengkap
        nama_layout = QHBoxLayout()
        self.txt_gelar_depan = QLineEdit()
        self.txt_gelar_depan.setPlaceholderText("Dr., Ir., ...")
        self.txt_gelar_depan.setMaximumWidth(80)
        nama_layout.addWidget(self.txt_gelar_depan)
        
        self.txt_nama = QLineEdit()
        self.txt_nama.setPlaceholderText("Nama lengkap tanpa gelar...")
        nama_layout.addWidget(self.txt_nama)
        
        self.txt_gelar_belakang = QLineEdit()
        self.txt_gelar_belakang.setPlaceholderText("S.T., M.M., ...")
        self.txt_gelar_belakang.setMaximumWidth(100)
        nama_layout.addWidget(self.txt_gelar_belakang)
        
        form1.addRow("Nama:", nama_layout)
        
        # Pangkat/Golongan
        pg_layout = QHBoxLayout()
        self.txt_pangkat = QComboBox()
        self.txt_pangkat.setEditable(True)
        self.txt_pangkat.addItems([
            '', 'Juru Muda', 'Juru Muda Tingkat I', 'Juru', 'Juru Tingkat I',
            'Pengatur Muda', 'Pengatur Muda Tingkat I', 'Pengatur', 'Pengatur Tingkat I',
            'Penata Muda', 'Penata Muda Tingkat I', 'Penata', 'Penata Tingkat I',
            'Pembina', 'Pembina Tingkat I', 'Pembina Utama Muda', 
            'Pembina Utama Madya', 'Pembina Utama'
        ])
        pg_layout.addWidget(self.txt_pangkat)
        
        self.txt_golongan = QComboBox()
        self.txt_golongan.setEditable(True)
        self.txt_golongan.addItems([
            '', 'I/a', 'I/b', 'I/c', 'I/d',
            'II/a', 'II/b', 'II/c', 'II/d',
            'III/a', 'III/b', 'III/c', 'III/d',
            'IV/a', 'IV/b', 'IV/c', 'IV/d', 'IV/e'
        ])
        self.txt_golongan.setMaximumWidth(80)
        pg_layout.addWidget(self.txt_golongan)
        
        form1.addRow("Pangkat/Gol:", pg_layout)
        
        self.txt_jabatan = QLineEdit()
        self.txt_jabatan.setPlaceholderText("Jabatan struktural/fungsional...")
        form1.addRow("Jabatan:", self.txt_jabatan)
        
        self.txt_unit_kerja = QLineEdit()
        self.txt_unit_kerja.setPlaceholderText("Unit kerja/bagian...")
        form1.addRow("Unit Kerja:", self.txt_unit_kerja)
        
        tabs.addTab(tab_utama, "Data Utama")
        
        # Tab 2: Kontak & Bank
        tab_kontak = QWidget()
        form2 = QFormLayout(tab_kontak)
        
        self.txt_email = QLineEdit()
        self.txt_email.setPlaceholderText("email@domain.com")
        form2.addRow("Email:", self.txt_email)
        
        self.txt_telepon = QLineEdit()
        self.txt_telepon.setPlaceholderText("08xxx...")
        form2.addRow("Telepon:", self.txt_telepon)
        
        self.txt_npwp = QLineEdit()
        self.txt_npwp.setPlaceholderText("XX.XXX.XXX.X-XXX.XXX")
        form2.addRow("NPWP:", self.txt_npwp)
        
        self.txt_no_rekening = QLineEdit()
        self.txt_no_rekening.setPlaceholderText("Nomor rekening bank...")
        form2.addRow("No. Rekening:", self.txt_no_rekening)
        
        self.txt_nama_bank = QComboBox()
        self.txt_nama_bank.setEditable(True)
        self.txt_nama_bank.addItems([
            '', 'Bank Mandiri', 'BRI', 'BNI', 'BTN', 'Bank Papua',
            'Bank Sulselbar', 'Bank DKI', 'Bank Jateng', 'Bank Jatim'
        ])
        form2.addRow("Nama Bank:", self.txt_nama_bank)
        
        tabs.addTab(tab_kontak, "Kontak & Bank")
        
        # Tab 3: Peran
        tab_peran = QWidget()
        peran_layout = QVBoxLayout(tab_peran)
        
        peran_info = QLabel("Centang peran yang dapat dijalankan pegawai ini:")
        peran_info.setStyleSheet("color: #666; font-style: italic;")
        peran_layout.addWidget(peran_info)
        
        self.chk_ppk = QCheckBox("Pejabat Pembuat Komitmen (PPK)")
        peran_layout.addWidget(self.chk_ppk)
        
        self.chk_pejabat_pengadaan = QCheckBox("Pejabat Pengadaan")
        peran_layout.addWidget(self.chk_pejabat_pengadaan)
        
        self.chk_ppspm = QCheckBox("Pejabat Penandatangan SPM (PPSPM)")
        peran_layout.addWidget(self.chk_ppspm)
        
        self.chk_bendahara = QCheckBox("Bendahara Pengeluaran")
        peran_layout.addWidget(self.chk_bendahara)
        
        self.chk_pemeriksa = QCheckBox("Pejabat Pemeriksa Hasil Pekerjaan (PPHP)")
        peran_layout.addWidget(self.chk_pemeriksa)
        
        peran_layout.addStretch()
        
        tabs.addTab(tab_peran, "Peran")
        
        layout.addWidget(tabs)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_save = QPushButton("💾 Simpan")
        btn_save.setObjectName("btnSuccess")
        btn_save.clicked.connect(self.save)
        btn_layout.addWidget(btn_save)
        
        btn_cancel = QPushButton("Batal")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)
        
        layout.addLayout(btn_layout)
        
        # Style
        self.setStyleSheet("""
            QPushButton#btnSuccess {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton#btnSuccess:hover {
                background-color: #219a52;
            }
        """)
    
    def load_data(self):
        """Load existing pegawai data"""
        if not self.pegawai_data:
            return
        
        self.txt_nip.setText(self.pegawai_data.get('nip', '') or '')
        self.txt_nama.setText(self.pegawai_data.get('nama', '') or '')
        self.txt_gelar_depan.setText(self.pegawai_data.get('gelar_depan', '') or '')
        self.txt_gelar_belakang.setText(self.pegawai_data.get('gelar_belakang', '') or '')
        self.txt_pangkat.setCurrentText(self.pegawai_data.get('pangkat', '') or '')
        self.txt_golongan.setCurrentText(self.pegawai_data.get('golongan', '') or '')
        self.txt_jabatan.setText(self.pegawai_data.get('jabatan', '') or '')
        self.txt_unit_kerja.setText(self.pegawai_data.get('unit_kerja', '') or '')
        self.txt_email.setText(self.pegawai_data.get('email', '') or '')
        self.txt_telepon.setText(self.pegawai_data.get('telepon', '') or '')
        self.txt_npwp.setText(self.pegawai_data.get('npwp', '') or '')
        self.txt_no_rekening.setText(self.pegawai_data.get('no_rekening', '') or '')
        self.txt_nama_bank.setCurrentText(self.pegawai_data.get('nama_bank', '') or '')
        
        self.chk_ppk.setChecked(bool(self.pegawai_data.get('is_ppk')))
        self.chk_pejabat_pengadaan.setChecked(bool(self.pegawai_data.get('is_pejabat_pengadaan')))
        self.chk_ppspm.setChecked(bool(self.pegawai_data.get('is_ppspm')))
        self.chk_bendahara.setChecked(bool(self.pegawai_data.get('is_bendahara')))
        self.chk_pemeriksa.setChecked(bool(self.pegawai_data.get('is_pemeriksa')))
    
    def get_data(self) -> dict:
        """Get form data"""
        return {
            'nip': self.txt_nip.text().strip() or None,
            'nama': self.txt_nama.text().strip(),
            'gelar_depan': self.txt_gelar_depan.text().strip() or None,
            'gelar_belakang': self.txt_gelar_belakang.text().strip() or None,
            'pangkat': self.txt_pangkat.currentText().strip() or None,
            'golongan': self.txt_golongan.currentText().strip() or None,
            'jabatan': self.txt_jabatan.text().strip() or None,
            'unit_kerja': self.txt_unit_kerja.text().strip() or None,
            'email': self.txt_email.text().strip() or None,
            'telepon': self.txt_telepon.text().strip() or None,
            'npwp': self.txt_npwp.text().strip() or None,
            'no_rekening': self.txt_no_rekening.text().strip() or None,
            'nama_bank': self.txt_nama_bank.currentText().strip() or None,
            'is_ppk': 1 if self.chk_ppk.isChecked() else 0,
            'is_pejabat_pengadaan': 1 if self.chk_pejabat_pengadaan.isChecked() else 0,
            'is_ppspm': 1 if self.chk_ppspm.isChecked() else 0,
            'is_bendahara': 1 if self.chk_bendahara.isChecked() else 0,
            'is_pemeriksa': 1 if self.chk_pemeriksa.isChecked() else 0,
        }
    
    def save(self):
        """Validate and save"""
        data = self.get_data()
        
        if not data['nama']:
            QMessageBox.warning(self, "Validasi", "Nama pegawai wajib diisi!")
            self.txt_nama.setFocus()
            return
        
        try:
            if self.pegawai_data:
                self.db.update_pegawai(self.pegawai_data['id'], data)
            else:
                self.db.create_pegawai(data)
            
            self.accept()
            
        except Exception as e:
            if 'UNIQUE constraint' in str(e) and 'nip' in str(e):
                QMessageBox.warning(self, "Error", "NIP sudah terdaftar!")
            else:
                QMessageBox.critical(self, "Error", f"Gagal menyimpan:\n{str(e)}")


# ============================================================================
# PEGAWAI MANAGER (Main Window)
# ============================================================================

class PegawaiManager(QDialog):
    """Main dialog for managing pegawai"""
    
    pegawai_changed = Signal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = get_db_manager_v4()
        
        self.setWindowTitle("📋 Master Data Pegawai")
        self.setMinimumSize(1000, 600)
        
        self.setup_ui()
        self.load_data()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Header
        header = QHBoxLayout()
        
        title = QLabel("👥 Master Data Pegawai")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        header.addWidget(title)
        
        header.addStretch()
        
        # Search
        header.addWidget(QLabel("🔍"))
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Cari nama/NIP/jabatan...")
        self.txt_search.setMinimumWidth(250)
        self.txt_search.textChanged.connect(self.filter_data)
        header.addWidget(self.txt_search)
        
        layout.addLayout(header)
        
        # Toolbar
        toolbar = QHBoxLayout()
        
        btn_add = QPushButton("+ Tambah Pegawai")
        btn_add.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        btn_add.clicked.connect(self.add_pegawai)
        toolbar.addWidget(btn_add)

        # Export button (green)
        btn_export = QPushButton("📤 Export")
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        btn_export.clicked.connect(self.export_excel)
        toolbar.addWidget(btn_export)

        # Import button (orange)
        btn_import = QPushButton("📥 Import")
        btn_import.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        btn_import.clicked.connect(self.import_data)
        toolbar.addWidget(btn_import)
        
        toolbar.addStretch()
        
        # Filter by role
        toolbar.addWidget(QLabel("Filter:"))
        self.cmb_filter = QComboBox()
        self.cmb_filter.addItem("Semua Pegawai", "")
        self.cmb_filter.addItem("PPK", "ppk")
        self.cmb_filter.addItem("Pejabat Pengadaan", "pejabat_pengadaan")
        self.cmb_filter.addItem("PPSPM", "ppspm")
        self.cmb_filter.addItem("Bendahara", "bendahara")
        self.cmb_filter.addItem("PPHP", "pemeriksa")
        self.cmb_filter.currentIndexChanged.connect(self.filter_data)
        toolbar.addWidget(self.cmb_filter)
        
        # Show inactive
        self.chk_show_inactive = QCheckBox("Tampilkan non-aktif")
        self.chk_show_inactive.stateChanged.connect(self.load_data)
        toolbar.addWidget(self.chk_show_inactive)
        
        layout.addLayout(toolbar)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "NIP", "Nama Lengkap", "Pangkat", "Gol", "Jabatan",
            "Unit Kerja", "Email", "Peran", "Status", "ID"
        ])
        
        # Hide ID column
        self.table.setColumnHidden(9, True)
        
        # Column widths
        self.table.setColumnWidth(0, 150)   # NIP
        self.table.setColumnWidth(1, 200)   # Nama
        self.table.setColumnWidth(2, 100)   # Pangkat
        self.table.setColumnWidth(3, 50)    # Gol
        self.table.setColumnWidth(4, 150)   # Jabatan
        self.table.setColumnWidth(5, 150)   # Unit Kerja
        self.table.setColumnWidth(6, 150)   # Email
        self.table.setColumnWidth(7, 120)   # Peran
        self.table.setColumnWidth(8, 80)    # Status
        
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self.edit_pegawai)
        
        # Context menu
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        layout.addWidget(self.table)
        
        # Summary
        summary = QHBoxLayout()
        self.lbl_total = QLabel("Total: 0 pegawai")
        self.lbl_total.setStyleSheet("font-weight: bold;")
        summary.addWidget(self.lbl_total)
        
        summary.addStretch()
        
        btn_close = QPushButton("Tutup")
        btn_close.clicked.connect(self.accept)
        summary.addWidget(btn_close)
        
        layout.addLayout(summary)
        
        # Style
        self.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QPushButton#btnPrimary {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton#btnPrimary:hover {
                background-color: #2980b9;
            }
        """)
    
    def load_data(self):
        """Load pegawai data into table"""
        active_only = not self.chk_show_inactive.isChecked()
        pegawai_list = self.db.get_all_pegawai(active_only=active_only)
        
        self.all_data = pegawai_list
        self.filter_data()
    
    def filter_data(self):
        """Filter displayed data based on search and role filter"""
        search = self.txt_search.text().lower()
        role_filter = self.cmb_filter.currentData()
        
        filtered = []
        for p in self.all_data:
            # Search filter
            if search:
                searchable = f"{p.get('nip', '')} {p.get('nama', '')} {p.get('jabatan', '')}".lower()
                if search not in searchable:
                    continue
            
            # Role filter
            if role_filter:
                role_col = f"is_{role_filter}"
                if not p.get(role_col):
                    continue
            
            filtered.append(p)
        
        self.display_data(filtered)
    
    def display_data(self, pegawai_list: List[Dict]):
        """Display data in table"""
        self.table.setRowCount(len(pegawai_list))
        
        for row, p in enumerate(pegawai_list):
            # NIP
            self.table.setItem(row, 0, QTableWidgetItem(p.get('nip', '') or '-'))
            
            # Nama lengkap
            nama = p.get('nama', '')
            if p.get('gelar_depan'):
                nama = f"{p['gelar_depan']} {nama}"
            if p.get('gelar_belakang'):
                nama = f"{nama}, {p['gelar_belakang']}"
            self.table.setItem(row, 1, QTableWidgetItem(nama))
            
            # Pangkat
            self.table.setItem(row, 2, QTableWidgetItem(p.get('pangkat', '') or ''))
            
            # Golongan
            self.table.setItem(row, 3, QTableWidgetItem(p.get('golongan', '') or ''))
            
            # Jabatan
            self.table.setItem(row, 4, QTableWidgetItem(p.get('jabatan', '') or ''))
            
            # Unit Kerja
            self.table.setItem(row, 5, QTableWidgetItem(p.get('unit_kerja', '') or ''))
            
            # Email
            self.table.setItem(row, 6, QTableWidgetItem(p.get('email', '') or ''))
            
            # Peran
            peran_list = []
            if p.get('is_ppk'):
                peran_list.append('PPK')
            if p.get('is_pejabat_pengadaan'):
                peran_list.append('PP')
            if p.get('is_ppspm'):
                peran_list.append('PPSPM')
            if p.get('is_bendahara'):
                peran_list.append('BPP')
            if p.get('is_pemeriksa'):
                peran_list.append('PPHP')
            self.table.setItem(row, 7, QTableWidgetItem(', '.join(peran_list)))
            
            # Status
            status_item = QTableWidgetItem("Aktif" if p.get('is_active') else "Non-aktif")
            if not p.get('is_active'):
                status_item.setForeground(QColor('#e74c3c'))
            else:
                status_item.setForeground(QColor('#27ae60'))
            self.table.setItem(row, 8, status_item)
            
            # ID (hidden)
            self.table.setItem(row, 9, QTableWidgetItem(str(p['id'])))
        
        self.lbl_total.setText(f"Total: {len(pegawai_list)} pegawai")
    
    def show_context_menu(self, pos):
        """Show context menu"""
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        
        menu = QMenu(self)
        
        action_edit = QAction("✏️ Edit", self)
        action_edit.triggered.connect(self.edit_pegawai)
        menu.addAction(action_edit)
        
        menu.addSeparator()
        
        # Check if active
        status_item = self.table.item(row, 8)
        is_active = status_item.text() == "Aktif"
        
        if is_active:
            action_deactivate = QAction("🚫 Nonaktifkan", self)
            action_deactivate.triggered.connect(self.deactivate_pegawai)
            menu.addAction(action_deactivate)
        else:
            action_activate = QAction("✅ Aktifkan Kembali", self)
            action_activate.triggered.connect(self.activate_pegawai)
            menu.addAction(action_activate)
        
        menu.exec(self.table.viewport().mapToGlobal(pos))
    
    def get_selected_id(self) -> Optional[int]:
        """Get selected pegawai ID"""
        row = self.table.currentRow()
        if row < 0:
            return None
        id_item = self.table.item(row, 9)
        return int(id_item.text()) if id_item else None
    
    def add_pegawai(self):
        """Add new pegawai"""
        dialog = PegawaiDialog(parent=self)
        if dialog.exec():
            self.load_data()
            self.pegawai_changed.emit()
    
    def edit_pegawai(self):
        """Edit selected pegawai"""
        pegawai_id = self.get_selected_id()
        if not pegawai_id:
            QMessageBox.warning(self, "Peringatan", "Pilih pegawai terlebih dahulu!")
            return
        
        pegawai = self.db.get_pegawai(pegawai_id)
        if not pegawai:
            return
        
        dialog = PegawaiDialog(pegawai, parent=self)
        if dialog.exec():
            self.load_data()
            self.pegawai_changed.emit()
    
    def deactivate_pegawai(self):
        """Deactivate selected pegawai"""
        pegawai_id = self.get_selected_id()
        if not pegawai_id:
            return
        
        reply = QMessageBox.question(
            self, "Konfirmasi",
            "Nonaktifkan pegawai ini?\n\n"
            "Pegawai yang dinonaktifkan tidak akan muncul di daftar pilihan pejabat.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.db.deactivate_pegawai(pegawai_id)
            self.load_data()
            self.pegawai_changed.emit()
    
    def activate_pegawai(self):
        """Reactivate pegawai"""
        pegawai_id = self.get_selected_id()
        if not pegawai_id:
            return
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE pegawai SET is_active = 1, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (pegawai_id,))
            conn.commit()
        
        self.load_data()
        self.pegawai_changed.emit()
    
    def import_csv(self):
        """Import pegawai from CSV file"""
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Pilih File CSV",
            "",
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if not filepath:
            return
        
        try:
            # Read CSV
            pegawai_list = []
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                # Map column names
                for row in reader:
                    data = {}
                    
                    # Try different column name variations
                    data['nip'] = row.get('NIP') or row.get('nip') or row.get('Nip') or ''
                    data['nama'] = row.get('NAMA') or row.get('nama') or row.get('Nama') or ''
                    data['pangkat'] = row.get('PANGKAT') or row.get('pangkat') or row.get('Pangkat') or ''
                    data['golongan'] = row.get('GOLONGAN') or row.get('golongan') or row.get('Gol') or row.get('GOL') or ''
                    data['jabatan'] = row.get('JABATAN') or row.get('jabatan') or row.get('Jabatan') or ''
                    data['unit_kerja'] = row.get('UNIT_KERJA') or row.get('unit_kerja') or row.get('Unit Kerja') or row.get('UNIT KERJA') or ''
                    data['email'] = row.get('EMAIL') or row.get('email') or row.get('Email') or ''
                    data['telepon'] = row.get('TELEPON') or row.get('telepon') or row.get('HP') or row.get('Telepon') or ''
                    
                    if data['nama']:  # Only add if nama exists
                        pegawai_list.append(data)
            
            if not pegawai_list:
                QMessageBox.warning(self, "Peringatan", "Tidak ada data yang dapat diimport!")
                return
            
            # Confirm
            reply = QMessageBox.question(
                self, "Konfirmasi Import",
                f"Ditemukan {len(pegawai_list)} pegawai.\n\n"
                f"Pegawai dengan NIP yang sama akan diupdate.\n"
                f"Lanjutkan import?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Import
            imported, updated, errors = self.db.bulk_import_pegawai(pegawai_list)
            
            # Show result
            msg = f"Import selesai!\n\n"
            msg += f"✅ Ditambahkan: {imported}\n"
            msg += f"🔄 Diupdate: {updated}\n"
            
            if errors:
                msg += f"\n⚠️ {len(errors)} error:\n"
                msg += "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n... dan {len(errors) - 5} lainnya"
            
            QMessageBox.information(self, "Hasil Import", msg)
            
            self.load_data()
            self.pegawai_changed.emit()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal import:\n{str(e)}")
    
    def export_excel(self):
        """Export pegawai to Excel"""
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan Export",
            f"Data_Pegawai_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filepath:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Pegawai"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['NIP', 'Nama', 'Gelar Depan', 'Gelar Belakang', 'Pangkat', 
                      'Golongan', 'Jabatan', 'Unit Kerja', 'Email', 'Telepon',
                      'PPK', 'PP', 'PPSPM', 'Bendahara', 'PPHP', 'Status']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row, p in enumerate(self.all_data, 2):
                ws.cell(row=row, column=1, value=p.get('nip', '')).border = border
                ws.cell(row=row, column=2, value=p.get('nama', '')).border = border
                ws.cell(row=row, column=3, value=p.get('gelar_depan', '')).border = border
                ws.cell(row=row, column=4, value=p.get('gelar_belakang', '')).border = border
                ws.cell(row=row, column=5, value=p.get('pangkat', '')).border = border
                ws.cell(row=row, column=6, value=p.get('golongan', '')).border = border
                ws.cell(row=row, column=7, value=p.get('jabatan', '')).border = border
                ws.cell(row=row, column=8, value=p.get('unit_kerja', '')).border = border
                ws.cell(row=row, column=9, value=p.get('email', '')).border = border
                ws.cell(row=row, column=10, value=p.get('telepon', '')).border = border
                ws.cell(row=row, column=11, value='Ya' if p.get('is_ppk') else '').border = border
                ws.cell(row=row, column=12, value='Ya' if p.get('is_pejabat_pengadaan') else '').border = border
                ws.cell(row=row, column=13, value='Ya' if p.get('is_ppspm') else '').border = border
                ws.cell(row=row, column=14, value='Ya' if p.get('is_bendahara') else '').border = border
                ws.cell(row=row, column=15, value='Ya' if p.get('is_pemeriksa') else '').border = border
                ws.cell(row=row, column=16, value='Aktif' if p.get('is_active') else 'Non-aktif').border = border
            
            # Auto-width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)
            
            wb.save(filepath)
            
            QMessageBox.information(
                self, "Sukses",
                f"Data berhasil diexport ke:\n{filepath}"
            )
            
            # Open file
            import subprocess
            import sys
            if sys.platform == 'win32':
                os.startfile(filepath)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', filepath])
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export:\n{str(e)}")

    def import_data(self):
        """Import pegawai from Excel or CSV file"""
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Import Data Pegawai",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )

        if not filepath:
            return

        reply = QMessageBox.question(
            self, "Konfirmasi Import",
            "Import data pegawai?\n\n"
            "Data dengan NIP yang sama akan diupdate.\n"
            "Data baru akan ditambahkan.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        try:
            pegawai_list = []

            if filepath.endswith('.xlsx'):
                # Import from Excel
                from openpyxl import load_workbook
                wb = load_workbook(filepath)
                ws = wb.active

                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip() if cell.value else '')

                # Map column indices
                col_map = {}
                for idx, h in enumerate(headers):
                    if h in ['nip', 'nomor_nip']:
                        col_map['nip'] = idx
                    elif h in ['nama', 'name', 'nama_pegawai']:
                        col_map['nama'] = idx
                    elif h in ['jabatan', 'position']:
                        col_map['jabatan'] = idx
                    elif h in ['golongan', 'gol']:
                        col_map['golongan'] = idx
                    elif h in ['pangkat', 'rank']:
                        col_map['pangkat'] = idx
                    elif h in ['rekening', 'no_rekening', 'norekening', 'no rekening']:
                        col_map['no_rekening'] = idx
                    elif h in ['bank', 'nama_bank', 'namabank']:
                        col_map['nama_bank'] = idx
                    elif h in ['unitkerja', 'unit_kerja', 'unit kerja', 'unit']:
                        col_map['unit_kerja'] = idx
                    elif h in ['email']:
                        col_map['email'] = idx
                    elif h in ['telepon', 'hp', 'phone', 'telp']:
                        col_map['telepon'] = idx

                # Read data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(v is None for v in row):
                        continue

                    data = {}
                    for field, col_idx in col_map.items():
                        if col_idx < len(row):
                            val = row[col_idx]
                            data[field] = str(val).strip() if val else ''

                    # Only add if nama exists
                    if data.get('nama'):
                        pegawai_list.append(data)

            else:
                # Import from CSV
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        data = {}
                        data['nip'] = row.get('nip') or row.get('NIP') or ''
                        data['nama'] = row.get('nama') or row.get('NAMA') or ''
                        data['jabatan'] = row.get('jabatan') or row.get('JABATAN') or ''
                        data['golongan'] = row.get('golongan') or row.get('GOLONGAN') or row.get('gol') or ''
                        data['pangkat'] = row.get('pangkat') or row.get('PANGKAT') or ''
                        data['no_rekening'] = row.get('rekening') or row.get('no_rekening') or ''
                        data['nama_bank'] = row.get('bank') or row.get('nama_bank') or ''
                        data['unit_kerja'] = row.get('unitKerja') or row.get('unit_kerja') or ''
                        data['email'] = row.get('email') or row.get('EMAIL') or ''
                        data['telepon'] = row.get('telepon') or row.get('TELEPON') or ''

                        if data.get('nama'):
                            pegawai_list.append(data)

            if not pegawai_list:
                QMessageBox.warning(self, "Peringatan", "Tidak ada data yang dapat diimport!")
                return

            # Import data
            imported, updated, errors = self.db.bulk_import_pegawai(pegawai_list)

            msg = f"Import selesai!\n\n"
            msg += f"✅ Ditambahkan: {imported}\n"
            msg += f"🔄 Diupdate: {updated}\n"

            if errors:
                msg += f"\n⚠️ {len(errors)} error:\n"
                msg += "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n... dan {len(errors)-5} error lainnya"

            QMessageBox.information(self, "Import Selesai", msg)
            self.load_data()
            self.pegawai_changed.emit()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal import:\n{str(e)}")


# ============================================================================
# HELPER FUNCTION
# ============================================================================

def show_pegawai_manager(parent=None) -> PegawaiManager:
    """Show pegawai manager dialog"""
    dialog = PegawaiManager(parent)
    dialog.exec()
    return dialog
